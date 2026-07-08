#!/usr/bin/env python3
"""
Cross-store Sold Inventory Performance workbook for Valley Pawn.
Window: 2025-05-17 .. 2026-05-17 (trailing 12 months).
Input: sold_inv_<STORE>_2025-05-17_to_2026-05-17.csv for CUL, HAR, LEX, ROA, WAY.
Output: Valley_Pawn_5Store_SoldInventory_2025-05-17_to_2026-05-17.xlsx
"""
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

INPUT_DIR = "/sessions/upbeat-amazing-cori/mnt/Deep KPI analysis"
OUTPUT = f"{INPUT_DIR}/Valley_Pawn_5Store_SoldInventory_2025-05-17_to_2026-05-17.xlsx"
STORES = ["CUL", "HAR", "LEX", "ROA", "WAY"]
WINDOW = "2025-05-17_to_2026-05-17"

# ---------- Load & clean ----------
def money(s):
    if pd.isna(s):
        return 0.0
    s = str(s).replace("$", "").replace(",", "").strip()
    if s in ("", "-"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

def load_one(store):
    path = f"{INPUT_DIR}/sold_inv_{store}_{WINDOW}.csv"
    if not os.path.exists(path):
        print(f"MISSING: {path}")
        return None
    df = pd.read_csv(path, dtype=str)
    df["Store"] = store
    df["Cost$"] = df["Cost"].apply(money)
    df["AskPrice$"] = df["Price"].apply(money)
    df["SoldPrice$"] = df["Last Sold Price"].apply(money)
    df["Margin$"] = df["SoldPrice$"] - df["Cost$"]
    df["Margin%"] = df.apply(
        lambda r: (r["Margin$"] / r["SoldPrice$"]) if r["SoldPrice$"] > 0 else 0.0,
        axis=1,
    )
    df["SoldDate"] = pd.to_datetime(df["Date"], errors="coerce")
    df["SoldMonth"] = df["SoldDate"].dt.to_period("M").astype(str)
    return df

frames = []
for s in STORES:
    f = load_one(s)
    if f is not None:
        frames.append(f)
if not frames:
    raise SystemExit("No store CSVs found.")
ALL = pd.concat(frames, ignore_index=True)
print(f"Loaded {len(ALL):,} sold items across {ALL['Store'].nunique()} stores")
print(ALL.groupby("Store").size())

# ---------- Workbook ----------
wb = Workbook()

# Style helpers
HDR_FILL = PatternFill("solid", start_color="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOTAL_FILL = PatternFill("solid", start_color="DDEBF7")
TOTAL_FONT = Font(bold=True, name="Arial", size=11)
BORDER = Border(*([Side(style="thin", color="BFBFBF")] * 4))
ARIAL = Font(name="Arial", size=10)
ARIAL_BOLD = Font(name="Arial", size=10, bold=True)

def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = HDR_ALIGN
        c.border = BORDER
        if widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.row_dimensions[row].height = 32

def fmt(ws, rng, code):
    for row in ws[rng]:
        for c in row:
            c.number_format = code
            c.font = ARIAL

# ===== Sheet 1: README =====
ws = wb.active
ws.title = "README"
ws.column_dimensions['A'].width = 110
notes = [
    ("Valley Pawn — 5-Store Sold Inventory Performance", 16, True),
    (f"Window: 5/17/2025 – 5/17/2026 (trailing 12 months)", 11, False),
    (f"Total items sold: {len(ALL):,}    Total revenue: ${ALL['SoldPrice$'].sum():,.0f}    Cost of goods: ${ALL['Cost$'].sum():,.0f}    Gross profit: ${ALL['Margin$'].sum():,.0f}", 11, True),
    ("", 10, False),
    ("Tabs:", 12, True),
    ("  2. Store Scorecard — Revenue, COGS, GP, GP%, item count, avg ticket by store", 10, False),
    ("  3. Category × Store Heatmap — Revenue per category × store", 10, False),
    ("  4. Margin Heatmap — GP% per category × store (color-scaled)", 10, False),
    ("  5. Top 100 Winners — Highest gross-profit single items (all stores)", 10, False),
    ("  6. Negative Margin Sales — Items sold for less than cost (immediate fringe-action list)", 10, False),
    ("  7. Tiny Tickets — Items sold under $10 (per-transaction cost likely exceeds value)", 10, False),
    ("  8. Long-Tail Categories — Categories with very few items / very low revenue (review for shelf-space displacement)", 10, False),
    ("  9. Monthly Trend — Revenue & GP by month × store", 10, False),
    (" 10. CFO Recommendations — Quantified fringe-cut actions per store", 12, True),
    ("", 10, False),
    ("Key terms:", 12, True),
    ("  Cost = inventory cost basis (acquired via loan forfeiture or direct purchase)", 10, False),
    ("  Last Sold Price = actual realized selling price (this is the revenue figure)", 10, False),
    ("  Margin $ = Last Sold Price − Cost    Margin % = Margin $ / Last Sold Price", 10, False),
    ("  Tiny ticket = sold ≤ $10 (excluding $0 inventory like firearms held for transfer)", 10, False),
    ("  Long tail = category with <0.5% of total revenue AND <50 items across all 5 stores", 10, False),
]
for i, (txt, sz, b) in enumerate(notes, 1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = Font(name="Arial", size=sz, bold=b)
    c.alignment = Alignment(wrap_text=True, vertical="top")

# ===== Sheet 2: Store Scorecard =====
ws = wb.create_sheet("Store Scorecard")
header_row(ws, 1,
    ["Store", "Items Sold", "Revenue ($)", "Cost ($)", "Gross Profit ($)", "GP %", "Avg Ticket ($)", "Avg Margin ($)", "Below-Cost Items", "Below-Cost $ Loss"],
    widths=[10, 12, 16, 16, 16, 10, 14, 14, 16, 18])
for i, st in enumerate(STORES, 2):
    sub = ALL[ALL["Store"] == st]
    below = sub[sub["Margin$"] < 0]
    ws.cell(row=i, column=1, value=st).font = ARIAL_BOLD
    ws.cell(row=i, column=2, value=len(sub))
    ws.cell(row=i, column=3, value=sub["SoldPrice$"].sum())
    ws.cell(row=i, column=4, value=sub["Cost$"].sum())
    ws.cell(row=i, column=5, value=f"=C{i}-D{i}")
    ws.cell(row=i, column=6, value=f"=IFERROR(E{i}/C{i},0)")
    ws.cell(row=i, column=7, value=f"=IFERROR(C{i}/B{i},0)")
    ws.cell(row=i, column=8, value=f"=IFERROR(E{i}/B{i},0)")
    ws.cell(row=i, column=9, value=len(below))
    ws.cell(row=i, column=10, value=below["Margin$"].sum())
tot = len(STORES) + 2
ws.cell(row=tot, column=1, value="TOTAL").font = TOTAL_FONT
for col in [2, 3, 4, 9, 10]:
    ws.cell(row=tot, column=col, value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{tot-1})")
ws.cell(row=tot, column=5, value=f"=C{tot}-D{tot}")
ws.cell(row=tot, column=6, value=f"=IFERROR(E{tot}/C{tot},0)")
ws.cell(row=tot, column=7, value=f"=IFERROR(C{tot}/B{tot},0)")
ws.cell(row=tot, column=8, value=f"=IFERROR(E{tot}/B{tot},0)")
for r in range(2, tot + 1):
    for col in [3, 4, 5, 7, 8, 10]:
        ws.cell(row=r, column=col).number_format = '"$"#,##0;("$"#,##0);"-"'
    ws.cell(row=r, column=6).number_format = '0.0%;(0.0%);"-"'
    ws.cell(row=r, column=2).number_format = '#,##0;(#,##0);"-"'
    ws.cell(row=r, column=9).number_format = '#,##0;(#,##0);"-"'
    if r == tot:
        for c in ws[r]:
            c.fill = TOTAL_FILL
            c.font = TOTAL_FONT
            c.border = BORDER

# ===== Sheet 3: Revenue Heatmap =====
ws = wb.create_sheet("Category x Store ($)")
piv = ALL.pivot_table(index="Category", columns="Store", values="SoldPrice$", aggfunc="sum", fill_value=0)
piv["Total"] = piv.sum(axis=1)
piv = piv.sort_values("Total", ascending=False)
header_row(ws, 1, ["Category"] + list(piv.columns), widths=[36] + [13] * len(piv.columns))
for i, (cat, row) in enumerate(piv.iterrows(), 2):
    ws.cell(row=i, column=1, value=cat).font = ARIAL_BOLD
    for j, col in enumerate(piv.columns, 2):
        c = ws.cell(row=i, column=j, value=float(row[col]))
        c.number_format = '"$"#,##0;("$"#,##0);"-"'
        c.font = ARIAL
last_row = len(piv) + 1
ws.cell(row=last_row + 1, column=1, value="TOTAL").font = TOTAL_FONT
for j in range(2, len(piv.columns) + 2):
    col_letter = get_column_letter(j)
    ws.cell(row=last_row + 1, column=j, value=f"=SUM({col_letter}2:{col_letter}{last_row})").number_format = '"$"#,##0;("$"#,##0);"-"'
    ws.cell(row=last_row + 1, column=j).font = TOTAL_FONT
    ws.cell(row=last_row + 1, column=j).fill = TOTAL_FILL
# Color scale on per-store columns
rule = ColorScaleRule(start_type="min", start_color="FFFFFF",
                     mid_type="percentile", mid_value=50, mid_color="9BC2E6",
                     end_type="max", end_color="1F4E78")
for col in range(2, 2 + len(STORES)):
    ws.conditional_formatting.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{last_row}", rule)
ws.freeze_panes = "B2"

# ===== Sheet 4: Margin Heatmap =====
ws = wb.create_sheet("Margin % by Cat x Store")
rev_piv = ALL.pivot_table(index="Category", columns="Store", values="SoldPrice$", aggfunc="sum", fill_value=0)
mar_piv = ALL.pivot_table(index="Category", columns="Store", values="Margin$", aggfunc="sum", fill_value=0)
gp_pct = (mar_piv / rev_piv.replace(0, pd.NA)).fillna(0)
gp_pct["AllStores"] = ALL.groupby("Category").apply(
    lambda g: g["Margin$"].sum() / g["SoldPrice$"].sum() if g["SoldPrice$"].sum() else 0
)
items = ALL.groupby("Category").size()
gp_pct["Items"] = items
gp_pct = gp_pct.sort_values("Items", ascending=False)
header_row(ws, 1, ["Category"] + list(STORES) + ["AllStores", "Items"], widths=[36] + [13] * (len(STORES) + 2))
for i, (cat, row) in enumerate(gp_pct.iterrows(), 2):
    ws.cell(row=i, column=1, value=cat).font = ARIAL_BOLD
    for j, col in enumerate(STORES + ["AllStores"], 2):
        v = float(row[col]) if pd.notna(row[col]) else 0
        c = ws.cell(row=i, column=j, value=v)
        c.number_format = '0.0%;(0.0%);"-"'
        c.font = ARIAL
    ws.cell(row=i, column=len(STORES) + 3, value=int(row["Items"])).number_format = '#,##0'
# Color scale on margin %
rule_m = ColorScaleRule(start_type="num", start_value=-0.5, start_color="F8696B",
                        mid_type="num", mid_value=0, mid_color="FFEB84",
                        end_type="num", end_value=0.7, end_color="63BE7B")
for col in range(2, 2 + len(STORES) + 1):
    ws.conditional_formatting.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{len(gp_pct)+1}", rule_m)
ws.freeze_panes = "B2"

# ===== Sheet 5: Top 100 Winners =====
ws = wb.create_sheet("Top 100 Winners")
winners = ALL[ALL["SoldPrice$"] > 0].nlargest(100, "Margin$")
header_row(ws, 1, ["Store", "Date", "Number", "Category", "Description", "Cost ($)", "Sold ($)", "Margin ($)", "Margin %"],
           widths=[8, 12, 14, 26, 60, 12, 12, 12, 10])
for i, (_, r) in enumerate(winners.iterrows(), 2):
    ws.cell(row=i, column=1, value=r["Store"]).font = ARIAL
    ws.cell(row=i, column=2, value=r["Date"]).font = ARIAL
    ws.cell(row=i, column=3, value=r["Number"]).font = ARIAL
    ws.cell(row=i, column=4, value=r["Category"]).font = ARIAL
    ws.cell(row=i, column=5, value=r["Description"]).font = ARIAL
    ws.cell(row=i, column=6, value=r["Cost$"])
    ws.cell(row=i, column=7, value=r["SoldPrice$"])
    ws.cell(row=i, column=8, value=r["Margin$"])
    ws.cell(row=i, column=9, value=r["Margin%"])
for col in range(6, 9):
    fmt(ws, f"{get_column_letter(col)}2:{get_column_letter(col)}{len(winners)+1}", '"$"#,##0;("$"#,##0);"-"')
fmt(ws, f"I2:I{len(winners)+1}", '0.0%')
ws.freeze_panes = "A2"

# ===== Sheet 6: Negative Margin Sales =====
ws = wb.create_sheet("Below-Cost Sales")
below = ALL[(ALL["Margin$"] < 0) & (ALL["SoldPrice$"] > 0)].sort_values("Margin$")
header_row(ws, 1, ["Store", "Date", "Number", "Category", "Description", "Cost ($)", "Sold ($)", "Loss ($)", "Loss %"],
           widths=[8, 12, 14, 26, 60, 12, 12, 12, 10])
for i, (_, r) in enumerate(below.iterrows(), 2):
    ws.cell(row=i, column=1, value=r["Store"])
    ws.cell(row=i, column=2, value=r["Date"])
    ws.cell(row=i, column=3, value=r["Number"])
    ws.cell(row=i, column=4, value=r["Category"])
    ws.cell(row=i, column=5, value=r["Description"])
    ws.cell(row=i, column=6, value=r["Cost$"])
    ws.cell(row=i, column=7, value=r["SoldPrice$"])
    ws.cell(row=i, column=8, value=r["Margin$"])
    ws.cell(row=i, column=9, value=r["Margin%"])
n = len(below) + 1
for col in [6, 7, 8]:
    fmt(ws, f"{get_column_letter(col)}2:{get_column_letter(col)}{n}", '"$"#,##0.00;("$"#,##0.00);"-"')
fmt(ws, f"I2:I{n}", '0.0%')
ws.freeze_panes = "A2"

# ===== Sheet 7: Tiny Tickets (sub-$10 sales) =====
ws = wb.create_sheet("Tiny Tickets")
tiny = ALL[(ALL["SoldPrice$"] > 0) & (ALL["SoldPrice$"] <= 10)].sort_values("SoldPrice$")
header_row(ws, 1, ["Store", "Date", "Number", "Category", "Description", "Cost ($)", "Sold ($)", "Margin ($)"],
           widths=[8, 12, 14, 26, 60, 12, 12, 12])
for i, (_, r) in enumerate(tiny.iterrows(), 2):
    ws.cell(row=i, column=1, value=r["Store"])
    ws.cell(row=i, column=2, value=r["Date"])
    ws.cell(row=i, column=3, value=r["Number"])
    ws.cell(row=i, column=4, value=r["Category"])
    ws.cell(row=i, column=5, value=r["Description"])
    ws.cell(row=i, column=6, value=r["Cost$"])
    ws.cell(row=i, column=7, value=r["SoldPrice$"])
    ws.cell(row=i, column=8, value=r["Margin$"])
n = len(tiny) + 1
for col in [6, 7, 8]:
    fmt(ws, f"{get_column_letter(col)}2:{get_column_letter(col)}{n}", '"$"#,##0.00;("$"#,##0.00);"-"')
ws.freeze_panes = "A2"

# ===== Sheet 8: Long-Tail Categories =====
ws = wb.create_sheet("Long-Tail Categories")
cat_rollup = ALL.groupby("Category").agg(
    Items=("Number", "count"),
    Revenue=("SoldPrice$", "sum"),
    Cost=("Cost$", "sum"),
    Margin=("Margin$", "sum"),
).reset_index()
cat_rollup["RevShare"] = cat_rollup["Revenue"] / ALL["SoldPrice$"].sum()
cat_rollup["MarginPct"] = cat_rollup.apply(
    lambda r: r["Margin"] / r["Revenue"] if r["Revenue"] else 0, axis=1)
tail = cat_rollup[(cat_rollup["RevShare"] < 0.005) & (cat_rollup["Items"] < 50)].sort_values("Items")
header_row(ws, 1, ["Category", "Items", "Revenue ($)", "Rev Share", "Cost ($)", "Margin ($)", "Margin %"],
           widths=[36, 10, 14, 12, 14, 14, 10])
for i, (_, r) in enumerate(tail.iterrows(), 2):
    ws.cell(row=i, column=1, value=r["Category"]).font = ARIAL
    ws.cell(row=i, column=2, value=int(r["Items"]))
    ws.cell(row=i, column=3, value=r["Revenue"])
    ws.cell(row=i, column=4, value=r["RevShare"])
    ws.cell(row=i, column=5, value=r["Cost"])
    ws.cell(row=i, column=6, value=r["Margin"])
    ws.cell(row=i, column=7, value=r["MarginPct"])
n = len(tail) + 1
for col in [3, 5, 6]:
    fmt(ws, f"{get_column_letter(col)}2:{get_column_letter(col)}{n}", '"$"#,##0;("$"#,##0);"-"')
fmt(ws, f"D2:D{n}", '0.00%')
fmt(ws, f"G2:G{n}", '0.0%')
fmt(ws, f"B2:B{n}", '#,##0')
ws.freeze_panes = "A2"

# ===== Sheet 9: Monthly Trend =====
ws = wb.create_sheet("Monthly Trend")
trend = ALL.groupby(["SoldMonth", "Store"]).agg(
    Revenue=("SoldPrice$", "sum"),
    Items=("Number", "count"),
    Margin=("Margin$", "sum"),
).reset_index()
rev_p = trend.pivot(index="SoldMonth", columns="Store", values="Revenue").fillna(0)
mar_p = trend.pivot(index="SoldMonth", columns="Store", values="Margin").fillna(0)
itm_p = trend.pivot(index="SoldMonth", columns="Store", values="Items").fillna(0)
rev_p["All"] = rev_p.sum(axis=1)
mar_p["All"] = mar_p.sum(axis=1)
itm_p["All"] = itm_p.sum(axis=1)
header_row(ws, 1, ["Month", "Metric"] + list(rev_p.columns), widths=[10, 12] + [12] * len(rev_p.columns))
row = 2
for m in rev_p.index:
    ws.cell(row=row, column=1, value=m).font = ARIAL_BOLD
    ws.cell(row=row, column=2, value="Revenue").font = ARIAL
    for j, c in enumerate(rev_p.columns, 3):
        cell = ws.cell(row=row, column=j, value=float(rev_p.loc[m, c]))
        cell.number_format = '"$"#,##0;("$"#,##0);"-"'
        cell.font = ARIAL
    row += 1
    ws.cell(row=row, column=2, value="GP $").font = ARIAL
    for j, c in enumerate(mar_p.columns, 3):
        cell = ws.cell(row=row, column=j, value=float(mar_p.loc[m, c]))
        cell.number_format = '"$"#,##0;("$"#,##0);"-"'
        cell.font = ARIAL
    row += 1
    ws.cell(row=row, column=2, value="Items").font = ARIAL
    for j, c in enumerate(itm_p.columns, 3):
        cell = ws.cell(row=row, column=j, value=int(itm_p.loc[m, c]))
        cell.number_format = '#,##0'
        cell.font = ARIAL
    row += 1
ws.freeze_panes = "C2"

# ===== Sheet 10: CFO Recommendations =====
ws = wb.create_sheet("CFO Recommendations")
ws.column_dimensions['A'].width = 90
recs = []

# Quantify fringe dollar impact
below_loss = ALL[ALL["Margin$"] < 0]["Margin$"].sum()
below_n = (ALL["Margin$"] < 0).sum()
tiny_rev = ALL[(ALL["SoldPrice$"] > 0) & (ALL["SoldPrice$"] <= 10)]["SoldPrice$"].sum()
tiny_n = ((ALL["SoldPrice$"] > 0) & (ALL["SoldPrice$"] <= 10)).sum()
tail_items = tail["Items"].sum()
tail_rev = tail["Revenue"].sum()
total_rev = ALL["SoldPrice$"].sum()
total_gp = ALL["Margin$"].sum()

recs.append(("CFO Recommendations — Sold Inventory 5/17/2025 – 5/17/2026", 14, True))
recs.append(("", 10, False))
recs.append((f"Total revenue: ${total_rev:,.0f}    Total gross profit: ${total_gp:,.0f}    Blended GP%: {total_gp/total_rev*100:.1f}%" if total_rev else "—", 12, True))
recs.append(("", 10, False))
recs.append(("1. STOP BELOW-COST SALES IMMEDIATELY", 13, True))
recs.append((f"   {below_n:,} items sold for less than cost. Total $ destroyed: ${abs(below_loss):,.0f}.", 10, False))
recs.append(("   Action: Require manager approval for any sale where Last Sold Price < Cost. Build a Bravo override-required rule for sub-cost sales.", 10, False))
recs.append((f"   Quick win: 100% prevention recovers ${abs(below_loss):,.0f}/year (≈ {abs(below_loss)/total_gp*100:.1f}% lift to total GP).", 10, False))
recs.append(("", 10, False))
recs.append(("2. KILL SUB-$10 TICKETS", 13, True))
recs.append((f"   {tiny_n:,} items sold for ≤$10, generating only ${tiny_rev:,.0f} ({tiny_rev/total_rev*100:.2f}% of revenue).", 10, False))
recs.append(("   Action: Bundle these into impulse-bin groupings (3-for-$10, 5-for-$15). Reduces transaction friction and frees employee time.", 10, False))
recs.append((f"   Bonus: Tighten loan acceptance for items projecting to a sub-$10 resale — these clog shelf and consume labor.", 10, False))
recs.append(("", 10, False))
recs.append(("3. PRUNE LONG-TAIL CATEGORIES", 13, True))
recs.append((f"   {len(tail)} categories have <0.5% revenue share AND <50 items sold across all 5 stores combined.", 10, False))
recs.append((f"   Combined: {int(tail_items):,} items, ${tail_rev:,.0f} revenue (only {tail_rev/total_rev*100:.2f}% of total).", 10, False))
recs.append(("   Action: Refuse pawn or buy in these categories unless cost basis < $25. They occupy shelf space disproportionate to revenue contribution.", 10, False))
recs.append(("   See 'Long-Tail Categories' tab for the full list ranked by item count.", 10, False))
recs.append(("", 10, False))
recs.append(("4. STORE-LEVEL DRIFT CHECK", 13, True))
recs.append(("   Compare each store's GP% in 'Store Scorecard' tab — flag any store >5 percentage points below the median.", 10, False))
recs.append(("   Underperforming stores typically have pricing-discipline issues or accept too much low-end intake.", 10, False))
recs.append(("", 10, False))
recs.append(("5. CATEGORY × STORE OPTIMIZATION", 13, True))
recs.append(("   In the Margin Heatmap tab: any cell red (<0%) means that store is losing money in that category.", 10, False))
recs.append(("   Action: Either bring pricing in line with sister stores OR stop accepting that category at that location.", 10, False))
recs.append(("", 10, False))
recs.append(("6. NEXT STEPS — DATA EXPANSION", 13, True))
recs.append(("   To unlock days-on-shelf analysis (the single most actionable cash-flow metric), expand the 'Claude Sold Inv Details' saved report to include Acquisition Date.", 10, False))
recs.append(("   With that, we can quantify exactly how much working capital is tied up in items that take >365 days to sell.", 10, False))
recs.append(("", 10, False))
recs.append(("Refresh: This workbook is regenerated monthly by the autonomous handler. Window rolls forward each month.", 10, True))

for i, (txt, sz, b) in enumerate(recs, 1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = Font(name="Arial", size=sz, bold=b)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = max(18, sz + 4)

wb.save(OUTPUT)
print(f"\nWrote: {OUTPUT}")
print(f"Sheets: {wb.sheetnames}")
