import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()

HDR = PatternFill('solid', fgColor='1F3864')
HDRF = Font(color='FFFFFF', bold=True, size=11)
TITLE = Font(bold=True, size=14, color='1F3864')
SUB = Font(italic=True, size=9, color='595959')
MONEY = '#,##0.00'
THIN = Border(bottom=Side(style='thin', color='BFBFBF'))
TOTFILL = PatternFill('solid', fgColor='DDEBF7')

def sheet(ws, title, subtitle, headers, rows, widths, total_col=None, notes=None):
    ws['A1'] = title; ws['A1'].font = TITLE
    ws['A2'] = subtitle; ws['A2'].font = SUB
    r = 4
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h); c.fill = HDR; c.font = HDRF
        c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 30
    r += 1
    first = r
    for row in rows:
        for i, v in enumerate(row, 1):
            c = ws.cell(r, i, v)
            c.border = THIN
            c.alignment = Alignment(vertical='top', wrap_text=(i == len(row)))
            if isinstance(v, (int, float)) and headers[i-1].lower().startswith('amount'):
                c.number_format = MONEY
        r += 1
    if total_col:
        ci = headers.index(total_col) + 1
        ws.cell(r, ci - 1, 'TOTAL').font = Font(bold=True)
        c = ws.cell(r, ci, f'=SUM({get_column_letter(ci)}{first}:{get_column_letter(ci)}{r-1})')
        c.font = Font(bold=True); c.number_format = MONEY; c.fill = TOTFILL
        ws.cell(r, ci-1).fill = TOTFILL
        r += 1
    if notes:
        r += 1
        for n in notes:
            ws.cell(r, 1, n).font = Font(size=9, italic=True, color='595959')
            r += 1
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(first, 1)

# ---------------- SHEET 1: NOT PROVEN ----------------
ws = wb.active; ws.title = '1. NOT PROVEN'
rows = [
 ['Valley Building Supply','Windows, patio + bi-parting doors, trim, AZEK porch/decking','2020-04 → 2022',28947.87,'Invoice','Open balances $950.75 / $3,181.73 (after $5,200 PlyGem credit) / $724.80, plus $6,232.64 where Balance Due is illegible on the scan. Reduced 8/31/26 by $2,161.00 of newly proven 2022 card payments.','Invoice request emailed to Adair Griffin 8/31/26. Also: OCR the 33-page invoices & credits PDF.'],
 ['Burns Builders Roofing','Roof — Est 1548, installed 7/6/2021','2021-05-28',21750.00,'Signed estimate + completion letter','Service Finance letters confirm the ROOFING PROJECT WAS COMPLETED 7/10/2021 on loan xxx3624. But only routine monthly "Payment Posted" emails exist — no payoff, no "paid in full", no $0 balance. First loan statement shows a payoff of $15,357.47 vs a $21,750 contract: a ~$6,400 gap. Loan is in Hillary Davis\'s name and went delinquent with collector letters into 2022.','Pull the Service Finance xxx3624 origination + payoff statement, and Burns\' final invoice. Highest-dollar single unresolved item.'],
 ['Red Rock Concrete','Engineered retaining walls — Inv #1244 balance','2022-08-07',18622.00,'Vendor invoice','$10,000 of the $28,622 invoice is proven on the invoice face. The remaining balance has no receipt, check, bank record or vendor acknowledgment anywhere after 2022-08-08.','Red Rock invoices through QuickBooks — a payment receipt should exist. Check Aug–Oct 2022. Also check Wells Fargo 2797.'],
 ['Signature Hardware','Bath/plumbing fixtures — all ship-to 282 Bald Rock','2020-04 → 2023-04',9847.50,'Orders + shipment emails','Net $11,730.49 invoiced. $1,139.88 proven on PayPal/AmEx, plus $743.11 newly proven 8/31/26 from the 2022 general ledger. Remainder still unproven.','Pull AmEx ****7115 / ****1005 and MC ****0305 statements — most of this converts.'],
 ['Royal Swimming Pools','Pool equipment/kit — #151398 + #152804','2022-11 → 2022-12',9919.02,'Customer-approved order + shipping','Was $15,511.88. Reduced 8/31/26 — the $5,592.86 Affirm-financed portion is now proven paid in full. What remains is the ~$9,388.07 paid at time of purchase (per the Affirm approval email) plus order #152804 ($530.95). No receipt found for either.','Find the card used for the down payment at checkout, 2022-11-03.'],
 ['Weaver Irrigation','Irrigation install — Inv 2248 / 2377 / 2773','2023',7250.00,'Vendor invoices','$4,400 install + $1,386 + $1,464. Joshua arranged 4 × $1,100 DuPont bill payments 2023-08-21 and the payee was confirmed — but the cleared payments themselves have never been pulled.','DuPont statements Aug–Oct 2023 would close this in one pull.'],
 ['Commonwealth Tile','Tile — Inv 1233 / 1228','2021–2022',1425.00,'Vendor invoices','Other Commonwealth Tile invoices (1294, 1252, 1265) are proven. These two are not.','Same card statements as Signature Hardware.'],
 ['Lowe\'s','Orders #761302963 + #771548245','2022-10 / 2023-03',1344.79,'Order confirmations','Bald Rock / Verona referenced on the orders.','MyLowe\'s Pro purchase history export, filtered to the Verona ship-to.'],
 ['Direct Door Hardware / Builders Warehouse','#203991, #23752, #23939','2020–2021',741.59,'Orders','All shipped to 282 Bald Rock.','Card statements.'],
]
sheet(ws, 'PAYMENTS NOT PROVEN — 282 Bald Rock Road, Verona VA',
 'Work is documented by invoice, signed estimate or order — but no proof of payment has been located. Current as of %s. Source: 282 Bald Rock — Full Evidence Log.md' % date.today().isoformat(),
 ['Vendor','What it covers','Date','Amount','Evidence we DO have','Why it is not proven','How to close it'],
 rows, [30, 38, 16, 14, 26, 62, 52], total_col='Amount',
 notes=['Ranked by dollar value. Everything here is a real documented obligation — the gap is payment evidence, not the work.',
        'The three biggest levers, in order: (1) card statements for MC ****0305, MC ****6246, MC ****1689, LCC ****1037, AmEx ****7115/****1005; (2) DuPont statements Feb 2022 onward; (3) DuPont check images for the 514 unattributed checks.'])

# ---------------- SHEET 2: NEWLY PROVEN ----------------
ws2 = wb.create_sheet('2. NEWLY PROVEN 8-31-26')
rows2 = [
 ['Shreckhise Landscape & Design','Landscaping','2023-01-18',8536.50,'DuPont bill-pay withdrawal "SHRECKHISE LANDS", captured in Lodestar Tax\'s Jan-2023 uncategorized-transaction workbook. Verified against the source spreadsheet, not a text reconstruction.','Log had ZERO dollars for this vendor. Joshua recalled ~$30K — only this one payment exists in any record searched.'],
 ['Pro Quality Property Maintenance ("282 Hardscape")','Front walkway / patio pavers','2024-12-31 → 2025-02-12',7800.00,'Check 5061 $3,000 (12/31/24) + four bill-pay installments of $1,200 (1/22, 1/29, 2/5, 2/12/25), all from Wells Fargo Checking 2797, per Lodestar\'s 2025 uncategorized workbooks.','Job was NOT "Valley Outside Services" — no such vendor exists. Invoice #4347 was $4,800; the $3,000 check was the deposit. Paid in full.'],
 ['Royal Swimming Pools','Pool equipment — Affirm-financed portion','2022-10-27 → 2023-09-28',5592.86,'Affirm loan NM3V-P2HC naming "your Royal Swimming Pools purchase", recurring payment stream, and completion email "Awesome, you\'re all done!" 2023-09-28.','Converts part of the previously fully-unproven $15,511.88.'],
 ['Fundamental Siteworks','Pool demo, dig, rough grade — Inv 670 balance','2023-03-01',2329.00,'QuickBooks payment confirmation email. Combined with the $3,000 already proven (11/22/22), the two payments exactly total the $5,329.00 invoice.','Invoice now fully proven.'],
 ['Valley Building Supply','Building materials — 2022 card charges','2022-02-03 / 2022-04-28',2161.00,'Full Circle Finance 2022 general ledger (Lodestar 1099 compilation workbook): $1,908.44 + $252.56, both coded Repairs & Maintenance.','These 2022 payments were outside the DuPont statement coverage the earlier sweep used.'],
 ['Signature Hardware','Fixtures — 2022 card charges','2022-11-20 / 2022-12-02',743.11,'Same 2022 general ledger: $588.63 + $154.48.','Converts part of the unproven Signature Hardware block.'],
]
sheet(ws2, 'NEWLY PROVEN — found 8/31/2026',
 'Payments located this pass that were previously unproven or entirely undocumented. Source: Lodestar Tax bookkeeping workbooks (32 spreadsheets pulled from 564 Lodestar emails) + Affirm/QuickBooks confirmations.',
 ['Vendor','What it covers','Date(s)','Amount','Proof','Note'],
 rows2, [34, 34, 24, 14, 70, 56], total_col='Amount',
 notes=['This moves proven-paid from $114,073.76 to $124,777.87.'])

# ---------------- SHEET 3: PAID, ATTRIBUTION TBD ----------------
ws3 = wb.create_sheet('3. PAID - WHICH PROPERTY')
rows3 = [
 ['Augusta Aluminum Gutterworks','Gutters','2022-04-08',4500.00,'Check 1206, Full Circle business checking, coded Repairs & Maintenance','HIGH','Master log flagged an unresolved gutter question (Burns Est 1557 $7,693.85 vs a Retex job). This looks like the answer: gutters were done by Augusta Aluminum for $4,500. Confirm the service address.'],
 ['Augusta Steel Corporation','Steel','2022-03-09',2652.50,'Debit card, Verona VA, Repairs & Maintenance','MEDIUM','Verona is Bald Rock\'s town, and the timing matches the retaining-wall / structural work. Could also be a store fixture job.'],
 ['Fiber Pro Insulation Inc','Insulation','2022-04-21',2500.00,'Check 1210, Repairs & Maintenance','MEDIUM','No insulation vendor appears anywhere in the Bald Rock file. Timing sits inside the renovation window.'],
 ['Augusta County','Building permit','2022-05-23',76.50,'Check 1219, memo reads "282", coded Licenses & Permits','VERY HIGH','The evidence log states no Augusta County permit exists anywhere. This check memo is literally "282". Small dollars, disproportionate audit value — a permit corroborates the scope of work.'],
 ['Zelle → "AVABI"','Gutters','2025-06-05',1745.00,'Wells Fargo Checking 2797, memo "GUTTERS"','MEDIUM','2025 gutter work. Payee is a first name only.'],
 ['Zelle → "Marlon"','Tile','2025-07-16 / 2025-07-18',3000.00,'Wells Fargo Checking 2797, memos "TILE BORO" and "TILE 1790"','LOW','"BORO" may mean the Waynesboro store, not Bald Rock. Needs Joshua to confirm.'],
 ['Zelle → "GC"','Pocket door frames','2025-09-03',832.03,'Wells Fargo Checking 2797, memo "POCKET DOOR FRAMES"','MEDIUM','Interior work, 2025.'],
 ['Bill Pay — "Flooring ON-LINE"','Flooring','2025-09-16',1000.00,'Wells Fargo Checking 2797','MEDIUM','Recurring payee — likely more instances beyond the two months sampled.'],
 ['Bill Pay — "Electric 282"','Electrical','2025-09-03 / 2025-09-17',594.00,'Wells Fargo Checking 2797, recurring payee named "Electric 282"','VERY HIGH','Payee is literally named for the property. Recurring — the two payments below are only what showed in the sampled months.'],
 ['Bill Pay — "282 Plumber"','Plumbing','2025-09-03 / 2025-09-17',376.00,'Wells Fargo Checking 2797, recurring payee named "282 Plumber"','VERY HIGH','Same — named for the property, and recurring. The evidence log currently says "Plumbing — nothing at all."'],
]
sheet(ws3, 'PAID — BUT WHICH PROPERTY?',
 'Real, cleared payments found in the Full Circle general ledger and the Wells Fargo 2797 uncategorized reports. The money moved; what needs confirming is that the work was at 282 Bald Rock and not a Valley Pawn store or 817 Richmond Rd.',
 ['Vendor / Payee','Work','Date(s)','Amount','How it was paid','Confidence it is 282','Note'],
 rows3, [32, 22, 24, 14, 52, 20, 70], total_col='Amount',
 notes=['Do NOT add these to basis until the service address is confirmed. Full Circle pays for all five pawn stores, 817 Richmond Rd and the rental out of the same accounts.',
        'The two recurring bill-pay payees named "Electric 282" and "282 Plumber" are the strongest — Joshua named them himself when setting up bill pay.'])

# ---------------- SHEET 4: THE 2025 GAP ----------------
ws4 = wb.create_sheet('4. THE BIG GAP')
rows4 = [
 ['Wells Fargo Checking 2797','2024 → present','NOT IN THE FILE AT ALL','This account pays Bald Rock vendors — Pro Quality, the hardscape installments, "Electric 282", "282 Plumber", the 2025 Zelles. The evidence log only ever covered DuPont. Everything found from this account so far came from Lodestar\'s uncategorized-transaction reports, which by definition only capture what the bookkeeper could not categorize — a fraction of the real activity.','Download WF 2797 statements 2024-01 → present. This is now the single highest-value pull on the property.'],
 ['DuPont Community CU 766518','Feb 2022 → present','MISSING','Statements on file stop at Jan 2022. Red Rock, Royal Swimming Pools, Fundamental Siteworks, Commonwealth Tile, Weaver Irrigation and R.E. Boggs were all paid after that date. Their absence from the sweep is a coverage gap, not evidence of non-payment.','Download DuPont statements Feb 2022 → Dec 2025.'],
 ['DuPont check images','2020 → 2021','514 checks, $535,113.23, no payee on any','DuPont statements print only the check number. Lodestar\'s 2022 general ledger DOES name check payees — which is how Augusta Aluminum Gutterworks, Fiber Pro Insulation and the Augusta County permit surfaced today. The same trick will not work for 2020–2021 because no equivalent ledger export has been found for those years.','Pull front/back check images from DuPont online banking. Most will be store operating checks; the renovation subset has to be identified visually.'],
 ['Card statements','2020 → 2023','MISSING','MC ****0305, MC ****6246, MC ****1689, LCC ****1037, AmEx ****7115 / ****1005. These cards already paid Bald Rock vendors on documents we hold.','Would convert most of the Signature Hardware, Commonwealth Tile and Royal Swimming Pools remainder.'],
 ['Lodestar 2020 / 2021 ledger exports','2020 → 2021','NOT FOUND','32 Lodestar workbooks were recovered and parsed. The oldest transaction-level export is 2022. A 2021 P&L exists but has no transaction detail — which is why Burns Builders (2021 roof) could not be resolved this way.','Ask Lodestar Tax directly for a 2020–2021 general ledger / transaction detail export. One email.'],
]
sheet(ws4, 'THE GAP — where the missing payments actually are',
 'Every unproven item above traces back to one of these five gaps. None of them require more searching — they require a download or one email.',
 ['Source','Period','Status','Why it matters','Action'],
 rows4, [30, 20, 26, 88, 62],
 notes=['Ranked. Wells Fargo 2797 is first because it is an entire account the evidence file has never seen, and it is demonstrably paying Bald Rock trades right now.'])

# ---------------- SHEET 5: PROVEN (reference) ----------------
ws5 = wb.create_sheet('5. PROVEN (reference)')
rows5 = [
 ['R.E. Boggs, Inc.','Two Rheem HVAC systems','2025-09-03',29187.00,'Service Finance loan 5977065 signed 2025-09-05; Certificate of Completion 2025-09-08'],
 ['Valley Building Supply','Windows, doors, trim, AZEK','2020-04 → 2022',29722.06,'Balance Due $0.00 / CREDIT CRD on invoice faces ($27,561.06) + $2,161.00 from the 2022 ledger'],
 ['Renu Therapy','Cold plunge — Cold Stoic 2.0','2023-05-27',10214.09,'Paid order + shipment 7/16/23'],
 ['Red Rock Concrete','Engineered retaining walls (partial)','2022-08-07',10000.00,'Payment line on the vendor\'s own final invoice'],
 ['Lowe\'s','Appliances — Order #717201671','2021-03-17',8757.88,'Payment LCC ending 1037'],
 ['Shreckhise Landscape & Design','Landscaping','2023-01-18',8536.50,'DuPont bill-pay withdrawal, per Lodestar Jan-2023 workbook'],
 ['Pro Quality Property Maintenance','Front walkway / patio','2024-12 → 2025-02',7800.00,'Check 5061 + 4 × $1,200 bill pay, WF Checking 2797'],
 ['Royal Swimming Pools','Pool equipment (financed portion)','2022-10 → 2023-09',5592.86,'Affirm loan NM3V-P2HC paid in full'],
 ['ProjectorScreen.com','SI 5 Series 160" screen','2020-02-14',5470.50,'Payment received, credit card'],
 ['Fundamental Siteworks','Pool demo, dig, grade — Inv 670','2022-11 / 2023-03',5329.00,'$3,000 MC ****6246 + $2,329 QuickBooks confirmation'],
 ['Commonwealth Tile','Tile — Inv 1294','2022-11-20',1685.00,'MC ****0305, Auth MQ0134421222'],
 ['Signature Hardware','Bath/plumbing fixtures','2020-04 → 2023-04',1883.00,'PayPal + AmEx ($1,139.88) + 2022 ledger ($743.11)'],
 ['Commonwealth Tile','Tile labor + material — Inv 1252 / 1265','2021',400.00,'QuickBooks payment confirmations'],
 ['Royal Swimming Pools','Pool niche — #151856','2022-12-01',199.99,'PayPal receipt'],
]
sheet(ws5, 'PROVEN PAID — for reference',
 'Payment proof exists on the face of a document. This is the number that feeds depreciable basis.',
 ['Vendor','What it covers','Date','Amount','Proof'],
 rows5, [34, 40, 22, 14, 74], total_col='Amount',
 notes=['Purchase price $405,000 + proven improvements = adjusted basis before the land allocation and the lesser-of-basis-or-FMV test.'])

out = '/root/x.xlsx'
import os
out = os.path.expanduser('~/mnt/Taxes 2026/282_Bald_Rock_Payments_Proven_and_Unproven.xlsx')
wb.save(out)
print('saved', out)
