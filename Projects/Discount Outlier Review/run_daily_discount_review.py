#!/usr/bin/env python3
"""
Valley Pawn — Daily Discount Outlier Review ("Discount Review")

Grades POINT-OF-SALE DISCOUNT behavior on sold goods — the amount knocked off
the ticketed/asking price at the register — as distinct from Sold Margin
Review's realized-margin math (cost vs sale price). Both signals matter and
can diverge: a high-margin item can still carry a heavy discount (leakage /
favoritism the margin math never surfaces); a thin-margin item can carry zero
discount (a pricing problem, not a discounting-behavior problem).

DATA SOURCE (confirmed 2026-07-28/29 by reading real CSVs, not assuming):
Bravo's existing "Claude Sold Inv Details" saved report (Inventory module) —
already pulled daily by the additive `jewelry-margin-sold` pipeline cell built
for an unrelated jewelry-scrap project — carries, per sold item, across ALL
categories (no jewelry filter in the AHK handler itself):
    Number,Status,Category,Description,Cost,Price,Last Sold Price,Date
`Price` = the ticketed/asking price. `Last Sold Price` = what it actually
sold for. Price - Last Sold Price IS the discount, no new Bravo report or AHK
handler needed. This script is additive/read-only: it never touches the
`jewelry-margin-sold` cell, its AHK handler, or the Sold Margin Review project.

Usage:
    python run_daily_discount_review.py [YYYY-MM-DD]
    Date defaults to yesterday. Pass a specific date for re-runs.

Output (daily/ subfolder next to this script):
    daily/{DATE}_discount_review.xlsx
    daily/{DATE}_discount_review_summary.json

Exit codes:
    0 — run complete (even if Slack token missing — check JSON)
    1 — critical error reading input
"""

from __future__ import annotations
import csv, glob, re, json, os, sys, datetime, statistics
import urllib.request

# ── Paths ─────────────────────────────────────────────────────────────────
HERE         = os.path.dirname(os.path.abspath(__file__))
BRAVO_OUTPUT = "/Users/joshuadavis/Documents/Claude/Projects/Bravo Data Extraction/output/"
DAILY_DIR    = os.path.join(HERE, "daily")

# ── Thresholds & config (Expert Board decision, 2026-07-28) ────────────────
# Dual threshold: a %-only rule misses a big $ knock on an expensive item at
# a modest %; a $-only rule misses a heavy % discount on cheap goods that is
# still a real behavioral signal. 20% mirrors "meaningful negotiation room
# becomes leakage" at the company's ~52% average retail margin; $50 is the
# rough daily materiality line for a 5-store chain this size. Both tunable
# after 2-4 weeks of live data.
SLACK_CHANNEL     = os.environ.get("DISCOUNT_REVIEW_SLACK_CHANNEL", "C0BQ6JA27MX")  # #discount-review,
                    # the private team channel Joshua created 2026-08-13. This report is now TEAM-VISIBLE
                    # (previously it defaulted to Joshua's own DM, D03BHQH5VGT, because no such channel
                    # existed yet). Keep the daily post plain and free of system/tool names per the Field
                    # Communication Standard — the team reads this, not just Joshua. FAILURE notices still
                    # go to Joshua's DM only, never here. Override with DISCOUNT_REVIEW_SLACK_CHANNEL.
FLAG_PCT          = 0.20     # flag items discounted >= 20% off ticket price
FLAG_DOLLARS      = 50.00    # OR discounted >= $50 off ticket price
GENERIC_SKU_RE     = re.compile(r'^\d+$')  # bare numeric "Number" = reused generic/bulk SKU (coins,
                                            # misc tools, bullion) — Price is not a real per-item
                                            # asking price for these; excluded from % ranking, kept
                                            # in a data-quality footnote only.
PLACEHOLDER_PRICE_MAX = 0.01  # firearms awaiting FFL paperwork/pricing show Price=$0.01, Cost=$0.00,
                               # Last Sold Price=$0.00 in Bravo — a data-entry placeholder, NOT a
                               # real $0.01 asking price that got "100% discounted." Confirmed via a
                               # live sample pull (2026-07-28/29): 5 of 40 CUL rows showed this exact
                               # pattern, all firearms. Treated like a generic SKU: excluded from
                               # ranking/flagging, counted in a data-quality footnote only.

STORES = ["CUL", "HAR", "LEX", "ROA", "WAY"]

# Candidate input filename patterns for a given date/store, tried in order.
# {d} = ISO date (single-day range, start==end, matching the cell's contract)
_FILENAME_CANDIDATES = [
    # Preferred source as of 2026-08-13: the sold-discount-detail cell, an
    # additive clone of jewelry-margin-sold that (a) writes a header-only CSV
    # on a genuine zero-sale day so "ran, no sales" is provable on disk, and
    # (b) validates grid identity so a stranded store-picker grid can never be
    # captured as sold-item data. Both bugs bit the jewelry-margin-sold path
    # live on 2026-08-13. The older patterns stay below as fallbacks so any
    # previously-pulled data still parses.
    "{d}_to_{d}_{store}_sold-discount-detail.csv",
    "{d}_{store}_sold-discount-detail.csv",
    "{d}_to_{d}_{store}_jewelry-margin-sold.csv",
    "{d}_{store}_jewelry-margin-sold.csv",
    "{d}_to_{d}_{store}_sold-discounts.csv",
    "{d}_{store}_sold-discounts.csv",
]


# ── Date resolution ─────────────────────────────────────────────────────────
def resolve_date(arg: str | None) -> datetime.date:
    if arg:
        try:
            return datetime.date.fromisoformat(arg)
        except ValueError:
            print(f"Bad date '{arg}' — expected YYYY-MM-DD", file=sys.stderr)
            sys.exit(1)
    return datetime.date.today() - datetime.timedelta(days=1)


# ── Slack token resolution (same pattern as sold-review / pawn-walk) ───────
def _get_slack_token() -> str | None:
    tok = os.environ.get("SLACK_BOT_TOKEN", "").strip()
    if tok.startswith("xoxb-"):
        return tok
    for cpath in [
        os.path.join(BRAVO_OUTPUT, "..", "slack_config.json"),
        os.path.expanduser("~/Documents/Claude/Projects/Bravo Data Extraction/slack_config.json"),
        os.path.expanduser("~/.vp_slack_config.json"),
    ]:
        try:
            with open(os.path.normpath(cpath)) as f:
                d = json.load(f)
            t = (d.get("SLACK_BOT_TOKEN") or d.get("slack_bot_token") or "").strip()
            if t.startswith("xoxb-"):
                return t
        except Exception:
            pass
    for profile in [
        os.path.expanduser("~/.bash_profile"),
        os.path.expanduser("~/.zshenv"),
        os.path.expanduser("~/.profile"),
    ]:
        try:
            txt = open(profile).read()
            m = re.search(r'(?:export\s+)?SLACK_BOT_TOKEN=["\']?(xoxb-[^\s"\']+)', txt)
            if m:
                return m.group(1).rstrip("\"'")
        except Exception:
            pass
    return None


# ── Money parsing ────────────────────────────────────────────────────────────
def money(s):
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace("$", "").replace(",", "").strip()
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None


def _find_col(fieldnames, *candidates):
    norm = {re.sub(r'[^a-z0-9]', '', fn.lower()): fn for fn in (fieldnames or [])}
    for c in candidates:
        key = re.sub(r'[^a-z0-9]', '', c.lower())
        if key in norm:
            return norm[key]
    return None


# ── Load sold items for a specific date, across all 5 stores ───────────────
def load_sold_for_date(date: datetime.date) -> tuple[list[dict], list[str]]:
    ds = date.isoformat()
    rows: list[dict] = []
    missing_files: list[str] = []

    for store in STORES:
        path = None
        for pat in _FILENAME_CANDIDATES:
            candidate = os.path.join(BRAVO_OUTPUT, pat.format(d=ds, store=store))
            if os.path.exists(candidate):
                path = candidate
                break
        if path is None:
            missing_files.append(store)
            continue

        try:
            with open(path, newline='', encoding='utf-8-sig') as fh:
                reader = csv.DictReader(fh)
                fns = reader.fieldnames or []
                col_num    = _find_col(fns, "Number", "Item Number", "Inventory #")
                col_status = _find_col(fns, "Status")
                col_cat    = _find_col(fns, "Category")
                col_desc   = _find_col(fns, "Description")
                col_cost   = _find_col(fns, "Cost")
                col_price  = _find_col(fns, "Price", "Ticket Price", "List Price", "Original Price")
                col_last   = _find_col(fns, "Last Sold Price", "Sale Price", "Sold Price")
                col_date   = _find_col(fns, "Date", "Date Sold", "Sale Date")

                for row in reader:
                    status = (row.get(col_status) or "").strip().upper() if col_status else "SOLD"
                    if status and status != "SOLD":
                        continue
                    number = (row.get(col_num) or "").strip() if col_num else ""
                    price  = money(row.get(col_price)) if col_price else None
                    last   = money(row.get(col_last)) if col_last else None
                    cost   = money(row.get(col_cost)) if col_cost else None
                    if price is None or last is None or price <= 0:
                        continue  # can't compute discount without both fields
                    row_date = (row.get(col_date) or "").strip() if col_date else ds
                    rows.append({
                        "store":        store,
                        "number":       number,
                        "category":     (row.get(col_cat) or "").strip() if col_cat else "",
                        "desc":         (row.get(col_desc) or "").strip() if col_desc else "",
                        "cost":         cost,
                        "price":        price,
                        "last_price":   last,
                        "date":         row_date,
                        "generic_sku":  bool(GENERIC_SKU_RE.match(number)) if number else False,
                        "placeholder_price": bool(price is not None and price <= PLACEHOLDER_PRICE_MAX),
                        "source_file":  os.path.basename(path),
                    })
        except Exception as e:
            print(f"WARNING: could not read {path}: {e}", file=sys.stderr)
            missing_files.append(store)

    return rows, missing_files


# ── Discount computation ────────────────────────────────────────────────────
def compute_discount(r: dict) -> dict:
    price, last, cost = r["price"], r["last_price"], r.get("cost")
    raw_discount = price - last
    sold_above_list = raw_discount < 0
    discount_dollars = max(0.0, raw_discount)
    discount_pct = (discount_dollars / price) if price else None
    into_loss = bool(cost is not None and last <= cost and discount_dollars > 0)
    excluded = r["generic_sku"] or r["placeholder_price"]
    flag = (not excluded) and (
        (discount_pct is not None and discount_pct >= FLAG_PCT) or discount_dollars >= FLAG_DOLLARS
    )
    out = dict(r, discount_dollars=discount_dollars, discount_pct=discount_pct,
               sold_above_list=sold_above_list, into_loss=into_loss, flag=flag or into_loss)
    return out


# ── Build Slack message ─────────────────────────────────────────────────────
def compute_ytd(date: datetime.date,
                today_store_totals: dict[str, float],
                today_company_total: float):
    """Running CALENDAR-YEAR total of discount dollars, by store and company.

    Added 2026-08-13 at Joshua's request so the team sees a cumulative annual
    number every day, not just the single-day figure.

    Reads the per-day summary JSONs already written in daily/ for the same year
    and sums their per-store `total_discount_dollars`. The TARGET DATE is
    deliberately EXCLUDED from the file scan and supplied from this run's
    in-memory numbers instead — so re-running a day recomputes rather than
    double-counts, and the YTD is always self-healing from the files on disk
    (no separate running-total ledger to drift out of sync).

    Returns (store_ytd: dict, company_ytd: float, days_counted: int, year: int).
    """
    year = date.year
    ds = date.isoformat()
    store_ytd: dict[str, float] = {}
    company_ytd = 0.0
    days = 0

    if os.path.isdir(DAILY_DIR):
        for fn in sorted(os.listdir(DAILY_DIR)):
            if not fn.endswith("_discount_review_summary.json"):
                continue
            fdate = fn.split("_")[0]
            if not fdate.startswith(f"{year}-") or fdate == ds:
                continue
            try:
                with open(os.path.join(DAILY_DIR, fn), "r", encoding="utf-8") as f:
                    d = json.load(f)
            except Exception:
                continue  # a malformed/partial day must never break today's post
            for st, sv in (d.get("stores") or {}).items():
                try:
                    store_ytd[st] = store_ytd.get(st, 0.0) + float(sv.get("total_discount_dollars") or 0.0)
                except (TypeError, ValueError):
                    pass
            try:
                company_ytd += float(d.get("total_discount_dollars") or 0.0)
            except (TypeError, ValueError):
                pass
            days += 1

    for st, v in (today_store_totals or {}).items():
        store_ytd[st] = store_ytd.get(st, 0.0) + float(v or 0.0)
    company_ytd += float(today_company_total or 0.0)
    days += 1

    return store_ytd, company_ytd, days, year


def build_slack_message(valued: list[dict], date: datetime.date, missing_stores: list[str],
                        store_ytd: dict[str, float] | None = None,
                        company_ytd: float | None = None,
                        ytd_days: int | None = None,
                        ytd_year: int | None = None) -> str | None:
    real = [r for r in valued if not (r["generic_sku"] or r["placeholder_price"])]
    if len(real) < 3:
        return None
    ds = date.isoformat()
    BAR = "━" * 22

    def _pct(m):
        return f"{m * 100:.0f}%" if m is not None else "—"

    lines = [f"🏷️ *Discount Review — {ds}*", "",
             f"> Ticket price vs actual sale price on yesterday's sold items. Flag: ≥{int(FLAG_PCT*100)}% off OR ≥${FLAG_DOLLARS:.0f} off.",
             ""]

    # Show EVERY store that has traded at all this year, not just the ones with sales
    # today (Joshua 2026-08-13 — the team should see the full per-store board daily).
    # On a closed day (Wed = CUL only) or a store's quiet day, it still shows with
    # Today $0 and its running YTD, so nobody's cumulative number silently vanishes.
    stores = sorted(set(r["store"] for r in real) | set(store_ytd or {}))

    # Rank the stores by TODAY's weighted-avg discount % (Joshua 2026-08-23):
    # lowest discount % = 1st place (best discipline), highest = last place.
    # Only stores with actual sales today can be ranked — a store with no sales
    # today has no discount % to rank on, so it's shown below the ranked list,
    # unranked, exactly as before.
    RANK_BADGE = {1: "🥇 1st", 2: "🥈 2nd", 3: "🥉 3rd"}

    def _ordinal(n):
        return RANK_BADGE.get(n, f"{n}th")

    ranked_stores = []
    no_sale_stores = []
    for st in stores:
        si = [r for r in real if r["store"] == st]
        disc_sum = sum(r["discount_dollars"] for r in si)
        price_sum = sum(r["price"] for r in si)
        fl = sum(1 for r in si if r["flag"])
        wavg = (disc_sum / price_sum) if price_sum else None
        if not si:
            no_sale_stores.append(st)
            continue
        ranked_stores.append({"store": st, "n": len(si), "disc_sum": disc_sum,
                              "price_sum": price_sum, "fl": fl, "wavg": wavg})
    ranked_stores.sort(key=lambda s: s["wavg"] if s["wavg"] is not None else 0.0)

    total_items = 0
    total_disc_dollars = 0.0
    total_price_dollars = 0.0
    total_flags = 0
    for rank, s in enumerate(ranked_stores, 1):
        st, disc_sum, price_sum, fl, wavg = s["store"], s["disc_sum"], s["price_sum"], s["fl"], s["wavg"]
        total_items += s["n"]
        total_disc_dollars += disc_sum
        total_price_dollars += price_sum
        total_flags += fl
        # Day AND year-to-date on the same line, per Joshua 2026-08-13 — the team should
        # see today's number next to the cumulative annual one, not in a separate block.
        ytd_part = f" | *YTD ${store_ytd[st]:,.0f}*" if (store_ytd and st in store_ytd) else ""
        stt = "✅" if (wavg is not None and wavg < FLAG_PCT) else "🚨"
        fw = "flag" if fl == 1 else "flags"
        lines.append(f"• {_ordinal(rank)} · *{st}* — {s['n']} items | Avg discount {_pct(wavg)} {stt} | "
                     f"Today ${disc_sum:,.0f}{ytd_part} | {fl} {fw}")
    for st in no_sale_stores:
        # Traded earlier this year but nothing sold today (or closed today) — unranked.
        ytd_part = f" | *YTD ${store_ytd[st]:,.0f}*" if (store_ytd and st in store_ytd) else ""
        lines.append(f"• *{st}* — no sales today{ytd_part}")
    if missing_stores:
        lines.append(f"_No data file for: {', '.join(missing_stores)}_")

    cwavg = (total_disc_dollars / total_price_dollars) if total_price_dollars else None
    lines.append("")
    co_ytd_part = f" | *YTD ${company_ytd:,.0f}*" if company_ytd is not None else ""
    lines.append(f"*COMPANY* — {total_items} items | Avg discount {_pct(cwavg)} | "
                 f"Today ${total_disc_dollars:,.0f}{co_ytd_part} | {total_flags} flags")
    if company_ytd is not None and ytd_days:
        day_word = "selling day" if ytd_days == 1 else "selling days"
        lines.append(f"_YTD = total discounted off ticket in {ytd_year}, across {ytd_days} {day_word}._")

    ranked_pct = sorted([r for r in real if r["discount_pct"] is not None],
                        key=lambda r: -r["discount_pct"])[:10]
    if ranked_pct:
        lines.append("")
        lines.append(BAR)
        lines.append("*Top 10 by discount % — company-wide*")
        lines.append(BAR)
        for r in ranked_pct:
            desc = (r.get("desc") or r.get("category") or "item").strip()
            if len(desc) > 40:
                desc = desc[:39] + "…"
            crit = " ⛔ into a loss" if r["into_loss"] else ""
            lines.append(f"• {r['store']} · {_pct(r['discount_pct'])} off (${r['discount_dollars']:,.0f}) · ${r['price']:,.0f}→${r['last_price']:,.0f} · {desc}{crit}")

    flagged = sorted([r for r in real if r["flag"]], key=lambda r: -(r["discount_pct"] or 0))
    if flagged:
        lines.append("")
        lines.append(BAR)
        lines.append(f"*🚨 HEAVY DISCOUNTS ({len(flagged)})*")
        lines.append(BAR)
        for r in flagged[:12]:
            desc = (r.get("desc") or r.get("category") or "item").strip()
            if len(desc) > 40:
                desc = desc[:39] + "…"
            crit = " ⛔CRITICAL(into a loss)" if r["into_loss"] else ""
            lines.append(f"• {r['store']} · {_pct(r['discount_pct'])} off · ${r['discount_dollars']:,.0f} · {desc}{crit}")
        if len(flagged) > 12:
            lines.append(f"…and {len(flagged) - 12} more — full detail in the spreadsheet")

    return "\n".join(lines)


# ── Post to Slack ────────────────────────────────────────────────────────────
def slack_post(text: str, token: str, channel: str) -> bool:
    payload = json.dumps({"channel": channel, "text": text, "mrkdwn": True}).encode("utf-8")
    req = urllib.request.Request(
        "https://slack.com/api/chat.postMessage",
        data=payload,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json; charset=utf-8"},
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            body = json.load(resp)
            if not body.get("ok"):
                print(f"Slack API error: {body.get('error', '?')}", file=sys.stderr)
            return bool(body.get("ok"))
    except Exception as e:
        print(f"Slack request failed: {e}", file=sys.stderr)
        return False


# ── Excel report ─────────────────────────────────────────────────────────────
def write_excel(valued: list[dict], date: datetime.date, path: str) -> bool:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxl not installed — Excel skipped.", file=sys.stderr)
        return False

    FILL_HDR  = PatternFill("solid", fgColor="1F3864")
    FILL_CRIT = PatternFill("solid", fgColor="FF9999")
    FILL_FLAG = PatternFill("solid", fgColor="FFCCCC")
    FILL_WARN = PatternFill("solid", fgColor="FFF0CC")
    FILL_GOOD = PatternFill("solid", fgColor="CCFFCC")
    FILL_GEN  = PatternFill("solid", fgColor="E7E6E6")
    FONT_HDR  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    FONT_BODY = Font(name="Calibri", size=10)
    FONT_BOLD = Font(name="Calibri", bold=True, size=10)
    THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = FILL_HDR; c.font = FONT_HDR; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _cell(ws, row, col, val, fmt=None, fill=None, bold=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = FONT_BOLD if bold else FONT_BODY
        c.border = THIN
        c.alignment = Alignment(vertical='center')
        if fmt: c.number_format = fmt
        if fill: c.fill = fill

    wb = openpyxl.Workbook()
    ds = date.isoformat()

    ws1 = wb.active
    ws1.title = "Items"
    ws1.freeze_panes = "A3"
    title = f"Valley Pawn — Discount Review  |  {ds}  |  Flag: >={int(FLAG_PCT*100)}% or >=${FLAG_DOLLARS:.0f} off"
    ws1.merge_cells("A1:L1")
    t = ws1["A1"]; t.value = title
    t.fill = PatternFill("solid", fgColor="0D1B40")
    t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 22

    hdrs1 = ["Store", "Item #", "Category", "Description", "Ticket Price", "Sale Price",
             "Discount $", "Discount %", "Sold Above List?", "Into a Loss?", "Generic SKU?", "Flag?"]
    for ci, h in enumerate(hdrs1, 1):
        _hdr(ws1, 2, ci, h)
    ws1.row_dimensions[2].height = 28

    for ri, r in enumerate(sorted(valued, key=lambda x: (x["store"], -(x["discount_pct"] or 0))), 3):
        if r["generic_sku"] or r["placeholder_price"]:
            rfill = FILL_GEN
        elif r["into_loss"]:
            rfill = FILL_CRIT
        elif r["flag"]:
            rfill = FILL_FLAG
        elif r["discount_pct"] is not None and r["discount_pct"] > 0:
            rfill = FILL_WARN
        else:
            rfill = FILL_GOOD
        _cell(ws1, ri, 1, r["store"])
        _cell(ws1, ri, 2, r["number"])
        _cell(ws1, ri, 3, r["category"])
        _cell(ws1, ri, 4, r["desc"])
        _cell(ws1, ri, 5, r["price"], fmt='"$"#,##0.00', fill=rfill)
        _cell(ws1, ri, 6, r["last_price"], fmt='"$"#,##0.00', fill=rfill)
        _cell(ws1, ri, 7, r["discount_dollars"], fmt='"$"#,##0.00', fill=rfill)
        _cell(ws1, ri, 8, r["discount_pct"], fmt='0.0%', fill=rfill)
        _cell(ws1, ri, 9, "YES" if r["sold_above_list"] else "")
        _cell(ws1, ri, 10, "⛔ YES" if r["into_loss"] else "")
        _cell(ws1, ri, 11, "generic" if r["generic_sku"] else ("placeholder" if r["placeholder_price"] else ""))
        _cell(ws1, ri, 12, "⛔ CRITICAL" if r["into_loss"] else ("🚨" if r["flag"] else ""))

    for ci, w in enumerate([7, 12, 20, 46, 12, 11, 11, 11, 15, 12, 12, 12], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Summary tab ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]; t2.value = f"Valley Pawn — Discount Review Summary  |  {ds}"
    t2.fill = PatternFill("solid", fgColor="0D1B40")
    t2.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    t2.alignment = Alignment(horizontal="center", vertical="center")

    real = [r for r in valued if not (r["generic_sku"] or r["placeholder_price"])]

    ws2["A3"] = "Store Leaderboard (real SKUs only) — Rank 1 = lowest discount %, Rank 5 = highest"
    ws2["A3"].font = FONT_BOLD
    hdrs2 = ["Rank", "Store", "Items Sold", "Wtd Avg Discount %", "Total Discount $", "Flags", "Into-a-Loss"]
    for ci, h in enumerate(hdrs2, 1):
        _hdr(ws2, 4, ci, h)
    stores = sorted(set(r["store"] for r in real))
    leaderboard_rows = []
    for store in stores:
        si = [r for r in real if r["store"] == store]
        disc_sum = sum(r["discount_dollars"] for r in si)
        price_sum = sum(r["price"] for r in si)
        wavg = (disc_sum / price_sum) if price_sum else None
        leaderboard_rows.append((store, len(si), wavg, disc_sum,
                                  sum(1 for r in si if r["flag"]),
                                  sum(1 for r in si if r["into_loss"])))
    # Rank ascending by discount % (Joshua 2026-08-23): lowest discount = 1st place,
    # highest discount = last place — same convention as the Slack post.
    leaderboard_rows.sort(key=lambda x: x[2] if x[2] is not None else 0.0)
    for rank, (store, n, wavg, disc_sum, flags, loss) in enumerate(leaderboard_rows, 1):
        ri = rank + 4
        _cell(ws2, ri, 1, rank, bold=True)
        _cell(ws2, ri, 2, store, bold=True)
        _cell(ws2, ri, 3, n)
        _cell(ws2, ri, 4, wavg, fmt='0.0%', fill=(FILL_FLAG if wavg is not None and wavg >= FLAG_PCT else None))
        _cell(ws2, ri, 5, disc_sum, fmt='"$"#,##0.00')
        _cell(ws2, ri, 6, flags, fill=(FILL_FLAG if flags else None))
        _cell(ws2, ri, 7, loss, fill=(FILL_CRIT if loss else None))
    for ci, w in enumerate([6, 10, 12, 18, 16, 9, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    next_row = 5 + len(leaderboard_rows) + 2
    ws2.cell(row=next_row, column=1, value="Top 10 by Discount % — company-wide").font = FONT_BOLD
    hdrs3 = ["Store", "Item #", "Description", "Ticket", "Sale", "Discount %", "Discount $"]
    for ci, h in enumerate(hdrs3, 1):
        _hdr(ws2, next_row + 1, ci, h)
    ranked_pct = sorted([r for r in real if r["discount_pct"] is not None], key=lambda r: -r["discount_pct"])[:10]
    for ri, r in enumerate(ranked_pct, next_row + 2):
        _cell(ws2, ri, 1, r["store"])
        _cell(ws2, ri, 2, r["number"])
        _cell(ws2, ri, 3, r["desc"])
        _cell(ws2, ri, 4, r["price"], fmt='"$"#,##0.00')
        _cell(ws2, ri, 5, r["last_price"], fmt='"$"#,##0.00')
        _cell(ws2, ri, 6, r["discount_pct"], fmt='0.0%', fill=FILL_FLAG if r["flag"] else None)
        _cell(ws2, ri, 7, r["discount_dollars"], fmt='"$"#,##0.00')

    next_row2 = next_row + 2 + len(ranked_pct) + 2
    ws2.cell(row=next_row2, column=1, value="Top 10 by Discount $ — company-wide").font = FONT_BOLD
    for ci, h in enumerate(hdrs3, 1):
        _hdr(ws2, next_row2 + 1, ci, h)
    ranked_dollars = sorted(real, key=lambda r: -r["discount_dollars"])[:10]
    for ri, r in enumerate(ranked_dollars, next_row2 + 2):
        _cell(ws2, ri, 1, r["store"])
        _cell(ws2, ri, 2, r["number"])
        _cell(ws2, ri, 3, r["desc"])
        _cell(ws2, ri, 4, r["price"], fmt='"$"#,##0.00')
        _cell(ws2, ri, 5, r["last_price"], fmt='"$"#,##0.00')
        _cell(ws2, ri, 6, r["discount_pct"], fmt='0.0%', fill=FILL_FLAG if r["flag"] else None)
        _cell(ws2, ri, 7, r["discount_dollars"], fmt='"$"#,##0.00')

    # ── Flags tab ────────────────────────────────────────────────────────────
    flags_list = sorted([r for r in real if r["flag"]], key=lambda r: -(r["discount_pct"] or 0))
    if flags_list:
        ws3 = wb.create_sheet("Flags")
        hdrs4 = ["Store", "Item #", "Category", "Description", "Ticket", "Sale",
                 "Discount %", "Discount $", "Into a Loss?"]
        for ci, h in enumerate(hdrs4, 1):
            _hdr(ws3, 1, ci, h)
        for ri, r in enumerate(flags_list, 2):
            _cell(ws3, ri, 1, r["store"])
            _cell(ws3, ri, 2, r["number"])
            _cell(ws3, ri, 3, r["category"])
            _cell(ws3, ri, 4, r["desc"], fill=(FILL_CRIT if r["into_loss"] else FILL_FLAG))
            _cell(ws3, ri, 5, r["price"], fmt='"$"#,##0.00')
            _cell(ws3, ri, 6, r["last_price"], fmt='"$"#,##0.00', fill=(FILL_CRIT if r["into_loss"] else FILL_FLAG))
            _cell(ws3, ri, 7, r["discount_pct"], fmt='0.0%', fill=(FILL_CRIT if r["into_loss"] else FILL_FLAG))
            _cell(ws3, ri, 8, r["discount_dollars"], fmt='"$"#,##0.00')
            _cell(ws3, ri, 9, "⛔ YES" if r["into_loss"] else "")
        for ci, w in enumerate([7, 12, 20, 46, 11, 11, 11, 11, 12], 1):
            ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(path)
    return True


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    date = resolve_date(sys.argv[1] if len(sys.argv) > 1 else None)
    date_str = date.isoformat()
    print(f"=== Discount Review  {date_str} ===")
    print(f"    Flag threshold: >={int(FLAG_PCT*100)}% off OR >=${FLAG_DOLLARS:.0f} off ticket price")

    os.makedirs(DAILY_DIR, exist_ok=True)
    summary_path = os.path.join(DAILY_DIR, f"{date_str}_discount_review_summary.json")
    xlsx_path = os.path.join(DAILY_DIR, f"{date_str}_discount_review.xlsx")

    raw, missing = load_sold_for_date(date)
    if not raw:
        msg = f"No sold-item data files found for {date_str} (missing: {', '.join(missing) or 'all stores'})"
        print(f"INFO: {msg}")
        summary = {
            "date": date_str, "items": 0, "avg_discount_pct": None, "flags": 0,
            "into_loss": 0, "stores": {}, "missing_stores": missing,
            "excel_path": None, "slack_posted": False, "slack_skipped": True,
            "info": msg,
        }
        with open(summary_path, "w") as f:
            json.dump(summary, f, indent=2)
        print(f"JSON -> {summary_path}")
        sys.exit(0)

    valued = [compute_discount(r) for r in raw]
    real = [r for r in valued if not (r["generic_sku"] or r["placeholder_price"])]
    generic_n = len(valued) - len(real)
    stores_seen = sorted(set(r["store"] for r in valued))
    print(f"Loaded {len(valued)} sold items from {len(stores_seen)} store(s): {', '.join(stores_seen)} "
          f"({generic_n} generic-SKU rows excluded from ranking)")

    store_summaries = {}
    total_disc = 0.0
    total_price = 0.0
    for store in stores_seen:
        si = [r for r in real if r["store"] == store]
        disc_sum = sum(r["discount_dollars"] for r in si)
        price_sum = sum(r["price"] for r in si)
        total_disc += disc_sum
        total_price += price_sum
        store_summaries[store] = {
            "items": len(si),
            "wtd_avg_discount_pct": round(disc_sum / price_sum, 4) if price_sum else None,
            "total_discount_dollars": round(disc_sum, 2),
            "flags": sum(1 for r in si if r["flag"]),
            "into_loss": sum(1 for r in si if r["into_loss"]),
        }

    flags_all = [r for r in real if r["flag"]]
    into_loss_all = [r for r in real if r["into_loss"]]

    summary = {
        "date": date_str,
        "items": len(real),
        "generic_sku_excluded": generic_n,
        "avg_discount_pct": round(total_disc / total_price, 4) if total_price else None,
        "total_discount_dollars": round(total_disc, 2),
        "flags": len(flags_all),
        "into_loss": len(into_loss_all),
        "stores": store_summaries,
        "missing_stores": missing,
        "excel_path": xlsx_path,
        "slack_posted": False,
        "slack_skipped": False,
        "slack_message": None,
    }

    xl_ok = write_excel(valued, date, xlsx_path)
    if xl_ok:
        print(f"Excel -> {xlsx_path}")
    summary["excel_path"] = xlsx_path if xl_ok else None

    # Running calendar-year discount totals, by store and company (added 2026-08-13).
    # Computed from the per-day summaries already on disk plus this run's numbers, so it
    # self-heals and never double-counts a re-run. Stored in the summary JSON too, so the
    # figure the team saw on any given day stays auditable after the fact.
    _today_store_totals = {st: sv["total_discount_dollars"] for st, sv in store_summaries.items()}
    store_ytd, company_ytd, ytd_days, ytd_year = compute_ytd(date, _today_store_totals, total_disc)
    summary["ytd_year"] = ytd_year
    summary["ytd_days_counted"] = ytd_days
    summary["ytd_total_discount_dollars"] = round(company_ytd, 2)
    summary["ytd_by_store"] = {st: round(v, 2) for st, v in store_ytd.items()}

    slack_msg = build_slack_message(valued, date, missing,
                                    store_ytd=store_ytd, company_ytd=company_ytd,
                                    ytd_days=ytd_days, ytd_year=ytd_year)
    summary["slack_message"] = slack_msg
    if slack_msg is None:
        print(f"Slack post skipped — only {len(real)} real item(s) (min 3 required).")
        summary["slack_skipped"] = True
    elif not SLACK_CHANNEL:
        print("Slack post skipped — #discount-review channel not yet created/configured "
              "(set DISCOUNT_REVIEW_SLACK_CHANNEL once it exists).")
        summary["slack_skipped"] = True
        summary["slack_error"] = "channel_not_configured"
    else:
        tok = _get_slack_token()
        if not tok:
            print("WARNING: SLACK_BOT_TOKEN not found -- Slack post skipped.")
            summary["slack_skipped"] = True
            summary["slack_error"] = "token_not_found"
        else:
            ok = slack_post(slack_msg, tok, SLACK_CHANNEL)
            summary["slack_posted"] = ok
            if ok:
                print(f"Slack -> {SLACK_CHANNEL}")
            else:
                print("Slack post failed -- check token/channel permissions", file=sys.stderr)
                summary["slack_error"] = "post_failed"

    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"JSON -> {summary_path}")

    if flags_all:
        print(f"\nFlagged items -- discount >= {int(FLAG_PCT*100)}% or >= ${FLAG_DOLLARS:.0f} "
              f"({len(flags_all)} total, {len(into_loss_all)} sold into a loss):")
        for r in sorted(flags_all, key=lambda x: -(x["discount_pct"] or 0))[:20]:
            pstr = f"{r['discount_pct']*100:.0f}%" if r["discount_pct"] is not None else "-"
            print(f"  {r['store']:4} {pstr:>5} off  ${r['discount_dollars']:>7,.0f}  "
                  f"${r['price']:>7,.0f}->${r['last_price']:>7,.0f}  {str(r.get('desc',''))[:40]}")
    else:
        print("\nNo flags -- no items discounted past threshold. OK")

    print("\nDone.")


if __name__ == "__main__":
    main()
