#!/usr/bin/env python3
"""
Valley Pawn — Daily Sales Tax Sweep engine (Phase 1: DRY-RUN, no money moves).

Sizes a FIXED daily sweep from the most recent monthly tax total (which already
has eBay backed out) and logs it to a running reserve ledger. Joshua's directive
2026-07-20: daily amount does NOT need to be exact; the monthly ST-9 filing is the
exact number and squares up the reserve. So NO daily Bravo pull, NO daily eBay pull.

  daily_sweep = (latest month's total Taxes Due, all 5 stores) / days_in_current_month

MODE is DRY-RUN until a programmable bank account exists (Phase 2). In dry-run it
only logs "would sweep $X, reserve would be $Y" — it never moves money.

Prints a JSON summary on stdout for the scheduled task to read and post to Slack.
"""
import json, sys, os, shutil, calendar
from datetime import date
import openpyxl
from openpyxl.comments import Comment

PROJ = "/Users/joshuadavis/Documents/Claude/Projects/Sales Tax"
SRC  = os.path.join(PROJ, "Sales Tax.xlsx")
LEDGER = os.path.join(PROJ, "Sales_Tax_Reserve_Ledger.xlsx")
DRIVE_DIR = "/Users/joshuadavis/Library/CloudStorage/GoogleDrive-jdavis@fcfpawn.com/Shared drives/Valley Pawn Drive/Bookkeeping/Sales Tax Automation"
DRIVE_LEDGER = os.path.join(DRIVE_DIR, "Sales_Tax_Reserve_Ledger.xlsx")

MODE = "DRY-RUN"          # flip to "LIVE" in Phase 2 when the bank API is wired
DUE_COLS = ["E", "H", "K", "N", "Q"]   # Taxes Due per store: CUL HAR LEX ROA WAY
MONTHS = ["JAN","FEB","MAR","APRIL","MAY","JUNE","JULY","AUG","SEP","OCT","NOV","DEC"]


def latest_month_total(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rate = ws["G1"].value
    last = None
    for r in range(5, ws.max_row + 1):
        lbl = ws.cell(row=r, column=1).value
        if lbl is None:
            continue
        vals = [ws[f"{c}{r}"].value for c in DUE_COLS]
        if any(isinstance(v, (int, float)) and v for v in vals):
            tot = sum(v for v in vals if isinstance(v, (int, float)))
            last = {"row": r, "label": str(lbl).strip(), "total": round(tot, 2)}
    return last, rate


def ensure_ledger():
    if os.path.exists(LEDGER):
        return openpyxl.load_workbook(LEDGER)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reserve Ledger"
    headers = ["Date", "Mode", "Daily Sweep $", "Cumulative Reserve $",
               "Source Month", "Source Month Tax $", "Days In Month", "Note"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
    return wb


def main():
    today = date.today()
    days_in_month = calendar.monthrange(today.year, today.month)[1]
    src, rate = latest_month_total(SRC)
    if not src:
        print(json.dumps({"ok": False, "error": "No populated month found in Sales Tax.xlsx"}))
        sys.exit(1)

    daily = round(src["total"] / days_in_month, 2)

    wb = ensure_ledger()
    ws = wb["Reserve Ledger"] if "Reserve Ledger" in wb.sheetnames else wb.active
    iso = today.isoformat()

    # idempotent: if today already logged, update that row instead of duplicating
    target = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value) == iso:
            target = r
            break
    if target is None:
        target = ws.max_row + 1
        ws.cell(row=target, column=1, value=iso)

    ws.cell(row=target, column=2, value=MODE)
    ws.cell(row=target, column=3, value=daily).number_format = '$#,##0.00'
    ws.cell(row=target, column=5, value=src["label"])
    ws.cell(row=target, column=6, value=src["total"]).number_format = '$#,##0.00'
    ws.cell(row=target, column=7, value=days_in_month)
    ws.cell(row=target, column=8,
            value=f"{MODE}: would sweep from operating to Sales Tax Reserve. eBay already netted in monthly source.")

    # recompute cumulative down the whole column (handles updates/backfills cleanly)
    cum = 0.0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=3).value
        if isinstance(v, (int, float)):
            cum += v
        ws.cell(row=r, column=4, value=round(cum, 2)).number_format = '$#,##0.00'
    cumulative = round(cum, 2)

    wb.save(LEDGER)
    # mirror to Google Drive (Joshua's directive: docs live in Valley Pawn Drive)
    try:
        os.makedirs(DRIVE_DIR, exist_ok=True)
        shutil.copy(LEDGER, DRIVE_LEDGER)
        drive_ok = True
    except Exception as e:
        drive_ok = False

    print(json.dumps({
        "ok": True, "mode": MODE, "date": iso,
        "daily_sweep": daily, "cumulative_reserve": cumulative,
        "source_month": src["label"], "source_total": src["total"],
        "days_in_month": days_in_month, "rate": rate,
        "ledger": LEDGER, "drive_mirror": drive_ok,
    }))


if __name__ == "__main__":
    main()
