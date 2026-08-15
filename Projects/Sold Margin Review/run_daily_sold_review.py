#!/usr/bin/env python3
"""
Valley Pawn — Daily Sold Margin Review ("Sold Review")

Reads yesterday's SOLD items per store from the Bravo Data Extraction pipeline
output (the "Claude Sold Inv Details" saved report, pipeline cell
`jewelry-margin-sold` — despite the jewelry-sounding cell name, the AHK
handler carries no category filter and returns ALL sold items), computes
realized margin % and margin $ per item directly from Bravo's own Cost vs
Last Sold Price fields (no external valuation needed — Bravo already knows
what we paid and what we actually got), flags items that sold too cheap, and
posts a per-store + company summary to #sold-review.

CHANGED 2026-08-13: originally targeted a separate saved report ("Claude Sold
Yesterday" / cell `sold-yesterday` / handler SoldYesterday.ahk) that proved
unreliable on its first live smoke test (UIA could not reliably select it —
0/5 stores on 2026-08-13). `discount-review` (a sibling task built the next
day, 2026-07-29) had already proven out "Claude Sold Inv Details" /
`jewelry-margin-sold` as a working, selectable report carrying the exact
columns needed (Cost, Price, Last Sold Price, Category, Description, Date) —
so sold-review was switched to reuse that same proven cell instead of
maintaining two near-duplicate "what sold yesterday" pulls. `sold-yesterday`
/ SoldYesterday.ahk is left in place untouched (additive-only) in case a
future project wants it; this script just no longer depends on it.

IMPORTANT — Price vs Last Sold Price: in this report, "Price" is the
ticketed/asking price, NOT what the customer actually paid. "Last Sold
Price" is the real realized sale price. Realized-margin math below uses
Last Sold Price. Do not swap these — that mistake would silently overstate
margin on anything sold at a discount.

This is the SALES-side counterpart to the daily-intake-margin / pawn-walk
pipeline (which grades what we PAY on the way IN, against an external market
estimate, because there's no internal cost basis yet at intake). Here the
direction is reversed: an item that already sold has a known, exact cost
(what we paid at intake or acquisition) and a known, exact sale price — the
realized margin is exact, no estimation required.

Usage:
    python run_daily_sold_review.py [YYYY-MM-DD]
    Date defaults to yesterday. Pass a specific date for re-runs.

Output (daily/ subfolder next to this script):
    daily/{DATE}_sold_review.xlsx
    daily/{DATE}_sold_review_summary.json

Exit codes:
    0 — run complete (even if Slack token missing — check JSON)
    1 — critical error reading input

Notes:
  - Additive only. Does not touch any other project's files.
  - Decoupled from the extraction side by design (Joshua may have a separate
    task/session producing the CSV): this loader tries several plausible
    filename patterns for the same store/date before giving up, so it works
    whether the data was produced by this project's own `sold-yesterday`
    pipeline cell or by another task that lands a CSV in the same output
    folder under a similar name.
"""

from __future__ import annotations
import csv, glob, re, json, os, sys, datetime, statistics
import urllib.request

# ── Paths ─────────────────────────────────────────────────────────────────────
HERE         = os.path.dirname(os.path.abspath(__file__))
BRAVO_OUTPUT = "/Users/joshuadavis/Documents/Claude/Projects/Bravo Data Extraction/output/"
DAILY_DIR    = os.path.join(HERE, "daily")

# ── Thresholds & config (Expert Board decision, 2026-07-23) ───────────────────
# Board panel: retail-operations analyst, pricing/margin analyst, data-integrity
# reviewer. Company retail margin has run 52-54% over the last several months
# (per monthly analytics: 52.0% Jun26, 53.5% YTD, 52.8% T12M) — so the 50%
# target below is not aspirational, it is where the company already sits on
# average. A flag threshold has to sit meaningfully BELOW that average or it
# fires on ordinary variance; 25% (half the target) was chosen to mirror the
# buy-side pawn-walk convention (30% flag vs 50% target — roughly 60% of
# target) while giving a bit more room, since retail sales have legitimate
# reasons to run a bit thinner (clearance, bundle deals, employee discounts)
# that acquisition-side "overpay" doesn't have.
SLACK_CHANNEL   = "C0BK802MP43"   # #sold-review
TARGET_MARGIN   = 0.50            # 50% gross-margin target (matches company retail benchmark)
FLAG_MARGIN     = 0.25            # flag items below this realized margin ("sold too cheap")
AGED_DAYS_MARKDOWN = 90           # items on shelf 90+ days get an "(aged clearance)" tag,
                                   # not suppressed — still shown, just contextualized so a
                                   # legitimate markdown-to-move doesn't read as a pricing mistake.

# Candidate input filename patterns, tried in order, for a given date/store.
# {d} = ISO date (also used as the single-day range start==end)
# Primary as of 2026-08-13 (evening): the `sold-discount-detail` cell — a strictly additive
# clone of jewelry-margin-sold (handler reports/SoldDiscountDetail.ahk) that fixes TWO real
# bugs the old handler still has:
#   1. Zero-sale days wrote NO csv at all, making "ran, no sales" indistinguishable on disk
#      from "never ran" — this task then reported those stores as missing_stores.
#   2. The grid-capture searched the whole UIA root for DataItems and could latch onto the
#      wrong grid entirely — on 2026-08-13 that wrote WAY's Global Access store picker
#      (DisplayCode,Store) to disk as if it were 5 rows of sold inventory.
# Both were proven fixed live on all 5 stores 2026-08-13. jewelry-margin-sold patterns are
# RETAINED below as fallbacks so previously-pulled data still parses; the old cell itself is
# untouched (the jewelry-scrap project still owns it). Legacy sold-yesterday patterns kept
# below that, in case that cell is ever fixed and used again by another task.
_FILENAME_CANDIDATES = [
    "{d}_to_{d}_{store}_sold-discount-detail.csv",
    "{d}_{store}_sold-discount-detail.csv",
    "{d}_to_{d}_{store}_jewelry-margin-sold.csv",
    "{d}_{store}_jewelry-margin-sold.csv",
    "{d}_to_{d}_{store}_sold-inv-details.csv",
    "{d}_{store}_sold-inv-details.csv",
    "{d}_to_{d}_{store}_sold-yesterday.csv",
    "{d}_{store}_sold-yesterday.csv",
    "{d}_to_{d}_{store}_claude-sold-yesterday.csv",
    "{d}_{store}_claude-sold-yesterday.csv",
]

STORES = ["CUL", "HAR", "LEX", "ROA", "WAY"]


# ── Open-stores gate (Joshua, 2026-08-12 pattern) ──────────────────────────────
# The scheduled task only ever REQUESTS a Bravo pull for open stores on a given
# date (Sunday = none, Wednesday = CUL only, else all 5) — see sold-review's
# SKILL.md STEP 0.5. This compile script mirrors that same logic purely for
# REPORTING clarity: a store with no CSV because it was legitimately closed
# should not be lumped into "missing_stores" next to a store that was open but
# whose pull genuinely failed. Added 2026-08-13 after the first live CUL-only
# rehearsal made the old wording ("No data file for: HAR, LEX, ROA, WAY") read
# like 4 failures on a Wednesday when it was correct, expected behavior.
def open_stores_for(date: datetime.date) -> list[str]:
    wd = date.weekday()  # Monday=0 ... Sunday=6
    if wd == 6:      # Sunday
        return []
    if wd == 2:       # Wednesday
        return ["CUL"]
    return list(STORES)


# ── Date resolution ────────────────────────────────────────────────────────────
def resolve_date(arg: str | None) -> datetime.date:
    if arg:
        try:
            return datetime.date.fromisoformat(arg)
        except ValueError:
            print(f"Bad date '{arg}' — expected YYYY-MM-DD", file=sys.stderr)
            sys.exit(1)
    return datetime.date.today() - datetime.timedelta(days=1)


# ── Slack token resolution (same pattern as pawn-walk) ────────────────────────
def _get_slack_token() -> str | None:
    tok = os.environ.get("SLACK_BOT_TOKEN", "").strip()
    if tok.startswith("xoxb-"):
        return tok
    for cpath in [
        os.path.join(BRAVO_OUTPUT, "..", "slack_config.json"),
        os.path.expanduser("~/Documents/Claude/Projects/Bravo Data Extraction/slack_config.json"),
        os.path.expanduser("~/.vp_slack_config.json"),
    ]:
        try:
            with open(os.path.normpath(cpath)) as f:
                d = json.load(f)
            t = (d.get("SLACK_BOT_TOKEN") or d.get("slack_bot_token") or "").strip()
            if t.startswith("xoxb-"):
                return t
        except Exception:
            pass
    for profile in [
        os.path.expanduser("~/.bash_profile"),
        os.path.expanduser("~/.zshenv"),
        os.path.expanduser("~/.profile"),
    ]:
        try:
            txt = open(profile).read()
            m = re.search(r'(?:export\s+)?SLACK_BOT_TOKEN=["\']?(xoxb-[^\s"\']+)', txt)
            if m:
                return m.group(1).rstrip("\"'")
        except Exception:
            pass
    return None


# ── Money parsing ──────────────────────────────────────────────────────────────
def money(s):
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace("$", "").replace(",", "").strip()
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None


def _find_col(fieldnames, *candidates):
    """Case/space-insensitive header lookup — returns the actual fieldname or None."""
    norm = {re.sub(r'[^a-z0-9]', '', fn.lower()): fn for fn in (fieldnames or [])}
    for c in candidates:
        key = re.sub(r'[^a-z0-9]', '', c.lower())
        if key in norm:
            return norm[key]
    return None


def _parse_mdy_or_iso(s):
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ── Load sold items for a specific date, across all 5 stores ─────────────────
def load_sold_for_date(date: datetime.date) -> tuple[list[dict], list[str]]:
    """Returns (rows, stores_with_no_file). Only checks stores that were OPEN on
    this date per open_stores_for() — a store closed that day (Sunday, or any
    non-CUL store on a Wednesday) is correctly excluded entirely, not reported
    as missing. Tries each filename candidate per store in order; first match
    wins. A store that WAS open but has no matching file is reported separately
    (not the same as a store with a file but zero rows — that's a legitimate
    no-sales day for an open store)."""
    ds = date.isoformat()
    rows: list[dict] = []
    missing_files: list[str] = []

    for store in open_stores_for(date):
        path = None
        for pat in _FILENAME_CANDIDATES:
            candidate = os.path.join(BRAVO_OUTPUT, pat.format(d=ds, store=store))
            if os.path.exists(candidate):
                path = candidate
                break
        if path is None:
            missing_files.append(store)
            continue

        try:
            with open(path, newline='', encoding='utf-8-sig') as fh:
                reader = csv.DictReader(fh)
                fns = reader.fieldnames or []
                col_cost   = _find_col(fns, "Cost", "Item Cost", "Unit Cost")
                # "Last Sold Price" (actual realized price) MUST win over "Price" (ticket/
                # asking price) — see module docstring. Sale Price/Selling Price/Sold Price
                # kept as fallbacks for any other CSV shape that lands in this same folder.
                col_price  = _find_col(fns, "Last Sold Price", "Sale Price", "Sold Price",
                                        "Selling Price", "Price")
                col_desc   = _find_col(fns, "Description", "Full Description", "Item Description")
                col_cat    = _find_col(fns, "Category")
                col_ticket = _find_col(fns, "Number", "Inventory #", "Inventory Number", "Item #",
                                        "Acquired Ticket #", "Ticket Number", "Ticket #")
                col_date   = _find_col(fns, "Date", "Date Sold", "Sold Date", "Sale Date")
                # jewelry-margin-sold's CSV carries no shelf-age column — this will
                # legitimately resolve to None and the aged-clearance tag simply won't
                # fire (graceful degradation, not an error).
                col_days   = _find_col(fns, "Days On Shelf", "Days on Shelf", "Age", "Days Held")

                for row in reader:
                    cost  = money(row.get(col_cost)) if col_cost else None
                    price = money(row.get(col_price)) if col_price else None
                    if cost is None or price is None or price <= 0:
                        continue  # can't compute margin without both fields
                    days_on_shelf = None
                    if col_days:
                        try:
                            days_on_shelf = int(float(str(row.get(col_days) or "").strip()))
                        except (ValueError, TypeError):
                            days_on_shelf = None
                    rows.append({
                        "store":         store,
                        "ticket":        (row.get(col_ticket) or "").strip() if col_ticket else "",
                        "category":      (row.get(col_cat) or "").strip() if col_cat else "",
                        "desc":          (row.get(col_desc) or "").strip() if col_desc else "",
                        "cost":          cost,
                        "price":         price,
                        "date_sold":     (row.get(col_date) or "").strip() if col_date else "",
                        "days_on_shelf": days_on_shelf,
                        "source_file":   os.path.basename(path),
                    })
        except Exception as e:
            print(f"WARNING: could not read {path}: {e}", file=sys.stderr)
            missing_files.append(store)

    return rows, missing_files


# ── Value / margin computation ────────────────────────────────────────────────
def compute_margin(r: dict) -> dict:
    cost, price = r["cost"], r["price"]
    margin = (price - cost) / price if price else None
    margin_dollars = price - cost
    aged = bool(r.get("days_on_shelf") is not None and r["days_on_shelf"] >= AGED_DAYS_MARKDOWN)
    critical = margin is not None and margin < 0   # sold at or below cost
    flag = margin is not None and margin < FLAG_MARGIN
    meets = margin is not None and margin >= TARGET_MARGIN
    out = dict(r, margin=margin, margin_dollars=margin_dollars,
               aged=aged, critical=critical, flag=flag, meets=meets)
    return out


# ── Market benchmark (added 2026-08-14) ────────────────────────────────────────
# WHY: cost-margin alone cannot answer "did we sell it too cheap." An item bought for $10,
# worth $200, sold for $60 posts an 83% margin and never flags — it looks like a win. This
# adds the missing benchmark: what this item normally sells for, from our OWN realized
# sales (pooled across stores, ~29k rows) plus optional eBay corroboration.
# See market_benchmark.py for the blend rules and why it deliberately leans conservative.
_BENCH_ENGINE = None

def attach_market(valued: list[dict], date: datetime.date) -> list[dict]:
    """
    Adds market_* fields to each item. Fails SOFT and loudly-in-the-log: if the benchmark
    engine can't load, the daily margin report must still go out — this is an added signal,
    never a new single point of failure for a report that already works.
    """
    global _BENCH_ENGINE
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from market_benchmark import BenchmarkEngine
        if _BENCH_ENGINE is None:
            # exclude_date: hold today's sales out of the comp index, or an item would
            # help set the very benchmark it is being judged against and mask its own flag.
            _BENCH_ENGINE = BenchmarkEngine(use_external=False, exclude_date=date.isoformat())
    except Exception as e:
        print(f"WARNING: market benchmark unavailable ({e}) — margin-only report.", file=sys.stderr)
        for r in valued:
            r["market_flag"] = False
            r["market_value"] = None
        return valued

    n_flag = 0
    for r in valued:
        try:
            res = _BENCH_ENGINE.evaluate(r.get("desc") or "", r.get("category") or "",
                                         r.get("price"), r.get("cost"))
            r["market_value"]      = res["value"]
            r["market_pct"]        = res["pct_of_benchmark"]
            r["market_gap"]        = res["gap_dollars"]
            r["market_conf"]       = res["confidence"]
            r["market_source"]     = res["source"]
            r["market_ebay"]       = res.get("external")   # real Terapeak sold comp, if any
            r["market_flag"]       = res["sold_too_cheap"]
            if res["sold_too_cheap"]:
                n_flag += 1
        except Exception:
            r["market_flag"] = False
            r["market_value"] = None
    print(f"    Market benchmark: {sum(1 for r in valued if r.get('market_value'))}/{len(valued)} "
          f"items benchmarked, {n_flag} sold below market")

    # DAILY CANARY — runs inside compile so it can never be skipped by a task edit.
    # Detects the failure mode that would otherwise be invisible: the eBay market data
    # silently disappearing while the report keeps publishing as if nothing changed.
    try:
        from terapeak import selfcheck
        sc = selfcheck()
        _MARKET_HEALTH.clear(); _MARKET_HEALTH.update(sc)
        if sc["ok"]:
            print(f"    Market feed health: OK (parser verified, {sc['with_value']} comps cached, "
                  f"{sc['fresh_7d']} added in last 7d)")
        else:
            print(f"    ⚠️  MARKET FEED: {sc['note']}", file=sys.stderr)
    except Exception as e:
        print(f"    ⚠️  MARKET FEED: selfcheck unavailable ({e})", file=sys.stderr)
    return valued


# Populated by attach_market()'s canary; surfaced into the summary JSON so STEP 7 can DM.
_MARKET_HEALTH: dict = {}


# ── Fair Value v2 (added 2026-08-14, BLEND_V2_PLAN.md) ────────────────────────
# SEPARATE from the market benchmark above, on purpose. The benchmark answers
# "is this sale bad enough to alert on" and stays deliberately conservative —
# flags are UNCHANGED. Fair value answers "what SHOULD this have sold for" and
# wants accuracy: time-decayed internal comps blended with channel-normalized
# (net-of-fees, used-condition) eBay sold comps, weighted by precision, with an
# uncertainty band. Disagreements >30% are surfaced, not averaged away, and
# feed the pricing-health aggregate (fair_value.py --health).
def attach_fair_value(valued: list[dict], date: datetime.date) -> list[dict]:
    """Adds fair_* fields per item. Fails SOFT: report still goes out without it."""
    try:
        sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
        from fair_value import FairValueEngine
        eng = FairValueEngine(exclude_date=date.isoformat(), ref_date=date.isoformat(),
                              allow_live_api=True)
    except Exception as e:
        print(f"WARNING: fair-value engine unavailable ({e}) — report without fair values.",
              file=sys.stderr)
        for r in valued:
            r["fair_value"] = None
        return valued

    n_ok = n_disp = 0
    for r in valued:
        try:
            est = eng.estimate(r.get("desc") or "", r.get("category") or "",
                               r.get("price"), store=r.get("store"),
                               date=date.isoformat())
            r["fair_value"]    = est["fair"]
            r["fair_band"]     = est["band"]
            r["fair_basis"]    = est["basis"]
            r["fair_disputed"] = est["disputed"]
            if est["fair"]:
                n_ok += 1
            if est["disputed"]:
                n_disp += 1
        except Exception:
            r["fair_value"] = None
    print(f"    Fair value v2: {n_ok}/{len(valued)} items valued, {n_disp} source disputes "
          f"(disputes feed pricing health)")
    _MARKET_HEALTH["fair_value_coverage"] = n_ok
    _MARKET_HEALTH["fair_value_disputes"] = n_disp
    return valued


def _soldcomps_usage() -> int | None:
    """Today's SoldComps API request count, for the summary JSON. Never raises."""
    try:
        from soldcomps import _usage_today
        return _usage_today()
    except Exception:
        return None


def _week_mix(sat: datetime.date) -> list[str]:
    """Ticket-size mix for the week ending Saturday `sat` (Mon-Sat), straight
    from the week's sold CSVs. Returns Slack lines, or [] on any trouble —
    a missing weekly section must never block the daily post."""
    try:
        buckets = [(0, 25, "Under $25"), (25, 100, "$25-100"),
                   (100, 300, "$100-300"), (300, float("inf"), "Over $300")]
        agg = {b[2]: [0, 0.0, 0.0] for b in buckets}   # n, rev, profit
        for i in range(6):  # Mon..Sat
            d = (sat - datetime.timedelta(days=5 - i)).isoformat()
            for path in glob.glob(os.path.join(BRAVO_OUTPUT, f"{d}_to_{d}_*_sold-discount-detail.csv")):
                with open(path, newline="", encoding="utf-8-sig") as f:
                    for r in csv.DictReader(f):
                        if (r.get("Status") or "").strip().upper() != "SOLD":
                            continue
                        sale = money(r.get("Last Sold Price"))
                        cost = money(r.get("Cost"))
                        if not sale:
                            continue
                        for lo, hi, nm in buckets:
                            if lo <= sale < hi:
                                agg[nm][0] += 1
                                agg[nm][1] += sale
                                agg[nm][2] += sale - (cost or 0)
                                break
        tot_rev = sum(v[1] for v in agg.values())
        if not tot_rev:
            return []
        lines = []
        for _, _, nm in buckets:
            n, rev, prof = agg[nm]
            lines.append(f"• {nm}: {n} items · ${rev:,.0f} ({rev / tot_rev * 100:.0f}% of revenue) "
                         f"· ${prof:,.0f} profit")
        tot_n = sum(v[0] for v in agg.values())
        tot_prof = sum(v[2] for v in agg.values())
        lines.append(f"*Week total: {tot_n} items · ${tot_rev:,.0f} revenue · ${tot_prof:,.0f} profit*")
        return lines
    except Exception:
        return []


# ── Build Slack message ────────────────────────────────────────────────────────
def build_slack_message(valued: list[dict], date: datetime.date, missing_stores: list[str]) -> str | None:
    if len(valued) < 3:
        return None
    ds = date.isoformat()
    BAR = "━" * 22

    def _pct(m):
        return f"{m * 100:.0f}%" if m is not None else "—"

    lines = [f"🛒 *Sold Review — {ds}*", "",
             f"> ⚠️ Target ~{int(TARGET_MARGIN*100)}% realized margin. 🚨 = below {int(FLAG_MARGIN*100)}% (sold too cheap).",
             ""]

    stores = sorted(set(r["store"] for r in valued))
    total_items = 0
    total_margins = []
    total_flags = 0
    for st in stores:
        si = [r for r in valued if r["store"] == st]
        mm = [r["margin"] for r in si if r["margin"] is not None]
        fl = sum(1 for r in si if r["flag"])
        total_items += len(si)
        total_margins += mm
        total_flags += fl
        avg = statistics.mean(mm) if mm else None
        stt = "—" if avg is None else ("✅" if avg >= TARGET_MARGIN else "🚨")
        fw = "flag" if fl == 1 else "flags"
        lines.append(f"*{st}* — {len(si)} items sold | Avg margin {_pct(avg)} {stt} | {fl} {fw}")
    if missing_stores:
        lines.append(f"_No data file for: {', '.join(missing_stores)}_")

    cavg = statistics.mean(total_margins) if total_margins else None
    lines.append("")
    lines.append(f"*Company:* {total_items} items sold | Avg margin {_pct(cavg)} | {total_flags} total flags")

    # ── Day scorecard (strategy decision 2026-08-14: 5-second read of the day) ─
    revenue = sum(r["price"] for r in valued)
    profit  = sum(r["price"] - r["cost"] for r in valued)
    # "left on table": graded items that sold >= $25 under fair value
    left = sum((r["fair_value"] - r["price"]) for r in valued
               if r.get("fair_value") and (r["fair_value"] - r["price"]) >= 25)
    lines.append(f"💰 Revenue ${revenue:,.0f} · Profit ${profit:,.0f} · "
                 f"Left on table ~${left:,.0f}")

    flagged = sorted([r for r in valued if r["flag"]],
                      key=lambda r: (r["margin"] if r["margin"] is not None else 0.0))
    if flagged:
        lines.append("")
        lines.append(BAR)
        lines.append(f"*🚨 SOLD TOO CHEAP ({len(flagged)})* — store · sale · cost · margin · item")
        lines.append(BAR)
        for r in flagged[:12]:
            mstr = _pct(r["margin"])
            desc = (r.get("desc") or r.get("category") or "item").strip()
            if len(desc) > 40:
                desc = desc[:39] + "…"
            crit = " ⛔CRITICAL(below cost)" if r["critical"] else ""
            aged_tag = f" (aged {r['days_on_shelf']}d — likely clearance)" if r.get("aged") else ""
            # Fair value ± band (BLEND_V2): what this SHOULD have brought, with
            # honest uncertainty — a thin estimate must LOOK thin.
            fv = ""
            if r.get("fair_value"):
                band = f" ±${r['fair_band']:,.0f}" if r.get("fair_band") else ""
                fv = f" · fair ~${r['fair_value']:,.0f}{band}"
            lines.append(f"• {r['store']} · ${r['price']:,.0f} sale · ${r['cost']:,.0f} cost · {mstr}{crit}{fv} · {desc}{aged_tag}")
        if len(flagged) > 12:
            lines.append(f"…and {len(flagged) - 12} more — full detail in the spreadsheet")

    # ── Sold below what we normally get (market benchmark, added 2026-08-14) ──────
    # Separate section on purpose: this is a DIFFERENT question from margin. An item can
    # post a healthy margin and still appear here, and that is exactly the case the margin
    # flag was blind to. Shows the evidence (how many of our own past sales back it) so a
    # thin comp reads as thin rather than as authority.
    below_mkt = sorted([r for r in valued if r.get("market_flag")],
                       key=lambda r: -(r.get("market_gap") or 0))
    if below_mkt:
        lines.append("")
        lines.append(BAR)
        lines.append(f"*📉 SOLD BELOW WHAT WE NORMALLY GET ({len(below_mkt)})* — sale · typical · left on table")
        lines.append(BAR)
        for r in below_mkt[:10]:
            desc = (r.get("desc") or r.get("category") or "item").strip()
            if len(desc) > 38:
                desc = desc[:37] + "…"
            src = (r.get("market_source") or "").replace("internal ", "")
            # Show the eBay sold figure alongside ours when they differ materially. The
            # flag itself uses the conservative (lower) benchmark, but hiding the market
            # number would bury the more important finding: where eBay is well above our
            # own typical price, we may be underpricing that whole category in-store.
            mkt = ""
            ev = r.get("market_ebay")
            if ev and r.get("market_value") and abs(ev - r["market_value"]) / max(ev, r["market_value"]) > 0.20:
                mkt = f" · eBay sold ~${ev:,.0f}"
            lines.append(
                f"• {r['store']} · ${r['price']:,.0f} sale · ${r['market_value']:,.0f} typical{mkt} · "
                f"−${r.get('market_gap', 0):,.0f} · {desc} _({src})_")
        if len(below_mkt) > 10:
            lines.append(f"…and {len(below_mkt) - 10} more — full detail in the spreadsheet")

    # ── Big-ticket review (strategy decision 2026-08-14) ──────────────────────
    # Items >=$100 carry ~80%+ of daily revenue and margin the thinnest, so every
    # one gets a verdict vs fair value. Small items are deliberately NOT graded
    # per-item here (below-cost criticals still flag above) — a $8 comic doesn't
    # earn counter attention; a $400 laptop does.
    big = sorted([r for r in valued if r["price"] >= 100], key=lambda r: -r["price"])
    if big:
        lines.append("")
        lines.append(BAR)
        lines.append(f"*💵 BIG-TICKET REVIEW ({len(big)} items ≥$100 — "
                     f"${sum(r['price'] for r in big):,.0f} of the day)*")
        lines.append(BAR)
        for r in big[:12]:
            desc = (r.get("desc") or r.get("category") or "item").strip()
            if len(desc) > 36:
                desc = desc[:35] + "…"
            fv, band = r.get("fair_value"), r.get("fair_band") or 0
            basis = (r.get("fair_basis") or "") + (r.get("market_source") or "")
            if not fv and "melt" in basis.lower():
                verdict = "🪙 gold/silver — melt-priced"
            elif not fv:
                verdict = "no benchmark"
            elif r["price"] >= fv - max(band, 25):
                verdict = f"✔ got market (fair ~${fv:,.0f})"
            else:
                verdict = f"▼ ${fv - r['price']:,.0f} under (fair ~${fv:,.0f})"
            lines.append(f"• {r['store']} · ${r['price']:,.0f} · {desc} · {verdict}")
        if len(big) > 12:
            lines.append(f"…and {len(big) - 12} more in the spreadsheet")

    # ── Saturday wrap: the week's ticket-size mix (posted Sunday morning) ─────
    # Weekly, not daily, on purpose: mix is a trend question. Answers "does the
    # small stuff actually matter" with a week of evidence at a time.
    if date.weekday() == 5:  # reporting a Saturday -> week just ended
        wk = _week_mix(date)
        if wk:
            lines.append("")
            lines.append(BAR)
            lines.append("*📅 WEEK IN REVIEW — where the money was*")
            lines.append(BAR)
            lines.extend(wk)

    # ── Pricing health (BLEND_V2 Phase 3) — the systematic-underpricing radar ─
    # Per-category: our realized prices vs eBay NET (fees + shipping removed,
    # used-condition). Gated at n>=30 pairs per category so this section only
    # ever prints statistically meaningful findings — it is silent for weeks
    # while the pairs accumulate, then starts telling the truth. Per-item flags
    # structurally cannot see store-wide underpricing; this can.
    try:
        from fair_value import health_lines
        hl = health_lines(min_n=30)
        if hl:
            lines.append("")
            lines.append(BAR)
            lines.append("*📊 PRICING HEALTH — our prices vs eBay net (categories with n≥30)*")
            lines.append(BAR)
            lines.extend(f"• {h}" for h in hl[:8])
    except Exception:
        pass

    return "\n".join(lines)


# ── Post to Slack ──────────────────────────────────────────────────────────────
def slack_post(text: str, token: str) -> bool:
    payload = json.dumps({"channel": SLACK_CHANNEL, "text": text, "mrkdwn": True}).encode("utf-8")
    req = urllib.request.Request(
        "https://slack.com/api/chat.postMessage",
        data=payload,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json; charset=utf-8"},
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            body = json.load(resp)
            if not body.get("ok"):
                print(f"Slack API error: {body.get('error', '?')}", file=sys.stderr)
            return bool(body.get("ok"))
    except Exception as e:
        print(f"Slack request failed: {e}", file=sys.stderr)
        return False


# ── Excel report ───────────────────────────────────────────────────────────────
def write_excel(valued: list[dict], date: datetime.date, path: str) -> bool:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxl not installed — Excel skipped.", file=sys.stderr)
        return False

    FILL_HDR  = PatternFill("solid", fgColor="1F3864")
    FILL_CRIT = PatternFill("solid", fgColor="FF9999")
    FILL_FLAG = PatternFill("solid", fgColor="FFCCCC")
    FILL_WARN = PatternFill("solid", fgColor="FFF0CC")
    FILL_GOOD = PatternFill("solid", fgColor="CCFFCC")
    FONT_HDR  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    FONT_BODY = Font(name="Calibri", size=10)
    FONT_BOLD = Font(name="Calibri", bold=True, size=10)
    THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = FILL_HDR; c.font = FONT_HDR; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _cell(ws, row, col, val, fmt=None, fill=None, bold=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = FONT_BOLD if bold else FONT_BODY
        c.border = THIN
        c.alignment = Alignment(vertical='center')
        if fmt: c.number_format = fmt
        if fill: c.fill = fill

    wb = openpyxl.Workbook()
    ds = date.isoformat()

    ws1 = wb.active
    ws1.title = "Items"
    ws1.freeze_panes = "A3"
    title = f"Valley Pawn — Sold Review  |  {ds}  |  Target {int(TARGET_MARGIN*100)}%  |  Flag<{int(FLAG_MARGIN*100)}%"
    ws1.merge_cells("A1:N1")
    t = ws1["A1"]; t.value = title
    t.fill = PatternFill("solid", fgColor="0D1B40")
    t.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 22

    hdrs1 = ["Store", "Ticket/Inv #", "Category", "Description", "Cost", "Sale Price",
              "Margin", "Margin $", "Days On Shelf", "Meets Target?", "Flag?",
              "Fair Value", "± Band", "Sale vs Fair"]
    for ci, h in enumerate(hdrs1, 1):
        _hdr(ws1, 2, ci, h)
    ws1.row_dimensions[2].height = 28

    for ri, r in enumerate(sorted(valued, key=lambda x: (x["store"], (x["margin"] if x["margin"] is not None else 0))), 3):
        if r["critical"]:
            rfill = FILL_CRIT
        elif r["flag"]:
            rfill = FILL_FLAG
        elif r["margin"] is not None and r["margin"] < TARGET_MARGIN:
            rfill = FILL_WARN
        else:
            rfill = FILL_GOOD
        _cell(ws1, ri, 1, r["store"])
        _cell(ws1, ri, 2, r["ticket"])
        _cell(ws1, ri, 3, r["category"])
        _cell(ws1, ri, 4, r["desc"])
        _cell(ws1, ri, 5, r["cost"], fmt='"$"#,##0.00', fill=rfill)
        _cell(ws1, ri, 6, r["price"], fmt='"$"#,##0.00', fill=rfill)
        _cell(ws1, ri, 7, r["margin"], fmt='0.0%', fill=rfill)
        _cell(ws1, ri, 8, r["margin_dollars"], fmt='"$"#,##0.00')
        _cell(ws1, ri, 9, r.get("days_on_shelf"))
        _cell(ws1, ri, 10, "YES" if r["meets"] else "NO")
        _cell(ws1, ri, 11, "⛔ CRITICAL" if r["critical"] else ("🚨" if r["flag"] else ""))
        # Fair Value v2 (BLEND_V2): the accurate "should have sold for", with
        # uncertainty. A disputed pair (internal vs eBay-net >30% apart) shows
        # its basis text so the disagreement is visible, not averaged away.
        fv = r.get("fair_value")
        _cell(ws1, ri, 12, fv, fmt='"$"#,##0.00',
              fill=(FILL_WARN if r.get("fair_disputed") else None))
        _cell(ws1, ri, 13, r.get("fair_band"), fmt='"$"#,##0.00')
        ratio = (r["price"] / fv) if fv else None
        _cell(ws1, ri, 14, ratio, fmt='0%',
              fill=(FILL_FLAG if (ratio is not None and ratio < 0.70) else None))

    for ci, w in enumerate([7, 16, 20, 46, 10, 11, 9, 11, 13, 12, 12, 11, 9, 11], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:F1")
    t2 = ws2["A1"]; t2.value = f"Valley Pawn — Sold Review Summary  |  {ds}"
    t2.fill = PatternFill("solid", fgColor="0D1B40")
    t2.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    t2.alignment = Alignment(horizontal="center", vertical="center")

    hdrs2 = ["Store", "Items Sold", "Avg Margin", "Meets Target", "Below Threshold Flags", "Critical (below cost)"]
    for ci, h in enumerate(hdrs2, 1):
        _hdr(ws2, 2, ci, h)
    stores = sorted(set(r["store"] for r in valued))
    for ri, store in enumerate(stores, 3):
        si = [r for r in valued if r["store"] == store]
        mm = [r["margin"] for r in si if r["margin"] is not None]
        avg_m = statistics.mean(mm) if mm else None
        _cell(ws2, ri, 1, store, bold=True)
        _cell(ws2, ri, 2, len(si))
        _cell(ws2, ri, 3, avg_m, fmt='0.0%', fill=(FILL_FLAG if avg_m is not None and avg_m < FLAG_MARGIN else None))
        _cell(ws2, ri, 4, sum(1 for r in si if r["meets"]))
        _cell(ws2, ri, 5, sum(1 for r in si if r["flag"]), fill=(FILL_FLAG if any(r["flag"] for r in si) else None))
        _cell(ws2, ri, 6, sum(1 for r in si if r["critical"]), fill=(FILL_CRIT if any(r["critical"] for r in si) else None))
    for ci, w in enumerate([12, 12, 12, 14, 20, 20], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    flags_list = sorted([r for r in valued if r["flag"]],
                         key=lambda r: (r["margin"] if r["margin"] is not None else 0.0))
    if flags_list:
        ws3 = wb.create_sheet("Flags")
        hdrs3 = ["Store", "Ticket/Inv #", "Category", "Description", "Cost", "Sale Price",
                 "Margin", "Days On Shelf", "Critical?"]
        for ci, h in enumerate(hdrs3, 1):
            _hdr(ws3, 1, ci, h)
        for ri, r in enumerate(flags_list, 2):
            _cell(ws3, ri, 1, r["store"])
            _cell(ws3, ri, 2, r["ticket"])
            _cell(ws3, ri, 3, r["category"])
            _cell(ws3, ri, 4, r["desc"], fill=(FILL_CRIT if r["critical"] else FILL_FLAG))
            _cell(ws3, ri, 5, r["cost"], fmt='"$"#,##0.00')
            _cell(ws3, ri, 6, r["price"], fmt='"$"#,##0.00', fill=(FILL_CRIT if r["critical"] else FILL_FLAG))
            _cell(ws3, ri, 7, r["margin"], fmt='0.0%', fill=(FILL_CRIT if r["critical"] else FILL_FLAG))
            _cell(ws3, ri, 8, r.get("days_on_shelf"))
            _cell(ws3, ri, 9, "⛔ YES" if r["critical"] else "")
        for ci, w in enumerate([7, 16, 20, 46, 10, 11, 9, 13, 10], 1):
            ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(path)
    return True


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    date = resolve_date(sys.argv[1] if len(sys.argv) > 1 else None)
    date_str = date.isoformat()
    print(f"=== Sold Review  {date_str} ===")
    print(f"    Target margin: {int(TARGET_MARGIN*100)}%  |  Flag threshold: {int(FLAG_MARGIN*100)}%")

    os.makedirs(DAILY_DIR, exist_ok=True)
    summary_path = os.path.join(DAILY_DIR, f"{date_str}_sold_review_summary.json")
    xlsx_path = os.path.join(DAILY_DIR, f"{date_str}_sold_review.xlsx")

    open_today = open_stores_for(date)
    if not open_today:
        msg = f"{date_str} — no stores open (Sunday), correct no-op"
        print(f"INFO: {msg}")
        summary = {
            "date": date_str, "items": 0, "avg_margin": None, "flags": 0,
            "critical": 0, "stores": {}, "missing_stores": [],
            "excel_path": None, "slack_posted": False, "slack_skipped": True,
            "info": msg,
        }
        with open(summary_path, "w") as f:
            json.dump(summary, f, indent=2)
        print(f"JSON → {summary_path}")
        sys.exit(0)

    raw, missing = load_sold_for_date(date)
    if not raw:
        msg = (f"No jewelry-margin-sold data files found for {date_str} "
               f"(open stores expected: {', '.join(open_today)}; "
               f"missing: {', '.join(missing) or 'none — pull returned 0 rows for all open stores'})")
        print(f"INFO: {msg}")
        summary = {
            "date": date_str, "items": 0, "avg_margin": None, "flags": 0,
            "critical": 0, "stores": {}, "missing_stores": missing,
            "excel_path": None, "slack_posted": False, "slack_skipped": True,
            "info": msg,
        }
        with open(summary_path, "w") as f:
            json.dump(summary, f, indent=2)
        print(f"JSON → {summary_path}")
        sys.exit(0)

    valued = [compute_margin(r) for r in raw]
    stores_seen = sorted(set(r["store"] for r in valued))
    print(f"Loaded {len(valued)} sold items from {len(stores_seen)} store(s): {', '.join(stores_seen)}")
    valued = attach_market(valued, date)
    valued = attach_fair_value(valued, date)

    store_summaries = {}
    all_margins = []
    for store in stores_seen:
        si = [r for r in valued if r["store"] == store]
        mm = [r["margin"] for r in si if r["margin"] is not None]
        all_margins.extend(mm)
        store_summaries[store] = {
            "items": len(si),
            "avg_margin": round(statistics.mean(mm), 4) if mm else None,
            "flags": sum(1 for r in si if r["flag"]),
            "critical": sum(1 for r in si if r["critical"]),
        }

    flags_all = [r for r in valued if r["flag"]]
    critical_all = [r for r in valued if r["critical"]]

    summary = {
        "date": date_str,
        "items": len(valued),
        "avg_margin": round(statistics.mean(all_margins), 4) if all_margins else None,
        "flags": len(flags_all),
        "critical": len(critical_all),
        "stores": store_summaries,
        "missing_stores": missing,
        "excel_path": xlsx_path,
        "slack_posted": False,
        "slack_skipped": False,
        "slack_message": None,
        # Canary result. `market_feed_ok: false` means the eBay market benchmark is
        # degraded — the report is still VALID (margin grading is unaffected and the
        # benchmark falls back to internal comps), but it is no longer market-informed.
        # STEP 7 DMs Joshua on this so it can't rot silently for weeks.
        "market_feed_ok": _MARKET_HEALTH.get("ok", None),
        "market_feed_note": _MARKET_HEALTH.get("note", ""),
        "market_comps_cached": _MARKET_HEALTH.get("with_value", 0),
        # Fair Value v2 (BLEND_V2_PLAN) — coverage + SoldComps quota state so
        # STEP 7 can DM once if the API quota guard tripped or coverage cratered.
        "fair_value_coverage": _MARKET_HEALTH.get("fair_value_coverage", 0),
        "fair_value_disputes": _MARKET_HEALTH.get("fair_value_disputes", 0),
        "soldcomps_used_today": _soldcomps_usage(),
    }

    xl_ok = write_excel(valued, date, xlsx_path)
    if xl_ok:
        print(f"Excel → {xlsx_path}")
    summary["excel_path"] = xlsx_path if xl_ok else None

    slack_msg = build_slack_message(valued, date, missing)
    summary["slack_message"] = slack_msg
    if slack_msg is None:
        print(f"Slack post skipped — only {len(valued)} item(s) (min 3 required).")
        summary["slack_skipped"] = True
    else:
        tok = _get_slack_token()
        if not tok:
            print("WARNING: SLACK_BOT_TOKEN not found — Slack post skipped.")
            summary["slack_skipped"] = True
            summary["slack_error"] = "token_not_found"
        else:
            ok = slack_post(slack_msg, tok)
            summary["slack_posted"] = ok
            if ok:
                print(f"✓ Slack → #sold-review ({SLACK_CHANNEL})")
            else:
                print("✗ Slack post failed — check token/channel permissions", file=sys.stderr)
                summary["slack_error"] = "post_failed"

    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"JSON → {summary_path}")

    if flags_all:
        print(f"\nFlagged items — margin < {int(FLAG_MARGIN*100)}% ({len(flags_all)} total, {len(critical_all)} below cost):")
        for r in sorted(flags_all, key=lambda x: (x["margin"] if x["margin"] is not None else 0.0))[:20]:
            mstr = f"{r['margin']*100:.0f}%" if r["margin"] is not None else "—"
            print(f"  {r['store']:4} ${r['price']:>7,.0f} sale  ${r['cost']:>7,.0f} cost  margin {mstr:>5}  {str(r.get('desc',''))[:40]}")
    else:
        print("\nNo flags — all items sold at or above target threshold. ✓")

    print("\nDone.")


if __name__ == "__main__":
    main()
