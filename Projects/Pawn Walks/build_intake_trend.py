#!/usr/bin/env python3
"""
build_intake_trend.py  —  Valley Pawn intake-margin TREND workbook (additive, 2026-06-18)

Rolls up every daily intake-margin summary JSON into one cumulative spreadsheet so
margin / volume / overpay-flag trends can be read by STORE and COMPANY over time.

Stateless by design: it re-reads ALL daily/*_intake_margin_summary.json each run and
regenerates the workbook from scratch — so it's always consistent with the source data
and safe to re-run any time (nothing to corrupt, nothing to append-dedupe).

Output: daily/intake_margin_trend.xlsx
Sheets:
  • Company Trend        — one row per day: items, trusted, avg margin %, flags
  • By Store (long)      — one row per day×store
  • Avg Margin by Store  — pivot: dates down, stores across, avg margin %
  • Items by Store       — pivot: dates down, stores across, item counts
"""
import glob, json, os, datetime

BASE      = os.path.dirname(os.path.abspath(__file__))
DAILY_DIR = os.path.join(BASE, "daily")
OUT_PATH  = os.path.join(DAILY_DIR, "intake_margin_trend.xlsx")
TARGET_MARGIN = 0.50   # 50% target — used to color cells below target


def load_summaries() -> list[dict]:
    rows = []
    for f in sorted(glob.glob(os.path.join(DAILY_DIR, "*_intake_margin_summary.json"))):
        try:
            with open(f, encoding="utf-8") as fh:
                d = json.load(fh)
            if d.get("date"):
                rows.append(d)
        except Exception as e:
            print(f"WARNING: could not read {f}: {e}")
    rows.sort(key=lambda d: d["date"])
    return rows


def main():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — run: pip3 install --break-system-packages openpyxl")
        return 1

    summaries = load_summaries()
    if not summaries:
        print(f"No summary JSONs found in {DAILY_DIR}")
        return 1

    # discover the full set of stores seen across all days
    stores = sorted({s for d in summaries for s in (d.get("stores") or {})})

    wb = openpyxl.Workbook()
    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="305496")
    below_fill = PatternFill("solid", fgColor="F8CBAD")   # margin below 50% target
    meets_fill = PatternFill("solid", fgColor="C6EFCE")   # at/above target
    center = Alignment(horizontal="center")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
        ws.freeze_panes = "A2"

    def pct(v):
        return None if v is None else round(v * 100, 1)

    # ── Sheet 1: Company Trend ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Company Trend"
    ws.append(["Date", "Items", "Trusted", "Avg Margin %", "Overpay Flags"])
    for d in summaries:
        m = pct(d.get("avg_margin"))
        ws.append([d["date"], d.get("items", 0), d.get("trusted", 0), m, d.get("flags", 0)])
        mc = ws.cell(row=ws.max_row, column=4)
        if m is not None:
            mc.fill = meets_fill if m >= TARGET_MARGIN * 100 else below_fill
    style_header(ws, 5)
    ws.column_dimensions["A"].width = 13
    for col in "BCDE":
        ws.column_dimensions[col].width = 14

    # ── Sheet 2: By Store (long) ────────────────────────────────────────────
    ws2 = wb.create_sheet("By Store (long)")
    ws2.append(["Date", "Store", "Items", "Trusted", "Avg Margin %", "Overpay Flags"])
    for d in summaries:
        for store in sorted(d.get("stores") or {}):
            s = d["stores"][store]
            m = pct(s.get("avg_margin"))
            ws2.append([d["date"], store, s.get("total_items", 0),
                        s.get("trusted_items", 0), m, s.get("flags", 0)])
            mc = ws2.cell(row=ws2.max_row, column=5)
            if m is not None:
                mc.fill = meets_fill if m >= TARGET_MARGIN * 100 else below_fill
    style_header(ws2, 6)
    ws2.column_dimensions["A"].width = 13
    for col in "BCDEF":
        ws2.column_dimensions[col].width = 13

    # ── Sheet 3: Avg Margin by Store (pivot) ────────────────────────────────
    ws3 = wb.create_sheet("Avg Margin by Store")
    ws3.append(["Date"] + stores + ["Company"])
    for d in summaries:
        row = [d["date"]]
        for store in stores:
            s = (d.get("stores") or {}).get(store)
            row.append(pct(s.get("avg_margin")) if s else None)
        row.append(pct(d.get("avg_margin")))
        ws3.append(row)
        for ci in range(2, len(stores) + 3):
            v = ws3.cell(row=ws3.max_row, column=ci).value
            if isinstance(v, (int, float)):
                ws3.cell(row=ws3.max_row, column=ci).fill = (
                    meets_fill if v >= TARGET_MARGIN * 100 else below_fill)
    style_header(ws3, len(stores) + 2)
    ws3.column_dimensions["A"].width = 13
    for i in range(2, len(stores) + 3):
        ws3.column_dimensions[get_column_letter(i)].width = 11

    # ── Sheet 4: Items by Store (pivot) ─────────────────────────────────────
    ws4 = wb.create_sheet("Items by Store")
    ws4.append(["Date"] + stores + ["Company"])
    for d in summaries:
        row = [d["date"]]
        for store in stores:
            s = (d.get("stores") or {}).get(store)
            row.append(s.get("total_items", 0) if s else 0)
        row.append(d.get("items", 0))
        ws4.append(row)
    style_header(ws4, len(stores) + 2)
    ws4.column_dimensions["A"].width = 13
    for i in range(2, len(stores) + 3):
        ws4.column_dimensions[get_column_letter(i)].width = 11

    # ── Sheet 5: Avg Margin by Group ────────────────────────────────────────
    GROUPS = ['Gold', 'Silver', 'Guns', 'Everything Else']
    ws5 = wb.create_sheet("Avg Margin by Group")
    ws5.append(["Date"] + GROUPS)
    for d in summaries:
        g = d.get("groups") or {}
        ws5.append([d["date"]] + [pct(g[grp]["avg_margin"]) if g.get(grp) else None
                                  for grp in GROUPS])
    style_header(ws5, len(GROUPS) + 1)
    ws5.column_dimensions["A"].width = 13
    for i in range(2, len(GROUPS) + 2):
        ws5.column_dimensions[get_column_letter(i)].width = 16

    # ── Sheet 6: Cost Paid by Group ─────────────────────────────────────────
    ws6 = wb.create_sheet("Cost by Group")
    ws6.append(["Date"] + GROUPS)
    for d in summaries:
        g = d.get("groups") or {}
        ws6.append([d["date"]] + [(g[grp].get("cost_paid") if g.get(grp) else None)
                                  for grp in GROUPS])
    style_header(ws6, len(GROUPS) + 1)
    ws6.column_dimensions["A"].width = 13
    for i in range(2, len(GROUPS) + 2):
        ws6.column_dimensions[get_column_letter(i)].width = 16

    wb.save(OUT_PATH)
    print(f"Trend workbook → {OUT_PATH}")
    print(f"  {len(summaries)} day(s): {summaries[0]['date']} … {summaries[-1]['date']}")
    print(f"  stores seen: {', '.join(stores)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
