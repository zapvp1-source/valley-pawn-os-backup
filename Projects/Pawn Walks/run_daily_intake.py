#!/usr/bin/env python3
"""
Valley Pawn — Daily Intake Margin Pipeline
Reads yesterday's buys-from-public CSVs (and intake-detail when available) from
the Bravo Data Extraction pipeline output, runs the full T1/T2/T3 valuation
engine, flags items below the 30% margin threshold (overpay risk), posts a
per-store summary to #pawn-walks, and saves a daily Excel report.

Usage:
    python run_daily_intake.py [YYYY-MM-DD]

    Date defaults to yesterday. Pass a specific date for re-runs.

Output (daily/ subfolder next to this script):
    daily/{DATE}_intake_margin.xlsx
    daily/{DATE}_intake_margin_summary.json

Exit codes:
    0  — run complete (even if Slack token missing — check JSON)
    1  — critical import / data error

Notes:
  - Additive only. Does NOT modify intake_valuation_engine.py or any other file.
  - USE_TIER3=True uses tier3_cache.json — safe on re-runs; won't hammer APIs.
  - When intake-detail.csv becomes available (loans + buys), this script
    auto-prefers it over buys-from-public (same date/store filename pattern).
"""

from __future__ import annotations
import csv, glob, re, json, os, sys, datetime, statistics, collections
import urllib.request, urllib.parse

# ── Paths ─────────────────────────────────────────────────────────────────────
HERE          = os.path.dirname(os.path.abspath(__file__))
BRAVO_OUTPUT  = "/Users/joshuadavis/Documents/Claude/Projects/Bravo Data Extraction/output/"
DAILY_DIR     = os.path.join(HERE, "daily")

# ── Thresholds & config ───────────────────────────────────────────────────────
SLACK_CHANNEL = "C0B8WR95N31"   # #pawn-walks (private)
TARGET_MARGIN = 0.50            # 50% gross-margin target (matches engine design)
FLAG_MARGIN   = 0.30            # flag trusted items below this (overpay risk)
IMPLAUSIBLE_MARGIN = -1.0       # margin below this (cost > 2x est value) = bad/incomplete comp -> demote to 'review', never flag
USE_TIER3     = True            # cache-backed; safe on re-runs

# Bullion runs structurally thin — we pay a premium over spot to acquire it — so it
# gets its own lower margin expectation (set 2026-06-18 per VP direction).
BULLION_TARGET = 0.25           # bullion "meets" target (vs 50% general)
BULLION_FLAG   = 0.25           # flag bullion if margin < 25% (i.e. we paid > 75% of value)
# Premiums we pay/realize OVER spot, by type — the "premium side" of bullion.
SILVER_EAGLE_PREMIUM = 4.0      # $/ozt over spot — gov .999 1oz coins (Eagle/Maple)
SILVER_BAR_PREMIUM   = 2.0      # $/ozt over spot — generic .999 bars/rounds
GOLD_BULLION_PREMIUM = 0.025    # +2.5% over spot — gold bullion

# ── PM detection (mirrors engine's main() local; engine doesn't export it) ───
_PM_RE = re.compile(r'\b(GOLD|SILVER|COIN|BULLION|PLATINUM)\b', re.I)

# ── Import pure engine utilities (do NOT call engine's main()) ────────────────
sys.path.insert(0, HERE)
try:
    from intake_valuation_engine import (
        money, melt_value, comp_value, build_comp_index,
        GOLD_SPOT, SILVER_SPOT,
    )
except ImportError as e:
    print(f"CRITICAL: Cannot import intake_valuation_engine — {e}", file=sys.stderr)
    sys.exit(1)

# ── Tier 3 ────────────────────────────────────────────────────────────────────
_get_t3 = None
if USE_TIER3:
    try:
        from tier3_valuation import get_tier3_value as _get_t3
    except ImportError:
        print("WARNING: tier3_valuation.py not found — T3 disabled for this run.",
              file=sys.stderr)
        USE_TIER3 = False


# ── Date resolution ────────────────────────────────────────────────────────────
def resolve_date(arg: str | None) -> datetime.date:
    if arg:
        try:
            return datetime.date.fromisoformat(arg)
        except ValueError:
            print(f"Bad date '{arg}' — expected YYYY-MM-DD", file=sys.stderr)
            sys.exit(1)
    return datetime.date.today() - datetime.timedelta(days=1)


# ── Slack token resolution ─────────────────────────────────────────────────────
def _get_slack_token() -> str | None:
    """Try env var → Bravo config JSON → shell profile. Returns xoxb-... or None."""
    # 1. Environment variable (preferred, set in launchd plist or shell)
    tok = os.environ.get("SLACK_BOT_TOKEN", "").strip()
    if tok.startswith("xoxb-"):
        return tok

    # 2. JSON config file alongside the Bravo Data Extraction scripts
    for cpath in [
        os.path.join(BRAVO_OUTPUT, "..", "slack_config.json"),
        os.path.expanduser("~/Documents/Claude/Projects/Bravo Data Extraction/slack_config.json"),
        os.path.expanduser("~/.vp_slack_config.json"),
    ]:
        try:
            with open(os.path.normpath(cpath)) as f:
                d = json.load(f)
            t = (d.get("SLACK_BOT_TOKEN") or d.get("slack_bot_token") or "").strip()
            if t.startswith("xoxb-"):
                return t
        except Exception:
            pass

    # 3. Shell profile (export SLACK_BOT_TOKEN=xoxb-...)
    for profile in [
        os.path.expanduser("~/.bash_profile"),
        os.path.expanduser("~/.zshenv"),
        os.path.expanduser("~/.profile"),
    ]:
        try:
            txt = open(profile).read()
            m = re.search(r'(?:export\s+)?SLACK_BOT_TOKEN=["\']?(xoxb-[^\s"\']+)', txt)
            if m:
                return m.group(1).rstrip("\"'")
        except Exception:
            pass

    return None


# ── Load intake for a specific date ───────────────────────────────────────────
def load_intake_for_date(date: datetime.date) -> list[dict]:
    """
    Load intake rows for a specific date from the Bravo pipeline output.

    Filename patterns expected:
        {DATE}_{STORE}_intake-detail.csv      ← preferred when present
                                                (loans + buys; future Bravo report)
        {DATE}_{STORE}_buys-from-public.csv   ← fallback (buys only; exists now)

    When intake-detail files exist for the date they take precedence entirely.
    Mixed (some stores detail, some buys) is not expected; detail is all-or-nothing.
    """
    rows = []
    ds = date.isoformat()   # YYYY-MM-DD

    detail_files = glob.glob(os.path.join(BRAVO_OUTPUT, f"{ds}_*_intake-detail.csv"))
    buys_files   = glob.glob(os.path.join(BRAVO_OUTPUT, f"{ds}_*_buys-from-public.csv"))

    use_files  = detail_files if detail_files else buys_files
    file_type  = "intake-detail" if detail_files else "buys-from-public"
    store_pat  = r'_([A-Z]{3})_' + re.escape(file_type)

    def _parse_mdy(s):
        # Bravo writes dates as M/D/YYYY (non-padded, e.g. 6/17/2026)
        try:
            return datetime.datetime.strptime((s or '').strip(), '%m/%d/%Y').date()
        except ValueError:
            return None

    seen = set()   # exact-row dedupe across all stores (additive fix 2026-06-18)

    for f in use_files:
        m = re.search(store_pat, os.path.basename(f))
        store = m.group(1) if m else "UNK"
        try:
            with open(f, newline='', encoding='utf-8-sig') as fh:
                for row in csv.DictReader(fh):
                    if file_type == "intake-detail":
                        # intake-detail columns (saved Bravo report "Claude Pawn Walks"):
                        #   Ticket Number, Disposition, Disposition Date, Due Date,
                        #   Pull Date, Customer, Loan Amount, Age, MobilePawn, SMS
                        # NOTE: this report currently has NO Category/Full Description
                        # column, so cat/desc come back empty and valuation cannot run.
                        # (Restoring margins requires a cloned Bravo report with those
                        #  columns — tracked separately; this loader is forward-compatible.)
                        amt    = money(row.get('Amount') or row.get('Loan Amount'))
                        ticket = (row.get('TicketNumber') or row.get('Ticket Number') or '').strip()
                        cat    = (row.get('Category') or '').strip()
                        desc   = (row.get('FullDescription') or row.get('Full Description') or '').strip()
                        tkind  = (row.get('Ticket Kind') or '').strip().upper()
                        disp = (row.get('Disposition') or '').strip().upper()
                        has_disp_date = bool((row.get('Disposition Date') or '').strip())

                        # (1) SINGLE-DAY SCOPE (backstop) — only applied when the layout
                        #     actually exports a Disposition Date column. The "Claude Pawn
                        #     Walks" report already scopes to the prior day at the source
                        #     (Active Loans & Buys + Age=1); this guards against a stray
                        #     backlog-style layout (the 1,241-item bug seen 2026-06-15).
                        #     If the column is absent, trust the source scope.
                        if has_disp_date:
                            if _parse_mdy(row.get('Disposition Date')) != date:
                                continue
                        # (2) INTAKE ONLY — when a Disposition column is present, keep new
                        #     buys + new pawn loans only (exclude redemptions/payments).
                        if disp and disp not in ('BUY', 'ON LOAN'):
                            continue
                        # (3) EXACT-ROW DEDUPE — the grid scraper double-captures
                        #     identical rows (~5-10% inflation). Layout-agnostic: key off
                        #     the whole row so it works under any column set.
                        key = (store,) + tuple((v or '').strip() for v in row.values())
                        if key in seen:
                            continue
                        seen.add(key)
                    else:
                        # buys-from-public columns (legacy fallback):
                        # Ticket Number, Category, Full Description, Loan Amount
                        amt    = money(row.get('Loan Amount'))
                        ticket = (row.get('Ticket Number') or '').strip()
                        cat    = (row.get('Category') or '').strip()
                        desc   = (row.get('Full Description') or '').strip()
                        tkind  = 'BUY'

                    if amt is None or amt <= 0:
                        continue

                    # Quantity column (Bravo, added 2026-06-18) — used to value coin lots.
                    try:
                        qty = int(float(str(row.get('Quantity') or row.get('Qty')
                                           or row.get('QTY') or '').strip()))
                    except (ValueError, TypeError):
                        qty = None

                    rows.append({
                        'store':       store,
                        'ticket':      ticket,
                        'category':    cat,
                        'desc':        desc,
                        'cost':        amt,
                        'qty':         qty,
                        'ticket_kind': tkind,
                        'source_file': file_type,
                    })
        except Exception as e:
            print(f"WARNING: Could not read {f}: {e}", file=sys.stderr)

    return rows


# ── Coin / bullion valuation (T1-COIN / BULLION) ──────────────────────────────
# The melt parser only fires on a weight+fineness string (e.g. "8.1DWT SILV-925").
# Coins describe themselves by NAME, so they fell through to PM-NEEDS-WEIGHT ($0).
# This table recovers them. Junk/constitutional silver is valued at melt floor;
# .999 government/bar bullion gets the acquisition premium (the "premium side").
# Per-unit precious-metal content in troy ounces.  (pattern, ozt, metal, bullion, gov-1oz)
_COIN_SPECS = [
    (r'SILVER EAGLE|AMERICAN SILVER EAGLE|\bASE\b', 1.0,      'AG', True,  True),
    (r'SILVER MAPLE|MAPLE LEAF',                    1.0,      'AG', True,  True),
    (r'MORGAN',                                     0.7734,   'AG', False, False),
    (r'PEACE DOLLAR|PEACE SILVER',                  0.7734,   'AG', False, False),
    (r'WALKING LIBERTY HALF|FRANKLIN HALF|SILVER HALF|90%?\s*HALF|HALF DOLLAR', 0.36169, 'AG', False, False),
    (r'KENNEDY',                                    0.1479,   'AG', False, False),  # 40% default; ambiguous -> medium
    (r'SILVER QUARTER|WASHINGTON QUARTER|90%?\s*QUARTER', 0.18084, 'AG', False, False),
    (r'SILVER DIME|MERCURY DIME|ROOSEVELT DIME|90%?\s*DIME', 0.07234, 'AG', False, False),
    (r'SILVER DOLLAR',                              0.7734,   'AG', False, False),  # generic, late
    (r'GOLD EAGLE|AMERICAN GOLD EAGLE|KRUGERRAND|GOLD MAPLE', 1.0, 'AU', True, True),
    (r'DOUBLE EAGLE',                               0.9675,   'AU', False, False),
]
_OZ_RE  = re.compile(r'(\d+(?:\.\d+)?)\s*(?:OZ|OUNCE|TROY)', re.I)
_LOT_RE = re.compile(r'\b(QUANTITY|QTY|LOT|BAG|ASSORT|ROLL|GROUP)\b', re.I)


def pm_market_value(cat, desc, qty):
    """Value coins & bullion the DWT melt parser misses.
    Returns (value, source, confidence, is_bullion) or None if not coin/bullion.
    Lots without a count return (None, 'COIN-NEEDS-COUNT', 'none', ...) so they
    surface for a manual count instead of silently valuing at $0."""
    d = (str(desc) + ' ' + str(cat)).upper()
    q = qty if (isinstance(qty, (int, float)) and qty and qty > 0) else None
    is_lot = bool(_LOT_RE.search(d))

    per_oz = None; metal = None; bullion = False; conf = 'high'; ag_prem = SILVER_BAR_PREMIUM

    # generic .999 bar / round / ingot with an explicit oz weight
    m = _OZ_RE.search(d)
    if m and (('BAR' in d) or ('ROUND' in d) or ('INGOT' in d)):
        oz1 = float(m.group(1))
        if 'GOLD' in d:
            per_oz, metal, bullion = oz1, 'AU', True
        else:
            per_oz, metal, bullion, ag_prem = oz1, 'AG', True, SILVER_BAR_PREMIUM

    if per_oz is None:
        for pat, po, met, bull, gov in _COIN_SPECS:
            if re.search(pat, d):
                per_oz, metal, bullion = po, met, bull
                if bull and met == 'AG':
                    ag_prem = SILVER_EAGLE_PREMIUM if gov else SILVER_BAR_PREMIUM
                if pat == r'KENNEDY':
                    conf = 'medium'   # could be 90% '64 or 40% '65-70 — flag uncertainty
                break

    if per_oz is None:
        return None
    if is_lot and q is None:
        return (None, 'COIN-NEEDS-COUNT', 'none', bullion)

    oz = per_oz * (q if q else 1)
    if metal == 'AG':
        value = (SILVER_SPOT + ag_prem) * oz if bullion else SILVER_SPOT * oz
        src   = 'BULLION-AG' if bullion else 'COIN-AG'
    else:
        value = GOLD_SPOT * (1 + GOLD_BULLION_PREMIUM) * oz if bullion else GOLD_SPOT * oz
        src   = 'BULLION-AU' if bullion else 'COIN-AU'
    return (round(value, 2), src, conf, bullion)


# ── Group classifier: Gold / Silver / Guns / Everything Else ───────────────────
_GUN_RE = re.compile(r'\b(RIFLE|PISTOL|SHOTGUN|REVOLVER|FIREARM|HANDGUN|CARBINE|AR-?15|AK-?47)\b', re.I)
_CURR_RE = re.compile(r'CERTIFICATE|PAPER MONEY|FEDERAL RESERVE|BANK NOTE|\bNOTE\b|CURRENCY', re.I)
_KARAT_RE = re.compile(r'\b(10K|14K|18K|22K|24K)\b', re.I)


def group_of(r: dict) -> str:
    blob = (str(r.get('category', '')) + ' ' + str(r.get('desc', ''))).upper()
    src  = str(r.get('source', ''))
    if _GUN_RE.search(blob):
        return 'Guns'
    if _CURR_RE.search(blob):                      # paper currency is not metal
        return 'Everything Else'
    if src in ('COIN-AU', 'BULLION-AU') or _KARAT_RE.search(blob) or 'GOLD' in blob:
        return 'Gold'
    if src in ('COIN-AG', 'BULLION-AG') or re.search(r'SILVER|STERLING|SILV-?925|\b925\b', blob):
        return 'Silver'
    return 'Everything Else'


# ── Value one item (T1 → T2 → T3 routing) ─────────────────────────────────────
def value_item(it: dict, by_cat: dict, cat_prices: dict) -> dict:
    """
    Apply T1/T2/T3 routing to a single intake item.
    Returns enriched dict with: value, source, conf, t3_range_lo/hi,
    margin, flag, overpay, meets, trusted.
    """
    cat   = it.get('category', '')
    desc  = it.get('desc', '')
    cost  = it['cost']
    is_pm = bool(_PM_RE.search(cat))

    # T1-COIN / BULLION: coins & .999 bullion the DWT parser misses (named by type).
    pm = pm_market_value(cat, desc, it.get('qty'))
    is_bullion = False
    if pm is not None:
        val, src, conf, is_bullion = pm
        n = 0
    else:
        # T1: precious-metal melt value (weight-bearing jewelry / bars)
        mv, _mnote = melt_value(cat, desc)
        if mv is not None and mv > 0:
            val, src, n, conf = mv, 'MELT', 0, 'high'
        elif is_pm:
            # PM but no parseable weight — don't token-comp (garbage result)
            val, src, n, conf = None, 'PM-NEEDS-WEIGHT', 0, 'none'
        else:
            # T2: internal comp index (model-number token match → high; cat median → low)
            val, src, n, conf = comp_value(cat, desc, by_cat, cat_prices)

    # UNQUANTIFIED COIN/BULLION LOT GUARD: coins/bullion are valued per single unit.
    # Until Bravo's Quantity column feeds in, a piece priced way under what we paid is a
    # multi-piece lot (e.g. paid $1,270 for "Silver Eagle" = ~18 coins valued as one), NOT
    # an overpay. Surface it for a count instead of emitting a false metal overpay flag.
    # (Weight-based MELT jewelry is single-item, so it is intentionally NOT caught here.)
    if (str(src).startswith(('COIN-', 'BULLION')) and it.get('qty') is None
            and val is not None and cost > val * 1.4):
        val, src, conf, n, is_bullion = None, 'COIN-NEEDS-COUNT', 'none', 0, False

    # T3: external market (eBay sold-proxy). For RESALE MERCH we always pull eBay and
    # reconcile against our internal comp, letting the MORE CONSERVATIVE (lower) trusted
    # value win — never overstate. For guns, external stays a fallback only. PM skips T3.
    t3_rlo = t3_rhi = None
    if USE_TIER3 and _get_t3 and not is_pm and src not in ('COIN-NEEDS-COUNT',):
        try:
            t3 = _get_t3(desc, cat, cost)   # routes merch→eBay, guns→GunBroker-style sites
            t3v = t3.get('value')
            t3_trusted  = bool(t3v) and t3.get('confidence') in ('high', 'medium')
            cur_trusted = (val is not None) and conf in ('high', 'medium')
            # Conservative cross-check: the LOWER trusted value wins. Internal comp is
            # historical and most non-metal merch depreciates, so a stale-high comp must
            # never overstate. Applies to merch (eBay) AND guns (external gun values).
            if cur_trusted and t3_trusted:
                if t3v < val:                       # external lower → take it
                    val, conf, n = t3v, t3['confidence'], 0
                    src = t3['source'] + '|cons<comp'
                    t3_rlo, t3_rhi = t3.get('range_low'), t3.get('range_high')
                else:                               # our comp was already lower/equal
                    src = src + '|cons<ext'
            elif t3_trusted:                        # no trusted internal comp → use external
                val, src, conf, n = t3v, t3['source'], t3['confidence'], 0
                t3_rlo, t3_rhi = t3.get('range_low'), t3.get('range_high')
        except Exception:
            pass    # T3 failure is non-fatal; keep T1/T2 result

    trusted = conf in ('high', 'medium')
    margin  = ((val - cost) / val) if (val and val > 0) else None

    # SANITY GUARD: for COMP/eBay an estimate worth far less than we paid is usually a
    # bad/incomplete comp — demote to 'review' so it doesn't false-flag. BUT metal values
    # (melt/coin/bullion) are EXACT, and gemstones are worth nothing to us — so a big
    # negative margin there is a REAL overpay we must flag, not suppress. Exempt metals.
    _is_metal = src.startswith(('MELT', 'COIN-', 'BULLION'))
    if trusted and margin is not None and margin < IMPLAUSIBLE_MARGIN and not _is_metal:
        conf = 'review'
        trusted = False

    # Bullion is measured against its own lower target (we pay a premium to acquire it).
    flag_thr = BULLION_FLAG   if is_bullion else FLAG_MARGIN
    target   = BULLION_TARGET if is_bullion else TARGET_MARGIN
    # flag = trusted item below its overpay-risk threshold
    flag    = 1 if (trusted and margin is not None and margin < flag_thr) else 0
    # overpay = trusted item where we paid MORE than the estimate
    overpay = 1 if (trusted and val is not None and cost > val) else 0
    meets   = 1 if (trusted and margin is not None and margin >= target) else 0

    out = dict(it,
        value=val, source=src, comp_n=n, conf=conf,
        t3_range_lo=t3_rlo, t3_range_hi=t3_rhi,
        margin=margin, flag=flag, overpay=overpay, meets=meets,
        trusted=1 if trusted else 0, is_bullion=1 if is_bullion else 0,
        target_margin=target,
    )
    out['group'] = group_of(out)
    return out


# ── Build Slack message ────────────────────────────────────────────────────────
def build_slack_message(valued: list[dict], date: datetime.date) -> str | None:
    """Daily Pawn Walk Slack post. Splits intake into BUY and LOAN sections
    (Ticket Kind column), each per-store vs the 50% target, plus a by-category
    breakdown and the overpay-flag detail. Returns None on a no-activity (closed)
    day so the caller skips the post."""
    if len(valued) < 3:
        return None
    ds = date.isoformat()
    BAR = "━" * 22

    def _pct(m):
        return f"{m * 100:.0f}%" if m is not None else "—"

    def _section(title, subset):
        out = [BAR, f"*{title}*", BAR, "", "> ⚠️ Target ~50% margin. 🚩 = below 30% (overpay risk)."]
        stores = sorted(set(r.get('store', '') for r in subset if r.get('store')))
        ci = 0
        cf = 0
        cm = []
        for st in stores:
            si = [r for r in subset if r.get('store') == st]
            if not si:
                continue
            tr = [r for r in si if r.get('trusted')]
            mm = [r['margin'] for r in tr if r.get('margin') is not None]
            fl = sum(r.get('flag', 0) for r in tr)
            ci += len(si)
            cf += fl
            cm += mm
            avg = statistics.mean(mm) if mm else None
            stt = "—" if avg is None else ("✅" if avg >= TARGET_MARGIN else "🚨")
            fw = "flag" if fl == 1 else "flags"
            out.append(f"*{st}* — {len(si)} items | Avg margin {_pct(avg)} {stt} | {fl} {fw}")
        if ci == 0:
            out.append("_No intake this day._")
        else:
            cavg = statistics.mean(cm) if cm else None
            out.append("")
            out.append(f"*Total:* {ci} items | Avg margin {_pct(cavg)} | {cf} overpay flags")
        return out

    buys = [r for r in valued if (r.get('ticket_kind') or 'BUY') == 'BUY']
    loans = [r for r in valued if (r.get('ticket_kind') or '') == 'LOAN']

    lines = [f"📋 *Daily Pawn Walk — {ds}*", ""]
    lines += _section("BUY INTAKE — MARGIN vs 50% TARGET", buys)
    lines.append("")
    lines += _section("LOAN INTAKE — MARGIN vs 50% TARGET", loans)

    lines.append("")
    lines.append(BAR)
    lines.append("*BY CATEGORY (all intake)*")
    lines.append(BAR)
    _emoji = {'Gold': '🥇', 'Silver': '🥈', 'Guns': '🔫', 'Everything Else': '📦'}
    for g in ['Gold', 'Silver', 'Guns', 'Everything Else']:
        gi = [r for r in valued if r.get('group') == g]
        if not gi:
            continue
        gtr = [r for r in gi if r.get('trusted')]
        gm = [r['margin'] for r in gtr if r.get('margin') is not None]
        gavg = statistics.mean(gm) if gm else None
        gpaid = sum(r.get('cost', 0) for r in gtr)
        gest = sum((r.get('value') or 0) for r in gtr)
        gflags = sum(r.get('flag', 0) for r in gtr)
        gstatus = " 🚨" if (gavg is not None and gavg < TARGET_MARGIN) else ""
        em = _emoji.get(g, '📦')
        lines.append(f"{em} *{g}* — {len(gi)} items | {_pct(gavg)} margin{gstatus} | ${gpaid:,.0f} → ${gest:,.0f} est | {gflags} flags")

    flagged = [r for r in valued if r.get('flag')]
    if flagged:
        flagged.sort(key=lambda r: (r['margin'] if r.get('margin') is not None else 0.0))
        lines.append("")
        lines.append(BAR)
        lines.append(f"*⚑ OVERPAY FLAGS ({len(flagged)})* — store · kind · paid · margin · item")
        lines.append(BAR)
        for r in flagged[:12]:
            mstr = _pct(r.get('margin'))
            desc = (r.get('desc') or r.get('category') or 'item').strip()
            if len(desc) > 42:
                desc = desc[:41] + "…"
            kind = "L" if (r.get('ticket_kind') == 'LOAN') else "B"
            lines.append(f"• {r.get('store', '')} · {kind} · ${r.get('cost', 0):,.0f} · {mstr} · {desc}")
        if len(flagged) > 12:
            lines.append(f"…and {len(flagged) - 12} more — full detail in the spreadsheet")

    return "\n".join(lines)


# ── Post to Slack ──────────────────────────────────────────────────────────────
def slack_post(text: str, token: str) -> bool:
    """POST to Slack chat.postMessage. Returns True on API ok:true."""
    payload = json.dumps({
        "channel": SLACK_CHANNEL,
        "text":    text,
        "mrkdwn":  True,
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://slack.com/api/chat.postMessage",
        data=payload,
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type":  "application/json; charset=utf-8",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            body = json.load(resp)
            if not body.get("ok"):
                print(f"Slack API error: {body.get('error', '?')}", file=sys.stderr)
            return bool(body.get("ok"))
    except Exception as e:
        print(f"Slack request failed: {e}", file=sys.stderr)
        return False


# ── Excel report ───────────────────────────────────────────────────────────────
def write_excel(valued: list[dict], date: datetime.date, path: str) -> bool:
    """
    Write three-tab Excel report:
      Items    — full item-level detail with colour coding
      Summary  — per-store roll-up
      Flags    — trusted items below 30% margin (sorted by overpaid $)

    Returns True on success, False if openpyxl missing (non-fatal).
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("WARNING: openpyxl not installed — Excel skipped. "
              "Install: pip install openpyxl --break-system-packages", file=sys.stderr)
        return False

    # ── Styles ────────────────────────────────────────────────────────────────
    FILL_HDR      = PatternFill("solid", fgColor="1F3864")
    FILL_FLAG     = PatternFill("solid", fgColor="FFCCCC")    # red  : < 30%
    FILL_WARN     = PatternFill("solid", fgColor="FFF0CC")    # amber: 30-50%
    FILL_GOOD     = PatternFill("solid", fgColor="CCFFCC")    # green: ≥ 50%
    FILL_NOTRUST  = PatternFill("solid", fgColor="F5F5F5")    # grey : untrusted
    FONT_HDR      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    FONT_BODY     = Font(name="Calibri", size=10)
    FONT_BOLD     = Font(name="Calibri", bold=True, size=10)
    THIN          = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )

    def _hdr(ws, row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = FILL_HDR; c.font = FONT_HDR; c.border = THIN
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _cell(ws, row, col, val, fmt=None, fill=None, bold=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = FONT_BOLD if bold else FONT_BODY
        c.border = THIN
        c.alignment = Alignment(vertical='center')
        if fmt:  c.number_format = fmt
        if fill: c.fill = fill

    wb   = openpyxl.Workbook()
    ds   = date.isoformat()

    # ── Tab 1: Items ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Items"
    ws1.freeze_panes = "A3"

    title = f"Valley Pawn — Daily Intake Margin  |  {ds}  |  T3={USE_TIER3}  |  Flag<{int(FLAG_MARGIN*100)}% (bullion<{int(BULLION_FLAG*100)}%)"
    ws1.merge_cells(f"A1:O1")
    t = ws1["A1"]
    t.value = title
    t.fill  = PatternFill("solid", fgColor="0D1B40")
    t.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 22

    hdrs1 = ["Store", "Ticket", "Category", "Description",
             "Cost Paid", "Est Value", "Value Source", "Confidence",
             "Margin", "Meets Target?", "Flag?", "Overpay",
             "T3 Range Lo", "T3 Range Hi", "Group"]
    for ci, h in enumerate(hdrs1, 1):
        _hdr(ws1, 2, ci, h)
    ws1.row_dimensions[2].height = 28

    _grp_order = {'Gold': 0, 'Silver': 1, 'Guns': 2, 'Everything Else': 3}
    for ri, r in enumerate(sorted(valued, key=lambda x: (
            _grp_order.get(x.get('group', 'Everything Else'), 9),
            x.get('store', ''), -(x.get('cost') or 0))), 3):
        margin  = r.get('margin')
        trusted = r.get('trusted', 0)
        is_flag = r.get('flag', 0)
        is_over = r.get('overpay', 0)

        if not trusted:
            rfill = FILL_NOTRUST
        elif is_flag or is_over:
            rfill = FILL_FLAG
        elif margin is not None and margin < TARGET_MARGIN:
            rfill = FILL_WARN
        else:
            rfill = FILL_GOOD

        _cell(ws1, ri, 1,  r.get('store', ''))
        _cell(ws1, ri, 2,  r.get('ticket', ''))
        _cell(ws1, ri, 3,  r.get('category', ''))
        _cell(ws1, ri, 4,  r.get('desc', ''))
        _cell(ws1, ri, 5,  r.get('cost'),        fmt='"$"#,##0.00', fill=rfill)
        _cell(ws1, ri, 6,  r.get('value'),       fmt='"$"#,##0.00', fill=rfill)
        _cell(ws1, ri, 7,  r.get('source', ''))
        _cell(ws1, ri, 8,  r.get('conf', ''))
        _cell(ws1, ri, 9,  margin,               fmt='0.0%',        fill=rfill)
        _cell(ws1, ri, 10, "YES" if r.get('meets') else ("NO" if trusted else "—"))
        _cell(ws1, ri, 11, "⚑" if is_flag else "")
        _cell(ws1, ri, 12, "⚑ OVERPAY" if is_over else "")
        _cell(ws1, ri, 13, r.get('t3_range_lo'), fmt='"$"#,##0.00')
        _cell(ws1, ri, 14, r.get('t3_range_hi'), fmt='"$"#,##0.00')
        _cell(ws1, ri, 15, r.get('group', ''))

    for ci, w in enumerate([7, 16, 22, 46, 10, 10, 26, 10, 9, 12, 8, 12, 10, 10, 16], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Tab 2: Summary ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]
    t2.value = f"Valley Pawn — Intake Margin Summary  |  {ds}"
    t2.fill  = PatternFill("solid", fgColor="0D1B40")
    t2.font  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    t2.alignment = Alignment(horizontal="center", vertical="center")

    hdrs2 = ["Store", "Total Items", "Trusted Items", "Avg Margin",
             "Meets 50%", "Below 30% Flags", "Overpays"]
    for ci, h in enumerate(hdrs2, 1):
        _hdr(ws2, 2, ci, h)

    stores = sorted(set(r.get('store', '') for r in valued if r.get('store')))
    for ri, store in enumerate(stores, 3):
        s_items   = [r for r in valued if r.get('store') == store]
        s_trusted = [r for r in s_items if r.get('trusted')]
        s_margins = [r['margin'] for r in s_trusted if r.get('margin') is not None]
        s_flags   = sum(r.get('flag', 0) for r in s_trusted)
        s_over    = sum(r.get('overpay', 0) for r in s_trusted)
        s_meets   = sum(r.get('meets', 0) for r in s_trusted)
        avg_m     = statistics.mean(s_margins) if s_margins else None

        _cell(ws2, ri, 1, store, bold=True)
        _cell(ws2, ri, 2, len(s_items))
        _cell(ws2, ri, 3, len(s_trusted))
        _cell(ws2, ri, 4, avg_m, fmt='0.0%',
              fill=(FILL_FLAG if avg_m is not None and avg_m < FLAG_MARGIN else None))
        _cell(ws2, ri, 5, s_meets)
        _cell(ws2, ri, 6, s_flags, fill=(FILL_FLAG if s_flags > 0 else None))
        _cell(ws2, ri, 7, s_over,  fill=(FILL_FLAG if s_over  > 0 else None))

    for ci, w in enumerate([12, 14, 14, 14, 12, 18, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── By Group (Gold / Silver / Guns / Everything Else) ─────────────────────
    grow = len(stores) + 4
    _cell(ws2, grow, 1, "BY GROUP", bold=True)
    grow += 1
    for ci, h in enumerate(["Group", "Items", "Trusted", "Avg Margin",
                            "Cost Paid", "Est Value (trusted)", "Flags"], 1):
        _hdr(ws2, grow, ci, h)
    for grw, g in enumerate(['Gold', 'Silver', 'Guns', 'Everything Else'], grow + 1):
        g_items = [r for r in valued if r.get('group') == g]
        if not g_items:
            continue
        g_tr = [r for r in g_items if r.get('trusted')]
        g_mg = [r['margin'] for r in g_tr if r.get('margin') is not None]
        _cell(ws2, grw, 1, g, bold=True)
        _cell(ws2, grw, 2, len(g_items))
        _cell(ws2, grw, 3, len(g_tr))
        _cell(ws2, grw, 4, (statistics.mean(g_mg) if g_mg else None), fmt='0.0%')
        _cell(ws2, grw, 5, sum(r.get('cost') or 0 for r in g_items), fmt='"$"#,##0.00')
        _cell(ws2, grw, 6, sum((r.get('value') or 0) for r in g_tr), fmt='"$"#,##0.00')
        _cell(ws2, grw, 7, sum(r.get('flag', 0) for r in g_tr))

    # ── Tab 3: Flags (trusted items below 30%, sorted by $ overpaid) ──────────
    flags_list = sorted(
        [r for r in valued if (r.get('flag') or r.get('overpay')) and r.get('trusted')],
        key=lambda r: (r.get('cost', 0) - (r.get('value') or r.get('cost', 0))),
        reverse=True,
    )
    if flags_list:
        ws3 = wb.create_sheet("Flags")
        hdrs3 = ["Store", "Ticket", "Category", "Description",
                 "Cost Paid", "Est Value", "Margin",
                 "Overpaid By", "Value Source", "Confidence"]
        for ci, h in enumerate(hdrs3, 1):
            _hdr(ws3, 1, ci, h)

        for ri, r in enumerate(flags_list, 2):
            v = r.get('value')
            overpaid_by = max(0.0, r.get('cost', 0) - (v or 0.0))
            _cell(ws3, ri, 1,  r.get('store', ''))
            _cell(ws3, ri, 2,  r.get('ticket', ''))
            _cell(ws3, ri, 3,  r.get('category', ''))
            _cell(ws3, ri, 4,  r.get('desc', ''),  fill=FILL_FLAG)
            _cell(ws3, ri, 5,  r.get('cost'),       fmt='"$"#,##0.00', fill=FILL_FLAG)
            _cell(ws3, ri, 6,  v,                   fmt='"$"#,##0.00')
            _cell(ws3, ri, 7,  r.get('margin'),     fmt='0.0%',        fill=FILL_FLAG)
            _cell(ws3, ri, 8,  overpaid_by if overpaid_by > 0 else None, fmt='"$"#,##0.00')
            _cell(ws3, ri, 9,  r.get('source', ''))
            _cell(ws3, ri, 10, r.get('conf', ''))

        for ci, w in enumerate([7, 16, 22, 46, 11, 11, 9, 11, 26, 10], 1):
            ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(path)
    return True


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    date     = resolve_date(sys.argv[1] if len(sys.argv) > 1 else None)
    date_str = date.isoformat()
    t3_label = f"T3={'ON' if USE_TIER3 else 'OFF'}"
    print(f"=== Daily Intake Margin  {date_str}  ({t3_label}) ===")
    print(f"    Spot: Au ${GOLD_SPOT}/ozt  Ag ${SILVER_SPOT}/ozt")
    print(f"    Flag threshold: margin < {int(FLAG_MARGIN*100)}%  (target {int(TARGET_MARGIN*100)}%)")

    os.makedirs(DAILY_DIR, exist_ok=True)

    # ── Load intake data ──────────────────────────────────────────────────────
    intake = load_intake_for_date(date)
    summary_path = os.path.join(DAILY_DIR, f"{date_str}_intake_margin_summary.json")
    xlsx_path    = os.path.join(DAILY_DIR, f"{date_str}_intake_margin.xlsx")

    if not intake:
        msg = f"No buys-from-public or intake-detail files found for {date_str}"
        print(f"INFO: {msg}")
        print("      This is normal on slow days / weekends with no buys.")
        summary = {
            "date": date_str, "items": 0, "trusted": 0, "flags": 0,
            "avg_margin": None, "stores": {}, "source_mix": {},
            "excel_path": None, "tier3_enabled": USE_TIER3,
            "gold_spot": GOLD_SPOT, "silver_spot": SILVER_SPOT,
            "slack_posted": False, "info": msg,
        }
        with open(summary_path, 'w') as f:
            json.dump(summary, f, indent=2)
        print(f"JSON → {summary_path}")
        sys.exit(0)    # not an error — no-activity day

    stores_seen = sorted(set(r['store'] for r in intake))
    print(f"Loaded {len(intake)} items from {len(stores_seen)} store(s): {', '.join(stores_seen)}")

    # ── Build comp index ──────────────────────────────────────────────────────
    print("Building comp index from historical SOLD items...")
    try:
        by_cat, cat_prices = build_comp_index()
    except Exception as e:
        print(f"WARNING: build_comp_index failed ({e}) — T2 comps will be empty.",
              file=sys.stderr)
        by_cat, cat_prices = {}, {}

    # ── Value each item ───────────────────────────────────────────────────────
    print(f"Valuing {len(intake)} items (T1/T2{'+ T3 (cached)' if USE_TIER3 else ''})...")
    valued = [value_item(it, by_cat, cat_prices) for it in intake]

    trusted_all = [r for r in valued  if r.get('trusted')]
    flags_all   = [r for r in trusted_all if r.get('flag') or r.get('overpay')]
    src_mix     = dict(collections.Counter(r['source'] for r in valued))

    print(f"  {len(valued)} items valued | {len(trusted_all)} trusted | "
          f"{len(flags_all)} flagged (<{int(FLAG_MARGIN*100)}%)")
    print(f"  Source mix: {src_mix}")

    # ── Per-store summary ─────────────────────────────────────────────────────
    store_summaries: dict[str, dict] = {}
    all_margins: list[float] = []
    for store in stores_seen:
        s_items   = [r for r in valued if r['store'] == store]
        s_trusted = [r for r in s_items if r.get('trusted')]
        s_margins = [r['margin'] for r in s_trusted if r.get('margin') is not None]
        s_flags   = sum(r.get('flag', 0) or r.get('overpay', 0) for r in s_trusted)
        all_margins.extend(s_margins)
        store_summaries[store] = {
            "total_items":   len(s_items),
            "trusted_items": len(s_trusted),
            "avg_margin":    round(statistics.mean(s_margins), 4) if s_margins else None,
            "flags":         s_flags,
        }

    # Per-group roll-up (Gold / Silver / Guns / Everything Else) for the trend workbook
    group_summaries: dict[str, dict] = {}
    for g in ('Gold', 'Silver', 'Guns', 'Everything Else'):
        g_items = [r for r in valued if r.get('group') == g]
        if not g_items:
            continue
        g_tr = [r for r in g_items if r.get('trusted')]
        g_mg = [r['margin'] for r in g_tr if r.get('margin') is not None]
        group_summaries[g] = {
            "total_items":   len(g_items),
            "trusted_items": len(g_tr),
            "avg_margin":    round(statistics.mean(g_mg), 4) if g_mg else None,
            "cost_paid":     round(sum(r.get('cost') or 0 for r in g_items), 2),
            "est_value":     round(sum((r.get('value') or 0) for r in g_tr), 2),
            "flags":         sum(r.get('flag', 0) for r in g_tr),
        }

    summary = {
        "date":           date_str,
        "items":          len(valued),
        "trusted":        len(trusted_all),
        "flags":          len(flags_all),
        "avg_margin":     round(statistics.mean(all_margins), 4) if all_margins else None,
        "stores":         store_summaries,
        "groups":         group_summaries,
        "source_mix":     src_mix,
        "excel_path":     xlsx_path,
        "tier3_enabled":  USE_TIER3,
        "gold_spot":      GOLD_SPOT,
        "silver_spot":    SILVER_SPOT,
        "slack_posted":   False,    # updated below if post succeeds
        "slack_skipped":  False,
        "slack_message":  None,     # formatted message saved for MCP-fallback posting
    }

    # ── Save Excel ────────────────────────────────────────────────────────────
    xl_ok = write_excel(valued, date, xlsx_path)
    if xl_ok:
        print(f"Excel → {xlsx_path}")
    summary["excel_path"] = xlsx_path if xl_ok else None

    # ── Slack ─────────────────────────────────────────────────────────────────
    slack_msg = build_slack_message(valued, date)
    summary["slack_message"] = slack_msg   # always saved — scheduled task posts via Slack MCP if needed
    if slack_msg is None:
        print(f"Slack post skipped — only {len(valued)} item(s) (min 3 required).")
        summary["slack_skipped"] = True
    else:
        tok = _get_slack_token()
        if not tok:
            print("WARNING: SLACK_BOT_TOKEN not found — Slack post skipped.")
            print("  Set SLACK_BOT_TOKEN env var (or place it in slack_config.json)")
            print("  and re-run to post. Summary JSON is ready for manual posting.")
            summary["slack_skipped"] = True
            summary["slack_error"] = "token_not_found"
        else:
            ok = slack_post(slack_msg, tok)
            summary["slack_posted"] = ok
            if ok:
                print(f"✓ Slack → #pawn-walks ({SLACK_CHANNEL})")
            else:
                print("✗ Slack post failed — check token/channel permissions",
                      file=sys.stderr)
                summary["slack_error"] = "post_failed"

    # ── Save JSON summary ─────────────────────────────────────────────────────
    with open(summary_path, 'w') as f:
        json.dump(summary, f, indent=2)
    print(f"JSON → {summary_path}")

    # ── Console flag list ─────────────────────────────────────────────────────
    if flags_all:
        print(f"\nFlagged items — trusted margin < {int(FLAG_MARGIN*100)}%"
              f" ({len(flags_all)} total):")
        for r in sorted(flags_all,
                        key=lambda x: x.get('cost', 0) - (x.get('value') or 0),
                        reverse=True)[:20]:
            mstr = f"{r['margin']*100:.0f}%" if r.get('margin') is not None else "—"
            print(f"  {r['store']:4} {str(r.get('ticket','')):<15} "
                  f"${r.get('cost',0):>6,.0f} paid  "
                  f"margin {mstr:>5}  [{r['source']:<20}]  "
                  f"{str(r.get('desc',''))[:40]}")
    else:
        print("\nNo flags — all trusted items at or above 30% margin. ✓")

    # Refresh the cumulative by-store / company TREND workbook (additive, non-fatal:
    # a failure here must never fail the daily report).
    try:
        import importlib.util
        _tp = os.path.join(os.path.dirname(os.path.abspath(__file__)), "build_intake_trend.py")
        _spec = importlib.util.spec_from_file_location("build_intake_trend", _tp)
        _mod = importlib.util.module_from_spec(_spec)
        _spec.loader.exec_module(_mod)
        _mod.main()
    except Exception as e:
        print(f"WARNING: trend workbook refresh failed (non-fatal): {e}", file=sys.stderr)

    print("\nDone.")


if __name__ == '__main__':
    main()
