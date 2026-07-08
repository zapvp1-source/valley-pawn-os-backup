#!/usr/bin/env python3
"""
Valley Pawn — Tier 3 Test Runner
Runs the Tier 3 external valuation on 50 low/no-confidence intake items and
produces an Excel report: Intake_Valuation_Tier3_Test_2026-06-09.xlsx

Input:  intake_valued_items.csv  (produced by intake_valuation_engine.py)
Output: Intake_Valuation_Tier3_Test_2026-06-09.xlsx

Item selection:
  - Source must be CAT or NONE (low / no confidence from T1+T2)
  - Skip PM categories already handled by MELT
  - Balance: ~15 firearms, ~35 general merch
  - For eBay items: prefer those with a model-number token (better chance of a hit)
  - Deduplicate by description prefix (no duplicate models in the sample)
  - Limit to 50 total to avoid hammering APIs
"""

import csv, re, os, sys, time, statistics, collections

BASE = os.path.dirname(os.path.abspath(__file__))
INPUT_CSV  = os.path.join(BASE, "intake_valued_items.csv")
OUTPUT_XLS = os.path.join(BASE, "Intake_Valuation_Tier3_Test_2026-06-09.xlsx")
LOG_FILE   = os.path.join(BASE, "tier3_test_run.log")

try:
    from tier3_valuation import get_tier3_value, _toks, _model_keys, GUN_CAT, PM_RE
except ImportError:
    print("ERROR: Could not import tier3_valuation. Run from the Pawn Walks directory.")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages")
    sys.exit(1)


TARGET_MARGIN = 0.50
TARGET_ITEMS  = 50
GUN_QUOTA     = 15  # how many firearm items to include


# ── Load T1/T2 valued items ──────────────────────────────────────────────────

def load_valued():
    rows = []
    with open(INPUT_CSV, newline='') as f:
        for r in csv.DictReader(f):
            rows.append(r)
    return rows


def money(x):
    x = (x or '').replace('$', '').replace(',', '').strip()
    try: return float(x)
    except: return None


# ── Select 50 Tier 3 candidates ──────────────────────────────────────────────

def select_candidates(rows):
    """
    Filter to low/no confidence, skip melt PM items, deduplicate,
    balance guns vs general merch, return up to TARGET_ITEMS.
    """
    guns, merch = [], []
    seen = set()

    for r in rows:
        src  = r.get('ValueSource', '')
        conf = r.get('Confidence', '')
        cat  = r.get('Category', '')
        desc = r.get('Description', '') or ''
        cost = money(r.get('Cost'))

        # Only process low/none confidence T1T2
        if src not in ('CAT', 'NONE') or conf not in ('low', 'none'):
            continue
        if cost is None or cost <= 0:
            continue
        # Skip precious metals already handled by MELT
        if PM_RE.search(cat) and not GUN_CAT.search(cat):
            continue

        # Deduplicate by first 35 chars of description
        key = re.sub(r'\s+', ' ', desc).upper()[:35]
        if key in seen:
            continue
        seen.add(key)

        rec = dict(r, cost_f=cost)

        if GUN_CAT.search(cat):
            guns.append(rec)
        else:
            merch.append(rec)

    # Sort by cost descending (test on most-capital-at-risk items first)
    guns.sort(key=lambda x: x['cost_f'], reverse=True)
    merch.sort(key=lambda x: x['cost_f'], reverse=True)

    # For merch: prefer items with model-number tokens (better eBay hit rate)
    merch_with_model = [r for r in merch if _model_keys(_toks(r.get('Description','')))]
    merch_other      = [r for r in merch if r not in merch_with_model]

    # Build balanced sample
    gun_sample   = guns[:GUN_QUOTA]
    merch_quota  = TARGET_ITEMS - len(gun_sample)
    merch_sample = (merch_with_model + merch_other)[:merch_quota]

    return gun_sample + merch_sample


# ── Excel styling helpers ────────────────────────────────────────────────────

FILL_HEADER  = PatternFill("solid", fgColor="1F3864")   # dark navy
FILL_GUN     = PatternFill("solid", fgColor="FFF2CC")   # pale gold
FILL_MERCH   = PatternFill("solid", fgColor="F0F8FF")   # alice blue
FILL_OVERPAY = PatternFill("solid", fgColor="FFCCCC")   # light red
FILL_GOOD    = PatternFill("solid", fgColor="CCFFCC")   # light green
FILL_WARN    = PatternFill("solid", fgColor="FFF0CC")   # pale amber

FONT_HEADER  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FONT_NORMAL  = Font(name="Calibri", size=10)
FONT_BOLD    = Font(name="Calibri", bold=True, size=10)

THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)


def _hdr(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = FILL_HEADER; c.font = FONT_HEADER
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = THIN
    return c


def _cell(ws, row, col, value, fmt=None, fill=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_BOLD if bold else FONT_NORMAL
    c.border = THIN
    c.alignment = Alignment(vertical='center', wrap_text=False)
    if fmt:   c.number_format = fmt
    if fill:  c.fill = fill
    return c


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    log_lines = []

    def log(msg):
        print(msg)
        log_lines.append(msg)

    rows      = load_valued()
    sample    = select_candidates(rows)
    log(f"\n{'='*70}")
    log(f"Valley Pawn — Tier 3 Test Run   {time.strftime('%Y-%m-%d %H:%M')}")
    log(f"{'='*70}")
    log(f"Total T1/T2 valued items:  {len(rows)}")
    log(f"Low/none confidence items: {sum(1 for r in rows if r.get('Confidence') in ('low','none'))}")
    log(f"Test sample selected:      {len(sample)}")
    guns_in_sample = sum(1 for r in sample if GUN_CAT.search(r.get('Category','')))
    log(f"  Firearms: {guns_in_sample}   General merch: {len(sample)-guns_in_sample}")
    log("")

    # ── Run Tier 3 on each item ──
    results = []
    conf_count = collections.Counter()
    src_count  = collections.Counter()
    hit_count  = 0
    total_cost_hit = 0.0
    total_val_hit  = 0.0

    log(f"{'#':>3} {'Cat':22} {'Paid':>7} {'T3Est':>8} {'Conf':7} {'Margin':>7}  Description / Source")
    log("-" * 100)

    for i, rec in enumerate(sample, 1):
        desc = rec.get('Description', '')
        cat  = rec.get('Category', '')
        cost = rec['cost_f']

        t3 = get_tier3_value(desc, cat, cost)
        time.sleep(0.35)   # rate-limit: ~3 req/sec

        val  = t3['value']
        conf = t3['confidence']
        src  = t3['source']
        rlo  = t3['range_low']
        rhi  = t3['range_high']

        conf_count[conf] += 1
        # Bucket the source into a clean tag
        src_tag = src.split(':')[0]
        src_count[src_tag] += 1

        # Compute margin
        margin = ((val - cost) / val) if (val and val > 0) else None
        is_high_conf = conf in ('high', 'medium')
        overpay = is_high_conf and val is not None and cost > val
        below_target = is_high_conf and margin is not None and margin < TARGET_MARGIN

        if val is not None:
            hit_count += 1
            total_cost_hit += cost
            total_val_hit  += val

        # Console row
        v_str = f"${val:>7,.0f}" if val else "        —"
        m_str = f"{margin*100:.0f}%" if margin is not None else "—"
        flag  = " ⚑OVERPAY" if overpay else (" <50%" if below_target else "")
        log(f"{i:>3} {cat[:22]:22} ${cost:>6,.0f} {v_str} {conf[:6]:6} {m_str:>7}  "
            f"{str(desc)[:38]}{flag}")
        if val:
            log(f"    src={src}  range=[${rlo:,.0f}–${rhi:,.0f}]")

        results.append({
            'store':       rec.get('Store',''),
            'ticket':      rec.get('Ticket',''),
            'category':    cat,
            'desc':        desc,
            'cost':        cost,
            't1t2_value':  money(rec.get('EstValue')),
            't1t2_source': rec.get('ValueSource',''),
            't1t2_conf':   rec.get('Confidence',''),
            't3_value':    val,
            't3_source':   src,
            't3_conf':     conf,
            't3_range_lo': rlo,
            't3_range_hi': rhi,
            'best_value':  val if val else money(rec.get('EstValue')),
            'best_conf':   conf if val else rec.get('Confidence',''),
            'margin':      margin,
            'overpay':     1 if overpay else 0,
            'below_target': 1 if below_target else 0,
            'is_gun':      1 if GUN_CAT.search(cat) else 0,
        })

    # ── Summary ──
    log("")
    log("=== RESULTS ===")
    log(f"T3 hit rate:  {hit_count}/{len(sample)}  ({hit_count/len(sample)*100:.0f}%)")
    log(f"Confidence:   {dict(conf_count)}")
    log(f"Source mix:   {dict(src_count)}")

    high_conf = [r for r in results if r['t3_conf'] in ('high','medium') and r['t3_value']]
    overpays  = [r for r in high_conf if r['overpay']]
    below_50  = [r for r in high_conf if r['below_target']]

    if total_val_hit > 0:
        blended_margin = (total_val_hit - total_cost_hit) / total_val_hit
        log(f"Blended margin (T3 hits): paid ${total_cost_hit:,.0f} vs est ${total_val_hit:,.0f} "
            f"=> {blended_margin*100:.1f}%")

    log(f"\nFlagged items (high/medium conf, T3):")
    log(f"  Outright overpays: {len(overpays)}")
    log(f"  Below 50% target:  {len(below_50)}")

    if overpays:
        log("\n  Top overpays:")
        for r in sorted(overpays, key=lambda x: x['cost']-x['t3_value'], reverse=True)[:6]:
            delta = r['cost'] - r['t3_value']
            log(f"    {r['store']:4} {r['ticket']:15} ${r['cost']:>6,.0f} paid  "
                f"vs ${r['t3_value']:>6,.0f} T3est  (overpaid ~${delta:,.0f})  {r['category'][:20]}")

    log(f"\nCache saved to: {os.path.join(BASE, 'tier3_cache.json')}")

    # ── Write Excel ──
    wb = openpyxl.Workbook()

    # ─ Sheet 1: Item-level results ─
    ws1 = wb.active
    ws1.title = "Tier3 Item Results"
    ws1.freeze_panes = "A3"

    headers = [
        "Store", "Ticket", "Category", "Description",
        "Cost Paid",
        "T1/T2 Value", "T1/T2 Source", "T1/T2 Conf",
        "T3 Value", "T3 Range Low", "T3 Range High",
        "T3 Source", "T3 Conf",
        "Best Value", "Best Conf",
        "Implied Margin", "Meets 50%?", "Overpay Flag",
    ]

    # Title row
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws1.cell(row=1, column=1,
        value="Valley Pawn — Tier 3 Valuation Test  |  50 Low-Confidence Items  |  2026-06-09")
    title_cell.fill   = PatternFill("solid", fgColor="0D1B40")
    title_cell.font   = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 22

    # Header row
    for col, h in enumerate(headers, 1):
        _hdr(ws1, 2, col, h)
    ws1.row_dimensions[2].height = 30

    # Data rows
    for row_i, r in enumerate(results, 3):
        is_gun  = r['is_gun']
        val     = r['t3_value']
        margin  = r['margin']
        overpay = r['overpay']

        row_fill = FILL_GUN if is_gun else FILL_MERCH
        if overpay:             row_fill = FILL_OVERPAY
        elif r['below_target']: row_fill = FILL_WARN
        elif val and margin and margin >= TARGET_MARGIN: row_fill = FILL_GOOD

        _cell(ws1, row_i, 1,  r['store'])
        _cell(ws1, row_i, 2,  r['ticket'])
        _cell(ws1, row_i, 3,  r['category'])
        _cell(ws1, row_i, 4,  r['desc'])
        _cell(ws1, row_i, 5,  r['cost'],        fmt='"$"#,##0.00', fill=row_fill)
        _cell(ws1, row_i, 6,  r['t1t2_value'],  fmt='"$"#,##0.00')
        _cell(ws1, row_i, 7,  r['t1t2_source'])
        _cell(ws1, row_i, 8,  r['t1t2_conf'])
        _cell(ws1, row_i, 9,  r['t3_value'],    fmt='"$"#,##0.00', fill=row_fill,
              bold=bool(val))
        _cell(ws1, row_i, 10, r['t3_range_lo'], fmt='"$"#,##0.00')
        _cell(ws1, row_i, 11, r['t3_range_hi'], fmt='"$"#,##0.00')
        _cell(ws1, row_i, 12, r['t3_source'])
        _cell(ws1, row_i, 13, r['t3_conf'])
        _cell(ws1, row_i, 14, r['best_value'],  fmt='"$"#,##0.00')
        _cell(ws1, row_i, 15, r['best_conf'])
        _cell(ws1, row_i, 16,
              margin, fmt='0.0%', fill=row_fill)
        _cell(ws1, row_i, 17, "YES" if (margin and margin >= TARGET_MARGIN) else
                               ("NO" if margin is not None else "—"))
        _cell(ws1, row_i, 18, "⚑ OVERPAY" if overpay else "")

    # Column widths
    widths = [7, 15, 24, 45, 11, 11, 14, 9, 11, 11, 11, 28, 9, 11, 9, 13, 9, 12]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ─ Sheet 2: Summary ─
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20

    summary_rows = [
        ("Tier 3 Test Run Summary", None),
        ("Date", time.strftime('%Y-%m-%d %H:%M')),
        (None, None),
        ("Total items tested", len(results)),
        ("  Firearms", sum(r['is_gun'] for r in results)),
        ("  General merch / other", sum(1-r['is_gun'] for r in results)),
        (None, None),
        ("T3 Hit rate (any value returned)", f"{hit_count}/{len(results)} ({hit_count/len(results)*100:.0f}%)"),
        ("  High confidence", conf_count.get('high', 0)),
        ("  Medium confidence", conf_count.get('medium', 0)),
        ("  None (failed / no match)", conf_count.get('none', 0)),
        (None, None),
        ("Source breakdown", None),
    ]
    for src_tag, cnt in src_count.most_common():
        summary_rows.append((f"  {src_tag}", cnt))
    summary_rows += [
        (None, None),
        ("Flagged (high/med conf only)", None),
        ("  Outright overpays (cost > T3 est)", len(overpays)),
        ("  Below 50% target margin",          len(below_50)),
        (None, None),
        ("eBay API",       "Browse (active), SOLD_HAIRCUT=0.88"),
        ("Gun search",     "DuckDuckGo → price extraction"),
        ("Cache TTL",      "7d general / 30d firearms"),
        ("Model matching", "Digit-bearing token ≥3 chars; brand-only NOT trusted"),
        ("Confidence HIGH", f"≥{3} model-matched comps"),
    ]

    for sr_i, (label, value) in enumerate(summary_rows, 1):
        if label is None:
            continue
        ws2.cell(row=sr_i, column=1, value=label).font = Font(name="Calibri", bold=(value is None), size=10)
        if value is not None:
            ws2.cell(row=sr_i, column=2, value=value).font = Font(name="Calibri", size=10)

    wb.save(OUTPUT_XLS)
    log(f"\nExcel saved → {OUTPUT_XLS}")

    # Save log
    with open(LOG_FILE, 'w') as lf:
        lf.write('\n'.join(log_lines))
    log(f"Log saved  → {LOG_FILE}")


if __name__ == '__main__':
    main()
