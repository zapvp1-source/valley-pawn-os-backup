"""
bonus_kpis_extract.py -- additive, standalone extractor for monthly-bonus-targets.

Does NOT modify store_kpis_compile.py. Duplicates its verified EOM-xlsx parsing
logic (see that file's 2026-07-02 comment: PSC = in-store Interest+Fees+Misc,
Net Revenue = PSC + Sales Revenue (Profit); verified to the penny against Bravo
Company Performance for all 5 stores) and emits a clean JSON keyed by store code
with exactly the 3 fields monthly-bonus-targets needs: Net Revenue MTD, Loan
Balance, Inventory Balance. Reads the same output/<ENDDATE>_<STORE>_end-of-month.xlsx
files the "end-of-month" pipeline cell already produces -- no new Bravo handler,
no new pipeline cell required.

Usage: python3 bonus_kpis_extract.py <ENDDATE YYYY-MM-DD>
Prints JSON to stdout: {"enddate": ..., "data": {"CUL": {"net_revenue": ..., "loan_balance": ..., "inventory_balance": ...}, ...}}
On missing/undersized files for any store, prints {"error": "missing", "missing": [...]} and exits 2.
"""
import openpyxl, sys, os, json

BASE = '/Users/joshuadavis/Documents/Claude/Projects/Bravo Data Extraction/output/'
ENDDATE = sys.argv[1] if len(sys.argv) > 1 else None
if not ENDDATE:
    print(json.dumps({"error": "usage", "message": "ENDDATE argument (YYYY-MM-DD) required"}))
    sys.exit(1)

STORES = ['CUL', 'HAR', 'LEX', 'ROA', 'WAY']


def N(v):
    if v is None:
        return None
    s = str(v).strip().replace(chr(36), '').replace(',', '')
    neg = s.startswith('(') and s.endswith(')')
    s = s.strip('()')
    try:
        return (-float(s) if neg else float(s))
    except Exception:
        return None


def load(store):
    p = BASE + ENDDATE + '_' + store + '_end-of-month.xlsx'
    if not os.path.exists(p) or os.path.getsize(p) < 500:
        return None
    ws = openpyxl.load_workbook(p, data_only=True).active
    MC = ws.max_column

    def nn(r):
        return [x for x in [ws.cell(r, c).value for c in range(1, MC + 1)] if x is not None]

    def find(lbl, col=1, after=0, exact=False):
        for r in range(after + 1, ws.max_row + 1):
            v = ws.cell(r, col).value
            if v is None:
                continue
            sv = str(v).strip()
            if (sv == lbl if exact else sv.startswith(lbl)):
                return r
        return 0

    L = nn(find('Ending Loan Base'))
    I = nn(find('Ending Inventory Base'))
    loan = N(L[2]) if len(L) > 2 else 0
    inv = N(I[2]) if len(I) > 2 else 0

    sub = nn(find('In-Store Subtotal'))
    isInt = (N(sub[4]) or 0) if len(sub) > 4 else 0
    isFee = (N(sub[5]) or 0) if len(sub) > 5 else 0
    isMisc = (N(sub[6]) or 0) if len(sub) > 6 else 0

    rev = nn(find('Sales Revenue (Profit)'))
    prof = (N(rev[-1]) or 0) if rev else 0

    # Verified formula (matches store_kpis_compile.py, verified to the penny
    # against Bravo Company Performance 2026-07-02): PSC = in-store Interest+Fees+Misc.
    # Net Revenue = PSC + Sales Revenue (Profit). This is the exact "Net Revenue MTD"
    # figure monthly-bonus-targets requires -- never a gross-sales substitute.
    psc = isInt + isFee + isMisc
    net = psc + prof

    return {
        'net_revenue': round(net, 2),
        'loan_balance': round(loan or 0, 2),
        'inventory_balance': round(inv or 0, 2),
    }


data = {}
missing = []
for s in STORES:
    d = load(s)
    if d is None:
        missing.append(s)
    else:
        data[s] = d

if missing:
    print(json.dumps({"error": "missing", "missing": missing, "enddate": ENDDATE}))
    sys.exit(2)

print(json.dumps({"enddate": ENDDATE, "data": data}))
