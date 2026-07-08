import openpyxl, sys, os, glob
BASE='/Users/joshuadavis/Documents/Claude/Projects/Bravo Data Extraction/output/'
ENDDATE=sys.argv[1] if len(sys.argv)>1 else None
if not ENDDATE:
    xs=sorted(glob.glob(BASE+'*_end-of-month.xlsx'))
    ENDDATE=os.path.basename(xs[-1]).split('_')[0] if xs else ''
STORES=['CUL','HAR','LEX','ROA','WAY']
FULL={'CUL':'Culpeper','HAR':'Harrisonburg','LEX':'Lexington','ROA':'Roanoke','WAY':'Waynesboro'}
def N(v):
    if v is None: return None
    s=str(v).strip().replace(chr(36),'').replace(',','')
    neg=s.startswith('(') and s.endswith(')'); s=s.strip('()')
    try: return (-float(s) if neg else float(s))
    except: return None
def load(s):
    p=BASE+ENDDATE+'_'+s+'_end-of-month.xlsx'
    if not os.path.exists(p) or os.path.getsize(p)<500: return None
    ws=openpyxl.load_workbook(p, data_only=True).active
    MC=ws.max_column
    def nn(r): return [x for x in [ws.cell(r,c).value for c in range(1,MC+1)] if x is not None]
    def find(lbl,col=1,after=0,exact=False):
        for r in range(after+1,ws.max_row+1):
            v=ws.cell(r,col).value
            if v is None: continue
            sv=str(v).strip()
            if (sv==lbl if exact else sv.startswith(lbl)): return r
        return 0
    L=nn(find('Ending Loan Base')); I=nn(find('Ending Inventory Base'))
    loan=N(L[2]) if len(L)>2 else 0; inv=N(I[2]) if len(I)>2 else 0
    sub=nn(find('In-Store Subtotal'))
    isInt=(N(sub[4]) or 0) if len(sub)>4 else 0; isFee=(N(sub[5]) or 0) if len(sub)>5 else 0; isMisc=(N(sub[6]) or 0) if len(sub)>6 else 0
    sa=find('Sales Activity',exact=True)
    tx=nn(find('Taxable Sales',after=sa,exact=True)); ntx=nn(find('Nontaxable Sales',after=sa,exact=True))
    taxT=(N(tx[-1]) or 0) if tx else 0; ntxT=(N(ntx[-1]) or 0) if ntx else 0
    rev=nn(find('Sales Revenue (Profit)')); prof=(N(rev[-1]) or 0) if rev else 0
    ref=nn(find('Refined',col=2)); scrap=abs(N(ref[-1])) if ref and N(ref[-1]) else 0
    mob=nn(find('Totals from')); mInt=(N(mob[3]) or 0) if len(mob)>3 else 0; mFee=(N(mob[4]) or 0) if len(mob)>4 else 0; mMisc=(N(mob[5]) or 0) if len(mob)>5 else 0
    conv=nn(find('MobilePawn Convenience Fees')); convV=(N(conv[-1]) or 0) if conv else 0
    layB=0
    for r in range(1,ws.max_row+1):
        if any(str(ws.cell(r,c).value or '').strip().startswith('Ending Balance') for c in range(2,MC+1)):
            for x in nn(r):
                xv=N(x)
                if xv is not None and abs(xv-round(xv))>0.001: layB=xv; break
            break
    # KPI-aligned 2026-07-02 (see monthly-analytics-report/parse_eom.py, verified to the penny vs Bravo Company Performance, all 5 stores):
    # PSC = in-store Interest+Fees+Misc. Net Revenue = PSC + Sales Revenue (Profit). Mobile int/fees/misc + conv fees are NOT in Bravo Net Revenue.
    psc=isInt+isFee+isMisc; net=psc+prof
    return {'Loan Balance':loan or 0,'Inventory Balance':inv or 0,'Total Assets':(loan or 0)+(inv or 0),'Retail Sales Total Amt':taxT+ntxT,'Pawn Service Charges':psc,'Scrap Sales':scrap,'Layaway Balance':layB,'Net Revenue MTD':net}
data={}; missing=[]
for s in STORES:
    d=load(s)
    if d is None: missing.append(s)
    else: data[s]=d
if missing:
    print('INCOMPLETE missing='+','.join(missing)); sys.exit(2)
mets=['Loan Balance','Inventory Balance','Total Assets','Retail Sales Total Amt','Pawn Service Charges','Scrap Sales','Layaway Balance','Net Revenue MTD']
def d2(x): return chr(36)+format(x,',.2f')
medal=['🥇','🥈','🥉','4th','5th']
avg={s:0 for s in STORES}; wins={s:0 for s in STORES}; catrank={}
for m in mets:
    order=sorted(STORES,key=lambda s:-data[s][m]); catrank[m]=order
    for i,s in enumerate(order): avg[s]+=i+1
    if data[order[0]][m]!=data[order[1]][m]: wins[order[0]]+=1
for s in STORES: avg[s]=avg[s]/8.0
overall=sorted(STORES,key=lambda s:avg[s])
w=overall[0]; sec=overall[1]; last=overall[4]
secwins=[FULL[s] for m in mets for s in [catrank[m][0]] if s==sec]
seccats=[m for m in mets if catrank[m][0]==sec]
narr='_'+FULL[w]+'_ led the month with '+str(wins[w])+' of 8 category wins, anchored by the top loan book ('+d2(data[w]['Loan Balance'])+') and inventory. _'+FULL[sec]+'_ pushed hardest on '+(' and '.join(seccats) if seccats else 'the sales floor')+' for 2nd. _'+FULL[last]+'_ finished 5th across the board — the focus for the week.'
m1=[]
m1.append('*Valley Pawn — Weekly Store Performance Rankings*')
m1.append('📊 Report Period: '+ENDDATE+' (month-to-date)')
m1.append('')
m1.append('*🏆 Overall Store Rankings:*')
for i,s in enumerate(overall): m1.append(medal[i]+' *'+FULL[s]+'* — Avg Rank '+format(avg[s],'.2f')+' | '+str(wins[s])+' category wins out of 8')
m1.append('')
m1.append('*💡 Quick Summary:*')
m1.append(narr)
m1.append('')
m1.append('Full ranked breakdown in thread 👇')
lbl={'Total Assets':'Total Assets (Inventory + Loan)'}
m2=['*📊 Full Category Rankings*','']
for m in mets:
    m2.append('*'+lbl.get(m,m)+'*')
    if m=='Scrap Sales' and all(abs(data[s][m])<0.005 for s in STORES):
        m2.append('All stores at '+chr(36)+'0.00 (no scrap activity for the period)'); m2.append(''); continue
    for i,s in enumerate(catrank[m]): m2.append(medal[i]+' '+FULL[s]+' — '+d2(data[s][m]))
    m2.append('')
tot=lambda m: sum(data[s][m] for s in STORES)
m2.append('*Company Totals*')
m2.append('Loan Balance: '+d2(tot('Loan Balance'))+' | Inventory Balance: '+d2(tot('Inventory Balance'))+' | Layaway Balance: '+d2(tot('Layaway Balance'))+' | Net Revenue MTD: '+d2(tot('Net Revenue MTD')))
open(BASE+ENDDATE+'_store_kpis_msg1.txt','w').write(chr(10).join(m1))
open(BASE+ENDDATE+'_store_kpis_msg2.txt','w').write(chr(10).join(m2))
print('OK enddate='+ENDDATE)
print('MSG1='+BASE+ENDDATE+'_store_kpis_msg1.txt')
print('MSG2='+BASE+ENDDATE+'_store_kpis_msg2.txt')
