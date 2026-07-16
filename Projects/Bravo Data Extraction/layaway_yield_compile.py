#!/usr/bin/env python3
"""
layaway_yield_compile.py  --  Layaway Yield % compiler (additive, net-new)
-----------------------------------------------------------------------------
Computes "Layaway Yield %" per store + company for Valley Pawn:

    Layaway Yield % = (Down Payments MTD + Payments MTD) / Layaway Balance

REV 2 (2026-07-15, same day as REV 1): EOM-ONLY DESIGN.

REV 1 pulled a separate "Layaway Deposits" Bravo report for the numerator.
That report hung on export repeatedly across a live 5-store run (6 hangs in
one session) even after the Continuous-Scrolling fix -- reliable per-store,
not reliable back-to-back. Investigation found that Bravo's own "End of
Month" export -- already pulled every Monday by weekly-store-kpis with ZERO
hangs recorded -- contains a full "Layaways" section with Down Payments MTD,
Payments MTD, AND Ending Balance all in one place. Verified byte-for-byte
against the REV 1 live pull for all 5 stores on 2026-07-14 (identical to the
cent). REV 2 drops the Layaway Deposits report entirely and reads
numerator + denominator from the SAME already-scheduled EOM export -- one
fewer live Bravo report pull, zero new hang surface, matches Rule #4
(reuse hardened infrastructure instead of adding a parallel one).

EXTRACTION METHOD (robust to per-store column drift -- confirmed the raw
column position of the "Layaways" block shifts by 1-2 columns between
stores, e.g. label at AE32 for LEX/ROA vs AF32 for CUL/HAR/WAY; see
reference-bravo-saved-reports-per-store in memory for why this matters):
  1. Find the row where a cell == 'Layaways' (the section header) --
     read that row for the columns where 'Layaway Balance' and
     'Layaway Deposits' group headers sit. The "Layaway Balance" column
     group span is [layaway_balance_col, layaway_deposits_col).
  2. Search rows below the anchor for the labels 'Down Payments',
     'Payments via In-Store Transactions', and 'Ending Balance...'
     (label cell can be in any column -- Bravo sometimes offsets it 1-2
     cols further right per store).
  3. For each target row, take the LAST non-None cell value within the
     "Layaway Balance" column span. This is robust to Bravo's inconsistent
     cell-merge behavior between rows (Down Payments/Ending Balance carry
     both a Qty and a $ cell in that span; Payments carries just one $
     cell at a different offset within the same span) -- confirmed correct
     for all 5 stores against known-good REV 1 values.

Down Payments and Payments come back NEGATIVE in the EOM sheet (they reduce
the balance owed) -- take abs().

Naming: labelled "Layaway Yield %" everywhere (never bare "Yield") to avoid
collision with the unrelated store-level "Yield" metric used in
monthly-bonus-targets.

ADDITIVE (Rule #4): net-new file, reads only the already-produced
end-of-month.xlsx from the hardened `end-of-month` pipeline cell. Modifies
no existing handler, saved report, or scheduled task. No live Bravo report
pull of its own at all as of REV 2.

USAGE:
    python3 layaway_yield_compile.py <ENDDATE_YYYY-MM-DD>

    <ENDDATE> is the MTD end date (typically yesterday). The script expects:
      output/<ENDDATE>_<STORE>_end-of-month.xlsx   (all 5 stores -- already
      produced weekly by weekly-store-kpis; this script never pulls its own)

OUTPUT:
    Prints "OK enddate=<ENDDATE>" and writes:
      output/<ENDDATE>_layaway_yield.json   -- machine-readable per-store + company
      output/<ENDDATE>_layaway_yield_table.txt -- preformatted table for Slack/Sheet
    On missing/undersized inputs for a store, that store is skipped and listed
    under "missing" -- never fabricate a number.
"""
import sys, os, glob, json
import openpyxl

BASE = '/Users/joshuadavis/Documents/Claude/Projects/Bravo Data Extraction/output/'
STORES = ['CUL', 'HAR', 'LEX', 'ROA', 'WAY']
FULL = {'CUL': 'Culpeper', 'HAR': 'Harrisonburg', 'LEX': 'Lexington', 'ROA': 'Roanoke', 'WAY': 'Waynesboro'}

ENDDATE = sys.argv[1] if len(sys.argv) > 1 else None
if not ENDDATE:
    xs = sorted(glob.glob(BASE + '*_end-of-month.xlsx'))
    ENDDATE = os.path.basename(xs[-1]).split('_')[0] if xs else ''


def money(v):
    """Parse a Bravo-formatted dollar value (str or numeric) to float."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('$', '').replace(',', '')
    if s == '' or s == '-':
        return 0.0
    neg = s.startswith('(') and s.endswith(')')
    s = s.strip('()')
    try:
        val = float(s)
        return -val if neg else val
    except ValueError:
        return None


def parse_eom_layaways(store):
    """Return (down_payments_mtd, payments_mtd, layaway_balance) all as
    positive floats, or None if the file/section is missing."""
    xlsx_path = BASE + ENDDATE + '_' + store + '_end-of-month.xlsx'
    if not os.path.exists(xlsx_path) or os.path.getsize(xlsx_path) < 500:
        return None
    ws = openpyxl.load_workbook(xlsx_path, data_only=True).active
    MR, MC = ws.max_row, ws.max_column

    layaways_row = lb_col = ld_col = None
    for r in range(1, MR + 1):
        row_has_anchor = False
        for c in range(1, MC + 1):
            v = ws.cell(r, c).value
            if v == 'Layaways':
                layaways_row = r
                row_has_anchor = True
            elif v == 'Layaway Balance' and layaways_row == r:
                lb_col = c
            elif v == 'Layaway Deposits' and layaways_row == r:
                ld_col = c
        if row_has_anchor and lb_col and ld_col:
            break
    if not (layaways_row and lb_col and ld_col):
        return None

    def last_val_in_span(row):
        if row is None:
            return None
        vals = [ws.cell(row, c).value for c in range(lb_col, ld_col)]
        nonnull = [v for v in vals if v is not None]
        return money(nonnull[-1]) if nonnull else None

    down_row = pay_row = end_row = None
    for r in range(layaways_row + 1, min(layaways_row + 25, MR + 1)):
        for c in range(1, MC + 1):
            v = ws.cell(r, c).value
            if v == 'Down Payments' and down_row is None:
                down_row = r
            elif v == 'Payments via In-Store Transactions' and pay_row is None:
                pay_row = r
            elif isinstance(v, str) and v.startswith('Ending Balance') and end_row is None:
                end_row = r

    down = last_val_in_span(down_row)
    pay = last_val_in_span(pay_row)
    bal = last_val_in_span(end_row)
    if down is None or pay is None or bal is None:
        return None
    return (abs(down), abs(pay), abs(bal))


def d2(x):
    return '$' + format(x, ',.2f')


def main():
    if not ENDDATE:
        print('ERROR no end-of-month.xlsx files found and no ENDDATE given')
        sys.exit(2)

    data = {}
    missing = []
    for s in STORES:
        parsed = parse_eom_layaways(s)
        if parsed is None:
            missing.append(s + ':eom-layaways')
            continue
        down, pay, bal = parsed
        mtd_collected = down + pay
        yield_pct = (mtd_collected / bal * 100.0) if bal else None
        data[s] = {
            'down_payments_mtd': down,
            'payments_mtd': pay,
            'collected_mtd': mtd_collected,
            'layaway_balance': bal,
            'layaway_yield_pct': yield_pct,
        }

    company_down = sum(d['down_payments_mtd'] for d in data.values())
    company_pay = sum(d['payments_mtd'] for d in data.values())
    company_collected = company_down + company_pay
    company_balance = sum(d['layaway_balance'] for d in data.values())
    company_yield = (company_collected / company_balance * 100.0) if company_balance else None

    result = {
        'enddate': ENDDATE,
        'source': 'end-of-month (EOM-only, REV 2, 2026-07-15)',
        'stores': data,
        'company': {
            'down_payments_mtd': company_down,
            'payments_mtd': company_pay,
            'collected_mtd': company_collected,
            'layaway_balance': company_balance,
            'layaway_yield_pct': company_yield,
        },
        'missing': missing,
    }

    with open(BASE + ENDDATE + '_layaway_yield.json', 'w') as f:
        json.dump(result, f, indent=2)

    lines = []
    lines.append('Store          Down Pmts MTD   Payments MTD   Collected MTD   Layaway Bal    Layaway Yield %')
    lines.append('-------------  --------------   ------------   -------------   -----------    ---------------')
    for s in STORES:
        if s not in data:
            lines.append('%-13s  MISSING' % FULL[s])
            continue
        d = data[s]
        lines.append('%-13s  %14s   %12s   %13s   %11s    %.2f%%' % (
            FULL[s], d2(d['down_payments_mtd']), d2(d['payments_mtd']),
            d2(d['collected_mtd']), d2(d['layaway_balance']), d['layaway_yield_pct']))
    lines.append('-------------  --------------   ------------   -------------   -----------    ---------------')
    lines.append('%-13s  %14s   %12s   %13s   %11s    %.2f%%' % (
        'Company', d2(company_down), d2(company_pay), d2(company_collected),
        d2(company_balance), company_yield))

    with open(BASE + ENDDATE + '_layaway_yield_table.txt', 'w') as f:
        f.write('\n'.join(lines) + '\n')

    if missing:
        print('PARTIAL enddate=' + ENDDATE + ' missing=' + ','.join(missing))
    else:
        print('OK enddate=' + ENDDATE)
    print('JSON=' + BASE + ENDDATE + '_layaway_yield.json')
    print('TABLE=' + BASE + ENDDATE + '_layaway_yield_table.txt')


if __name__ == '__main__':
    main()
