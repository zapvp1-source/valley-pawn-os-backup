#!/usr/bin/env python3
"""
scrap_trend_sheet.py — rebuild the rolling gold-scrap trend workbook.

Lands in the ONE trends home: Valley Pawn Drive / Trends /.
Per FILING_GUIDE.md, store-operations output belongs on the Valley Pawn SHARED
drive (never My Drive), and those folders are automation targets.

The workbook is rewritten in full every run, so it is always a complete
restatement of scrap_history.csv — there is no append logic that can drift.

Period = the month a bucket was POSTED (closed / sent out). Buckets are posted
the month AFTER the gold is collected, so each row also names the collection
month it represents.

Usage:  python3 scrap_trend_sheet.py
"""
import os
import csv as _csv
import scrap_rankings as S

TRENDS_DIR = ("/Users/joshuadavis/Library/CloudStorage/GoogleDrive-jdavis@fcfpawn.com"
              "/Shared drives/Valley Pawn Drive/Trends")
SHEET = os.path.join(TRENDS_DIR, "Valley Pawn - Gold Scrap Trend.xlsx")

MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]


def prev_month(y, m):
    return (y - 1, 12) if m == 1 else (y, m - 1)


def build_sheet():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    per = S.load_history()
    periods = sorted(per)
    if not periods:
        print("no history; nothing written")
        return

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2A78D6")
    bold = Font(bold=True)

    def style_header(ws, ncols):
        for i in range(1, ncols + 1):
            c = ws.cell(row=1, column=i)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"

    # ---------- Sheet 1: Monthly Trend ----------
    ws = wb.active
    ws.title = "Monthly Trend"
    cols = (["Posted Month", "Gold Collected"]
            + [S.STORE_NAMES[s] for s in S.STORES]
            + ["Company Total", "vs Same Month Last Yr"])
    ws.append(cols)
    style_header(ws, len(cols))

    for p in periods:
        y, m = int(p[:4]), int(p[5:7])
        cy, cm = prev_month(y, m)
        tot = sum(per[p].get(s, 0) for s in S.STORES)
        prior = per.get("%d-%02d" % (y - 1, m), {})
        ptot = sum(prior.get(s, 0) for s in S.STORES)
        yoy = ((tot - ptot) / ptot) if ptot else None
        ws.append([p, "%s %d" % (MONTH_NAMES[cm], cy)]
                  + [round(per[p].get(s, 0), 1) or None for s in S.STORES]
                  + [round(tot, 1), yoy])

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=len(cols)).number_format = "0%"
        ws.cell(row=r, column=len(cols) - 1).font = bold
    for i, w in enumerate([14, 18] + [13] * len(S.STORES) + [14, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---------- Sheet 2: Year over Year by store ----------
    ws2 = wb.create_sheet("Year over Year")
    years = sorted({int(p[:4]) for p in periods})
    ws2.append(["Store", "Year"] + MONTH_NAMES[1:] + ["Total"])
    style_header(ws2, 15)
    for s in S.STORES:
        for y in years:
            vals = [per.get("%d-%02d" % (y, m), {}).get(s, 0) for m in range(1, 13)]
            ws2.append([S.STORE_NAMES[s], y]
                       + [round(v, 1) or None for v in vals] + [round(sum(vals), 1)])
    for y in years:
        vals = [sum(per.get("%d-%02d" % (y, m), {}).get(s, 0) for s in S.STORES)
                for m in range(1, 13)]
        ws2.append(["COMPANY", y] + [round(v, 1) or None for v in vals]
                   + [round(sum(vals), 1)])
        for c in range(1, 16):
            ws2.cell(row=ws2.max_row, column=c).font = bold
    ws2.column_dimensions["A"].width = 16
    for i in range(2, 16):
        ws2.column_dimensions[get_column_letter(i)].width = 11

    # ---------- Sheet 3: Bucket Detail (audit trail) ----------
    ws3 = wb.create_sheet("Bucket Detail")
    ws3.append(["Store", "Posted Month", "Bucket Name", "Weight (dwt)", "Date Source"])
    style_header(ws3, 5)
    if os.path.exists(S.HISTORY):
        for r in _csv.DictReader(open(S.HISTORY)):
            ws3.append([r["store"], r["period"], r["bucket"],
                        float(r["dwt"]) if r["dwt"] else None, r["date_source"]])
    for col, w in zip("ABCDE", [8, 14, 44, 13, 18]):
        ws3.column_dimensions[col].width = w

    os.makedirs(TRENDS_DIR, exist_ok=True)
    wb.save(SHEET)
    print("wrote %s" % SHEET)
    print("  %d posted months, %d stores, %d..%d"
          % (len(periods), len(S.STORES), years[0], years[-1]))


if __name__ == "__main__":
    build_sheet()
