#!/usr/bin/env python3
"""Build Valley Pawn Store Rankings xlsx from Company KPI data."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Brand colors
PURPLE = "2D1A5E"
BLUE = "0099DD"
CORAL = "F58C8A"
LIGHT_BLUE = "3DB8E8"
GOLD = "FFD700"
SILVER = "C0C0C0"
BRONZE = "CD7F32"
WHITE = "FFFFFF"

DATA_FILE = "/sessions/blissful-lucid-hawking/mnt/Scheduled/_shared-bravo-data/2026-05-04/company_kpi.json"
OUT_FILE = "/sessions/blissful-lucid-hawking/mnt/Scheduled/Valley_Pawn_Store_Rankings_May2026.xlsx"

with open(DATA_FILE) as f:
    data = json.load(f)

# Map store codes to display names
DISPLAY = {
    "CUL": "Culpeper",
    "HAR": "Harrisonburg",
    "LEX": "Lexington",
    "ROA": "Roanoke",
    "WAY": "Waynesboro",
}

# Categories — for each, the metric key in stores dict, sorted high to low
CATEGORIES = [
    ("Loan Balance", "loan_balance"),
    ("Inventory Balance", "inventory_balance"),
    ("Total Assets", "total_assets"),
    ("Retail Sales Total Amt", "retail_sales_total_amt"),
    ("Pawn Service Charges", "pawn_service_charges"),
    ("Scrap Sales", "scrap_sales"),
    ("Layaway Balance", "layaway_balance"),
    ("Net Revenue MTD", "net_revenue_mtd"),
]

# Compute rankings
stores = data["stores"]
rankings = {}  # category -> [(code, value), ...] sorted high to low
for label, key in CATEGORIES:
    pairs = [(code, sd[key]) for code, sd in sorted(stores.items(), key=lambda kv: -kv[1][key])]
    rankings[label] = pairs

# Per-store stats: avg rank, #1 finishes
per_store_stats = {code: {"ranks": [], "wins": 0} for code in stores}
for label, pairs in rankings.items():
    for rank, (code, _) in enumerate(pairs, start=1):
        per_store_stats[code]["ranks"].append(rank)
        if rank == 1:
            per_store_stats[code]["wins"] += 1

for code in per_store_stats:
    rs = per_store_stats[code]["ranks"]
    per_store_stats[code]["avg_rank"] = sum(rs) / len(rs) if rs else 0

# Sort overall — by avg rank ascending (lower = better), wins as tiebreaker desc
overall = sorted(
    per_store_stats.items(),
    key=lambda kv: (kv[1]["avg_rank"], -kv[1]["wins"]),
)

# === Build workbook ===
wb = Workbook()
ws = wb.active
ws.title = "Store Rankings"

thin = Side(border_style="thin", color="888888")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title
ws["A1"] = "VALLEY PAWN — Store Performance Rankings"
ws["A1"].font = Font(name="Arial", bold=True, size=18, color=WHITE)
ws["A1"].fill = PatternFill("solid", fgColor=PURPLE)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells("A1:G1")
ws.row_dimensions[1].height = 32

# Subtitle
period = f"Report Period: {data['report_period']['start']} → {data['report_period']['end']}"
ws["A2"] = period
ws["A2"].font = Font(name="Arial", italic=True, size=11, color="555555")
ws["A2"].alignment = Alignment(horizontal="center")
ws.merge_cells("A2:G2")
ws.row_dimensions[2].height = 20

# Header row at row 4
headers = ["Category", "#1", "#2", "#3", "#4", "#5", "Company Total"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=4, column=col, value=h)
    c.font = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_all
ws.row_dimensions[4].height = 22

# Column widths
ws.column_dimensions["A"].width = 28
for col in range(2, 8):
    ws.column_dimensions[get_column_letter(col)].width = 17

# Data rows: each category gets two rows (store names, then $ values)
row = 5
RANK_FILLS = {1: GOLD, 2: SILVER, 3: BRONZE}
for label, key in CATEGORIES:
    pairs = rankings[label]
    # Row 1: Category label + store names by rank
    ws.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="EEEEEE")
    ws.cell(row=row, column=1).border = border_all
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for i, (code, _) in enumerate(pairs, start=1):
        c = ws.cell(row=row, column=i + 1, value=DISPLAY[code])
        c.font = Font(name="Arial", bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_all
        if i in RANK_FILLS:
            c.fill = PatternFill("solid", fgColor=RANK_FILLS[i])
    # Company Total cell on this row stays empty/merged
    ws.cell(row=row, column=7, value="").border = border_all

    # Row 2: blank label + $ values + grand total
    row += 1
    ws.cell(row=row, column=1, value="").border = border_all
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="EEEEEE")
    for i, (code, val) in enumerate(pairs, start=1):
        c = ws.cell(row=row, column=i + 1, value=val)
        c.number_format = '"$"#,##0.00'
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_all
        if i in RANK_FILLS:
            c.fill = PatternFill("solid", fgColor=RANK_FILLS[i])
    grand = data["grand_total"][key]
    c = ws.cell(row=row, column=7, value=grand)
    c.number_format = '"$"#,##0.00'
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=PURPLE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_all
    row += 1

# Overall Rankings section
row += 1
ws.cell(row=row, column=1, value="🏆 Overall Store Rankings").font = Font(name="Arial", bold=True, size=14, color=WHITE)
ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=CORAL)
ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
ws.row_dimensions[row].height = 26
row += 1

ovh = ["Rank", "Store", "Avg Rank", "#1 Finishes", "", "", ""]
for col, h in enumerate(ovh, start=1):
    c = ws.cell(row=row, column=col, value=h)
    c.font = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if h:
        c.border = border_all
row += 1

medals = {1: "🥇", 2: "🥈", 3: "🥉", 4: "4th", 5: "5th"}
for i, (code, stat) in enumerate(overall, start=1):
    ws.cell(row=row, column=1, value=medals[i]).font = Font(name="Arial", bold=True, size=12)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=1).border = border_all
    ws.cell(row=row, column=2, value=DISPLAY[code]).font = Font(name="Arial", bold=True, size=11)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=row, column=2).border = border_all
    c = ws.cell(row=row, column=3, value=round(stat["avg_rank"], 2))
    c.number_format = "0.00"
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_all
    c.font = Font(name="Arial", size=11)
    c = ws.cell(row=row, column=4, value=stat["wins"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border_all
    c.font = Font(name="Arial", size=11)
    if i in RANK_FILLS:
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=RANK_FILLS[i])
    row += 1

wb.save(OUT_FILE)
print(f"Saved: {OUT_FILE}")

# Print rankings summary for Slack
print("\n=== OVERALL RANKINGS ===")
for i, (code, stat) in enumerate(overall, start=1):
    print(f"{i}. {DISPLAY[code]:14s}  avg={stat['avg_rank']:.2f}  wins={stat['wins']}")

print("\n=== CATEGORY RANKINGS ===")
for label, key in CATEGORIES:
    print(f"\n{label}:")
    for rank, (code, val) in enumerate(rankings[label], start=1):
        medal = ["🥇", "🥈", "🥉", "4th", "5th"][rank - 1]
        print(f"  {medal} {DISPLAY[code]:14s}  ${val:>12,.2f}")
