import openpyxl, json, sys, glob, os
def money(v):
    if v is None: return 0.0
    s=str(v).strip().replace(chr(36),'').replace(',','').replace(' ','')
    neg=s.startswith('(') and s.endswith(')')
    if neg: s=s[1:-1]
    if not s or s=='-': return 0.0
    try: f=float(s)
    except: return 0.0
    return -f if neg else f
def parse_store_xlsx(path):
    ws=openpyxl.load_workbook(str(path), data_only=True).active
    MC=ws.max_column
    def nn(r): return [x for x in [ws.cell(r,c).value for c in range(1,MC+1)] if x is not None]
    def find(lbl,col=1,after=0,exact=False):
        for r in range(after+1,ws.max_row+1):
            v=ws.cell(r,col).value
            if v is None: continue
            sv=str(v).strip()
            if (sv==lbl if exact else sv.startswith(lbl)): return r
        return 0
    L=nn(find('Ending Loan Base')); I=nn(find('Ending Inventory Base'))
    loan=money(L[2]) if len(L)>2 else 0.0
    inv=money(I[2]) if len(I)>2 else 0.0
    sub=nn(find('In-Store Subtotal'))
    isInt=money(sub[4]) if len(sub)>4 else 0
    isFee=money(sub[5]) if len(sub)>5 else 0
    isMisc=money(sub[6]) if len(sub)>6 else 0
    sa=find('Sales Activity',exact=True)
    tx=nn(find('Taxable Sales',after=sa,exact=True)); ntx=nn(find('Nontaxable Sales',after=sa,exact=True))
    taxT=money(tx[-1]) if tx else 0
    ntxT=money(ntx[-1]) if ntx else 0
    rev=nn(find('Sales Revenue (Profit)')); profit=money(rev[-1]) if rev else 0
    ref=nn(find('Refined',col=2)); scrap=abs(money(ref[-1])) if ref else 0
    stot=nn(find('Sales Total',after=sa,exact=True)); total_sales=money(stot[-1]) if stot else 0
    mob=nn(find('Totals from'))
    mInt=money(mob[3]) if len(mob)>3 else 0
    mFee=money(mob[4]) if len(mob)>4 else 0
    mMisc=money(mob[5]) if len(mob)>5 else 0
    conv=nn(find('MobilePawn Convenience Fees')); convV=money(conv[-1]) if conv else 0
    rep=''
    for r in range(1,min(ws.max_row,25)+1):
        for c in range(1,MC+1):
            v=ws.cell(r,c).value
            if v is not None:
                sv=str(v)
                if '/' in sv and ' - ' in sv:
                    rep=sv.strip(); break
        if rep: break
    # KPI-aligned 2026-07-02: Bravo Company Performance Net Revenue = in-store service charges (Interest+Fees+Misc) + Sales Revenue (Profit).
    # Verified to the penny vs KPI report for CUL (66,649.27) and HAR (61,666.31), June 2026. Mobile int/fees/misc + conv fees are NOT in KPI Net Revenue.
    retail=taxT+ntxT; instore_sc=isInt+isFee+isMisc; psc=instore_sc; net=instore_sc+profit
    return dict(inventory_balance=round(inv,2),loan_balance=round(loan,2),total_sales=round(total_sales,2),scrap_cost=round(scrap,2),psc=round(psc,2),net_revenue=round(net,2),reporting_dates=rep)
def parse_folder(folder):
    out={}
    for p in sorted(glob.glob(os.path.join(folder,'*.xlsx'))):
        if os.path.getsize(p)<2048: continue
        out[os.path.basename(p)]=parse_store_xlsx(p)
    return out
if __name__=='__main__':
    if len(sys.argv)<2:
        sys.stderr.write('usage: parse_eom.py <xlsx-or-folder>'+chr(10)); sys.exit(2)
    t=sys.argv[1]
    if os.path.isdir(t): print(json.dumps(parse_folder(t),indent=2))
    else: print(json.dumps(parse_store_xlsx(t),indent=2))
